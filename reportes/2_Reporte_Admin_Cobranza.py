"""
Módulo para el procesamiento del 2_Reporte_Admin_Cobranza

Este módulo procesa archivos de administración de cobranza y asistencia para generar
reportes con clasificación automatizada de carteras y formato Excel profesional.
"""
import pandas as pd
import io
import re
from typing import Optional, Union
from flask import request, send_file
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from utils.file_utils import allowed_file

# ==============================================================================
# CONSTANTES Y CONFIGURACIÓN
# ==============================================================================
MESES_ESPANOL = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

ROBOTS_FILTROS = [
    'M1-1rboot-94AI', 'M1-2rboot-94AI', 'M1-2rboot-RISKAI', 
    'M1-2rboot', 'rboot-94AI', 'rboot'
]

ASESORES_FILTROS = [
    'liwenzhen', 'prueba', 'Karol Sanchez', 'Daniela Arias', 
    'Brayan Murcia', 'M1-2rboot-RISKAI', 'Yesid Espitia', 'William Cabiativa'
]

POSIBLES_NOMBRES_CONEXION = [
    'Primera conexión', 'Primera Conexión', 'Primera conexion', 'Primera Conexion',
    'PRIMERA CONEXIÓN', 'PRIMERA CONEXION', 'Primer Login', 'Primer login',
    'First Login', 'Login', 'Hora Ingreso', 'Hora de Ingreso'
]

RANGOS_FRS_M11A = [
    'RM1-1A', 'R M1-1A', 'RM11A', 'R M11A', 
    'RM1-1', 'R M1-1', 'RM11', 'R M11'
]

# ==============================================================================
# FUNCIONES DE UTILIDAD
# ==============================================================================

def formatear_nombre(nombre: str) -> str:
    """
    Formatea nombres poniendo la primera letra de cada palabra en mayúscula
    y el resto en minúscula (formato título)
    
    Args:
        nombre (str): Nombre a formatear
        
    Returns:
        str: Nombre formateado
        
    Example:
        formatear_nombre("JUAN CARLOS pérez") -> "Juan Carlos Pérez"
    """
    if not nombre or pd.isna(nombre):
        return ''
    
    return str(nombre).title()

def _validar_frs_con_rango(aplicativo: str, rango_upper: str, cartera_base: str) -> str:
    """
    Valida cartera FRS con rango específico
    
    Args:
        aplicativo: Nombre del aplicativo
        rango_upper: Rango en mayúsculas para validación
        cartera_base: Cartera base si no es FRS
        
    Returns:
        str: Cartera final después de validación
    """
    if aplicativo == 'FRS':
        return 'M1-1A-FRS' if any(rango in rango_upper for rango in RANGOS_FRS_M11A) else 'M0-FRS'
    elif aplicativo:
        # Para M0 con aplicativo, siempre usar M0-Aplicativo (no M0-PP-BT sino M0-BT)
        if cartera_base.startswith('M0'):
            return f'M0-{aplicativo}'
        # Para otros casos (M1-1A, etc.), mantener el prefijo completo
        return f'{cartera_base}-{aplicativo}'
    else:
        return cartera_base

def generar_nombre_archivo_admin(report: pd.DataFrame) -> str:
    """
    Genera nombre dinámico para el archivo Excel del reporte admin
    
    Args:
        report (pd.DataFrame): DataFrame con la columna day_key
        
    Returns:
        str: Nombre del archivo generado con formato de fecha español
        
    Example:
        generar_nombre_archivo_admin(df) -> "Admin-Cobranza_01_a_31_ene_2025"
    """
    fechas = pd.to_datetime(report['day_key'].unique())
    fecha_min = fechas.min()
    fecha_max = fechas.max()
    
    if fecha_min.month == fecha_max.month and fecha_min.year == fecha_max.year:
        # Mismo mes y año
        mes_nombre = MESES_ESPANOL[fecha_min.month]
        return f"Reporte Admin ({fecha_min.day}-{fecha_max.day} {mes_nombre} {fecha_min.year}).xlsx"
    else:
        # Diferentes meses o años
        mes_min = MESES_ESPANOL[fecha_min.month]
        mes_max = MESES_ESPANOL[fecha_max.month]
        if fecha_min.year == fecha_max.year:
            return f"Reporte Admin ({fecha_min.day} {mes_min} - {fecha_max.day} {mes_max} {fecha_min.year}).xlsx"
        else:
            return f"Reporte Admin ({fecha_min.day} {mes_min} {fecha_min.year} - {fecha_max.day} {mes_max} {fecha_max.year}).xlsx"


# ==============================================================================
# FUNCIONES DE PROCESAMIENTO DE DATOS
# ==============================================================================
def detectar_aplicativo(gerencia_text: str) -> Optional[str]:
    """
    Detecta el tipo de aplicativo en el texto de gerencia
    
    Args:
        gerencia_text (str): Texto de la gerencia
        
    Returns:
        Optional[str]: Tipo de aplicativo detectado ('FRS', 'PN', 'BT') o None
        
    Example:
        detectar_aplicativo("GERENCIA FRS EJEMPLO") -> "FRS"
        detectar_aplicativo("GERENCIA NORMAL") -> None
    """
    gerencia_upper = str(gerencia_text).upper()
    
    if 'FRS' in gerencia_upper:
        return 'FRS'
    elif 'PN' in gerencia_upper:
        return 'PN'
    elif 'BT' in gerencia_upper:
        return 'BT'
    
    return None


def extract_cartera_from_gerencia_and_rango(gerencia_text: str, rango_text: str) -> str:
    """
    Sistema simplificado de clasificación de carteras
    Solo usa GERENCIA y RANGO (elimina validaciones de CARTERA)
    
    Args:
        gerencia_text (str): Texto de la gerencia
        rango_text (str): Texto del rango
        
    Returns:
        str: Cartera clasificada según las reglas de negocio
        
    Example:
        extract_cartera_from_gerencia_and_rango("GERENCIA FRS", "91_120") -> "M1-1A-FRS"
        extract_cartera_from_gerencia_and_rango("GERENCIA PN", "1_30") -> "M0-PN"
    """
    # Validar y normalizar entradas
    gerencia_text = str(gerencia_text) if pd.notna(gerencia_text) else ""
    rango_text = str(rango_text) if pd.notna(rango_text) else ""
    
    if not gerencia_text.strip():
        return ""
    
    gerencia_upper = gerencia_text.upper().strip()
    rango_upper = rango_text.upper().strip()
    gerencia_lower = gerencia_text.lower().strip()
    
    # Detectar aplicativo
    aplicativo = detectar_aplicativo(gerencia_text)
    
    # ========== REGLAS DE CLASIFICACIÓN (PRIORIDAD DESCENDENTE) ==========
    
    # REGLA 1: M1-2 (Mayor prioridad - se queda tal cual)
    if ('M1-2' in gerencia_upper or 'M12' in gerencia_upper or 
        'm1-2' in gerencia_lower or 'm12' in gerencia_lower):
        return 'M1-2'
    
    # REGLA 2: M1-1B (se queda tal cual)
    if ('M1-1B' in gerencia_upper or 'M11B' in gerencia_upper or 
        'm1-1b' in gerencia_lower or 'm11b' in gerencia_lower):
        return 'M1-1B'
    
    # REGLA 3: M0-1 PP (se queda tal cual)
    if (('M0-1' in gerencia_upper and 'PP' in gerencia_upper) or 'M0-1PP' in gerencia_upper or
        ('m0-1' in gerencia_lower and 'pp' in gerencia_lower) or 'm0-1pp' in gerencia_lower):
        return 'M0-1 PP'
    
    # REGLA 4: Caso especial M1-1A con BEATRIZ Y NANCY
    if ('M1-1' in gerencia_upper and 'A' in gerencia_upper and 
        'BEATRIZ' in gerencia_upper and 'NANCY' in gerencia_upper):
        return 'M1-1A-PN'
    
    # REGLA 5: M1-1A (con validación especial para FRS)
    if ('M1-1A' in gerencia_upper or 'M11A' in gerencia_upper or 
        'm1-1a' in gerencia_lower or 'm11a' in gerencia_lower):
        return _validar_frs_con_rango(aplicativo, rango_upper, 'M1-1A')
    
    # REGLA 6: M0 PP (con validación especial para FRS)
    if (('M0' in gerencia_upper and 'PP' in gerencia_upper) or 'M0PP' in gerencia_upper or
        ('m0' in gerencia_text and 'pp' in gerencia_lower) or 'm0pp' in gerencia_lower):
        return _validar_frs_con_rango(aplicativo, rango_upper, 'M0-PP')
    
    # REGLA 7: M0 VP (con validación especial para FRS)
    if (('M0' in gerencia_upper and 'VP' in gerencia_upper) or 'M0VP' in gerencia_upper or
        ('m0' in gerencia_text and 'vp' in gerencia_lower) or 'm0vp' in gerencia_lower):
        return _validar_frs_con_rango(aplicativo, rango_upper, 'M0-VP')
    
    # Si no coincide con ninguna regla, devolver el texto original
    return gerencia_text.strip()

# ==============================================================================
# FUNCIONES DE DETECCIÓN Y CLASIFICACIÓN
# ==============================================================================
def filtrar_por_estado_asistencia(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra el archivo de asistencia eliminando registros con estado 'ausente' o similar
    
    Args:
        df (pd.DataFrame): DataFrame de asistencia
        
    Returns:
        pd.DataFrame: DataFrame filtrado sin registros ausentes
    """
    df_filtrado = df.copy()
    
    # Buscar la columna Estado (columna G, que es la posición 6 en base 0)
    estado_col = None
    
    # Primero intentar buscar por nombre de columna
    posibles_nombres_estado = ['Estado', 'ESTADO', 'estado', 'Status', 'STATUS']
    for nombre in posibles_nombres_estado:
        if nombre in df_filtrado.columns:
            estado_col = nombre
            break
    
    # Si no se encuentra por nombre, usar la posición G (columna 6)
    if estado_col is None:
        if len(df_filtrado.columns) > 6:
            estado_col = df_filtrado.columns[6]  # Columna G (posición 6)
        else:
            print(f"ADVERTENCIA: No se encontró columna Estado. Columnas disponibles: {list(df_filtrado.columns)}")
            return df_filtrado
    
    if estado_col in df_filtrado.columns:
        # Valores a filtrar (ausente y similares)
        estados_a_excluir = [
            'ausente', 'AUSENTE', 'Ausente',
            'ausencia', 'AUSENCIA', 'Ausencia',
            'falta', 'FALTA', 'Falta',
            'no asistio', 'NO ASISTIO', 'No Asistio',
            'no asistió', 'NO ASISTIÓ', 'No Asistió'
        ]
        
        # Filtrar registros antes del procesamiento
        filas_antes = len(df_filtrado)
        df_filtrado = df_filtrado[~df_filtrado[estado_col].astype(str).str.strip().isin(estados_a_excluir)]
        filas_despues = len(df_filtrado)
        
        print(f"INFO: Filtro por Estado aplicado. Registros eliminados: {filas_antes - filas_despues}")
        print(f"INFO: Registros restantes para procesamiento: {filas_despues}")
    
    return df_filtrado


def detectar_archivo_asistencia(df: pd.DataFrame) -> bool:
    """
    Detecta si un DataFrame corresponde al archivo de asistencia
    
    Args:
        df (pd.DataFrame): DataFrame a evaluar
        
    Returns:
        bool: True si es archivo de asistencia, False si no
        
    Example:
        detectar_archivo_asistencia(df_con_conexion) -> True
        detectar_archivo_asistencia(df_admin) -> False
    """
    return any(col in df.columns for col in POSIBLES_NOMBRES_CONEXION)


def buscar_columna_conexion(df: pd.DataFrame) -> Optional[str]:
    """
    Busca y retorna el nombre de la columna de primera conexión
    
    Args:
        df (pd.DataFrame): DataFrame donde buscar
        
    Returns:
        Optional[str]: Nombre de la columna encontrada o None
        
    Example:
        buscar_columna_conexion(df) -> "First Login"
        buscar_columna_conexion(df_sin_conexion) -> None
    """
    for nombre in POSIBLES_NOMBRES_CONEXION:
        if nombre in df.columns:
            return nombre
    return None


# ==============================================================================
# FUNCIONES DE FILTRADO Y LIMPIEZA
# ==============================================================================


def aplicar_filtros_exclusion(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica filtros de exclusión para robots y asesores específicos
    
    Args:
        df (pd.DataFrame): DataFrame a filtrar
        
    Returns:
        pd.DataFrame: DataFrame filtrado excluyendo robots y asesores específicos
        
    Example:
        aplicar_filtros_exclusion(df_con_robots) -> df_limpio
    """
    # Aplicar filtros de robots en columna 6 (Gerencia)
    cond_robots_gerencia = df.get(6, pd.Series(dtype=str)).astype(str).str.contains(
        '|'.join(ROBOTS_FILTROS), case=False, na=False)
    
    # Aplicar filtros de robots en columna 4 (Nombre) - por si aparecen como nombres
    cond_robots_nombre = df.get(4, pd.Series(dtype=str)).astype(str).str.contains(
        '|'.join(ROBOTS_FILTROS), case=False, na=False)
    
    # Aplicar filtros de asesores (columna 4 = Nombre)
    cond_asesores = df.get(4, pd.Series(dtype=str)).astype(str).str.contains(
        '|'.join(ASESORES_FILTROS), case=False, na=False)
    
    # Combinar todas las condiciones de exclusión
    condicion_total = cond_robots_gerencia | cond_robots_nombre | cond_asesores
    
    # Retornar datos filtrados
    return df[~condicion_total]


# ==============================================================================
# FUNCIONES DE FORMATO Y CONVERSIÓN
# ==============================================================================
def convert_time_format(value: Union[str, pd.Timestamp, None]) -> str:
    """
    Convierte diferentes formatos de tiempo a formato de 12 horas con AM/PM y segundos
    
    Args:
        value (Union[str, pd.Timestamp, None]): Valor de tiempo a convertir
        
    Returns:
        str: Tiempo formateado como H:MM:SS AM/PM o cadena vacía si no válido
        
    Example:
        convert_time_format("14:30:45") -> "2:30:45 PM"
        convert_time_format("08:15:30") -> "8:15:30 AM"
        convert_time_format("2025-09-06 07:28:24") -> "7:28:24 AM"
    """
    if pd.isna(value) or value == '' or value is None:
        return ''
    
    try:
        value_str = str(value).strip()
        
        # Si ya tiene AM/PM, devolverlo como está
        if 'AM' in value_str.upper() or 'PM' in value_str.upper():
            return value_str
        
        # Si es timestamp completo como "2025-09-06 07:28:24"
        if ' ' in value_str and ':' in value_str:
            time_part = value_str.split(' ')[1]  # "07:28:24"
            parts = time_part.split(':')
            if len(parts) >= 3:
                # Formato completo con segundos
                hour = int(parts[0])
                minute = parts[1].zfill(2)
                second = parts[2].zfill(2)
                period = 'AM' if hour < 12 else 'PM'
                hour_12 = hour if hour <= 12 else hour - 12
                hour_12 = 12 if hour_12 == 0 else hour_12
                return f"{hour_12}:{minute}:{second} {period}"
            elif len(parts) == 2:
                # Sin segundos, agregar :00
                hour = int(parts[0])
                minute = parts[1].zfill(2)
                period = 'AM' if hour < 12 else 'PM'
                hour_12 = hour if hour <= 12 else hour - 12
                hour_12 = 12 if hour_12 == 0 else hour_12
                return f"{hour_12}:{minute}:00 {period}"
        
        # Si es solo hora como "08:34:58" o "8:34:58"
        elif ':' in value_str:
            parts = value_str.split(':')
            if len(parts) >= 3:
                # Formato completo con segundos
                hour = int(parts[0])
                minute = parts[1].zfill(2)
                second = parts[2].zfill(2)
                period = 'AM' if hour < 12 else 'PM'
                hour_12 = hour if hour <= 12 else hour - 12
                hour_12 = 12 if hour_12 == 0 else hour_12
                return f"{hour_12}:{minute}:{second} {period}"
            elif len(parts) == 2:
                # Sin segundos, agregar :00
                hour = int(parts[0])
                minute = parts[1].zfill(2)
                period = 'AM' if hour < 12 else 'PM'
                hour_12 = hour if hour <= 12 else hour - 12
                hour_12 = 12 if hour_12 == 0 else hour_12
                return f"{hour_12}:{minute}:00 {period}"
        
        # Si no hay formato reconocible pero no está vacío, devolver tal como está
        if value_str and value_str != 'nan':
            return value_str
        
        return ''
        
    except Exception:
        return str(value) if str(value) != 'nan' else ''


# ==============================================================================
# FUNCIONES DE ASIGNACIÓN DE GERENCIA
# ==============================================================================
def get_gerente(gerencia_str: str, team_leader_str: str = "") -> str:
    """
    Determina el gerente basado en gerencia y team leader
    
    Args:
        gerencia_str (str): Texto de gerencia
        team_leader_str (str): Texto de team leader
        
    Returns:
        str: Nombre del gerente asignado
        
    Example:
        get_gerente("M1-1A FRS", "Camilo") -> "Camilo Arciniegas"
        get_gerente("M0 VP", "") -> "Clara Navarro"
    """
    gerencia = str(gerencia_str).lower()
    team_leader = str(team_leader_str).lower()
    
    # Revisar primero por Team Leader específicos
    if 'andres acevedo' in team_leader or 'andres' in team_leader:
        return 'Yesid Espitia'
    if 'lizethe rodriguez' in team_leader or 'lizethe' in team_leader:
        return 'Daniela Arias'
    # Específico: Nancy Rodriguez va a William
    if 'nancy rodriguez' in team_leader:
        return 'William Cabiativa'
    # Específico: Nancy Cruz va a Daniela (después de verificar Nancy Rodriguez)
    if 'nancy cruz' in team_leader or ('nancy' in team_leader and 'rodriguez' not in team_leader):
        return 'Daniela Arias'
    
    # Revisar por contenido en gerencia
    if 'william' in gerencia:
        return 'William Cabiativa'
    if 'daniela' in gerencia:
        return 'Daniela Arias'
    if 'gerencia luis' in gerencia:
        return 'Luis Alzate'
    if 'yesid' in gerencia:
        return 'Yesid Espitia'
    
    return ''


# ==============================================================================
# FUNCIONES DE GENERACIÓN DE EXCEL
# ==============================================================================
def generar_archivo_excel(merged_df: pd.DataFrame):
    """
    Genera el archivo Excel con formato profesional
    
    Args:
        merged_df (pd.DataFrame): DataFrame con los datos procesados
        
    Returns:
        Response: Archivo Excel para descarga
        
    Example:
        generar_archivo_excel(df) -> Flask Response object
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for day, data in merged_df.groupby('day_key'):
            # Crear DataFrame del reporte
            report_sheet_df = pd.DataFrame({
                "ID": data.get('ID_str'),
                "Nombre": data.get(4).apply(formatear_nombre) if 4 in data else '',
                "Logueo": data.get('Logueo', ''),
                "CARTERA": data.get('cartera'),
                "ASIGNACION": data.get(13),
                "TOCADAS 11 AM": data.get(14),
                "ASIGNADO": data.get(17),
                "RECUPERADO": data.get(18),
                "PAGOS": data.get(15),
                "% RECUPERADO": data.get('% RECUPERADO'),
                "% CUENTAS": data.get('% CUENTAS'),
                "TOQUES": data.get(21),
                "ULTIMO TOQUE": data.get('ULTIMO TOQUE', ''),
                "Gerente": data.get('Gerente').apply(formatear_nombre) if 'Gerente' in data else '',
                "Team Leader": data.get('Team Leader').apply(formatear_nombre) if 'Team Leader' in data else '',
                "Ubicación": data.get('Ubicación')
            })
            
            # Eliminar duplicados manteniendo el de mayor RECUPERADO
            if not report_sheet_df.empty:
                report_sheet_df['RECUPERADO'] = pd.to_numeric(
                    report_sheet_df['RECUPERADO'], errors='coerce').fillna(0)
                
                report_sheet_df = (report_sheet_df
                                 .sort_values('RECUPERADO', ascending=False)
                                 .drop_duplicates(subset=['ID'], keep='first')
                                 .sort_values(['ASIGNACION', 'ASIGNADO', 'RECUPERADO'], 
                                            ascending=[False, False, False]))
            
            # Generar nombre de hoja (válido para Excel)
            dt = pd.to_datetime(day)
            sheet_name = f"{dt.day:02d}-{dt.month:02d}-{dt.year}"
            
            # Escribir datos
            report_sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Aplicar formato
            aplicar_formato_excel(writer.sheets[sheet_name], report_sheet_df)
    
    # Generar nombre dinámico del archivo
    nombre_archivo = generar_nombre_archivo_admin(merged_df)
    
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )


def aplicar_formato_excel(worksheet, report_sheet_df: pd.DataFrame) -> None:
    """
    Aplica formato profesional al worksheet de Excel
    
    Args:
        worksheet: Worksheet de openpyxl
        report_sheet_df (pd.DataFrame): DataFrame con los datos
        
    Example:
        aplicar_formato_excel(ws, df)
    """
    num_rows, num_cols = report_sheet_df.shape
    
    if num_rows == 0 or num_cols == 0:
        return
    
    try:
        # Crear nombre de tabla válido (solo letras, números y underscore)
        clean_title = re.sub(r'[^a-zA-Z0-9_]', '_', worksheet.title)
        table_name = f"Tabla_{clean_title}"
        
        # Calcular rango de tabla correctamente para cualquier número de columnas
        end_col = get_column_letter(num_cols)
        table_range = f"A1:{end_col}{num_rows + 1}"
        
        # Crear tabla con estilo púrpura
        table = Table(displayName=table_name, ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium12",  # Tabla Medio 12 (púrpura)
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        worksheet.add_table(table)
        
    except Exception as e:
        # Continuar sin tabla si hay problemas
        pass
    
    try:
        # Ajustar ancho de columnas
        for col_obj in worksheet.columns:
            max_length = 0
            column_letter = col_obj[0].column_letter
            column_name = col_obj[0].value
            
            for cell in col_obj:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            
            # Dar ancho extra a la columna ASIGNADO
            if column_name == 'ASIGNADO':
                adjusted_width = max(max_length + 5, 18)
            else:
                adjusted_width = (max_length + 2) if max_length < 50 else 50
            
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Centrar todo el contenido
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
        
        # Aplicar formatos numéricos
        aplicar_formatos_numericos(worksheet, report_sheet_df)
        
    except Exception as e:
        # Continuar sin formato si hay problemas
        pass


def aplicar_formatos_numericos(worksheet, report_sheet_df: pd.DataFrame) -> None:
    """
    Aplica formatos numéricos específicos a las columnas
    
    Args:
        worksheet: Worksheet de openpyxl
        report_sheet_df (pd.DataFrame): DataFrame con los datos
        
    Example:
        aplicar_formatos_numericos(ws, df)
    """
    # Mapear nombres de columnas a letras usando get_column_letter
    column_letters = {}
    for col_num, col_name in enumerate(report_sheet_df.columns, 1):
        column_letters[col_name] = get_column_letter(col_num)
    
    # Formato de moneda COP para ASIGNADO y RECUPERADO
    if 'ASIGNADO' in column_letters and 'RECUPERADO' in column_letters:
        peso_format = '"$"#,##0'
        for row_num in range(2, worksheet.max_row + 1):
            asignado_cell = f"{column_letters['ASIGNADO']}{row_num}"
            worksheet[asignado_cell].number_format = peso_format
            
            recuperado_cell = f"{column_letters['RECUPERADO']}{row_num}"
            worksheet[recuperado_cell].number_format = peso_format
    
    # Formato de porcentaje
    if '% RECUPERADO' in column_letters and '% CUENTAS' in column_letters:
        percent_format = '0.00%'
        for row_num in range(2, worksheet.max_row + 1):
            percent_recup_cell = f"{column_letters['% RECUPERADO']}{row_num}"
            worksheet[percent_recup_cell].number_format = percent_format
            
            percent_cuentas_cell = f"{column_letters['% CUENTAS']}{row_num}"
            worksheet[percent_cuentas_cell].number_format = percent_format


# ==============================================================================
# FUNCIÓN PRINCIPAL DE PROCESAMIENTO
# ==============================================================================
def procesar_admin_cobranza() -> Union[str, tuple]:
    """
    Procesa los archivos de Admin Cobranza y genera el 2_Reporte_Admin_Cobranza
    
    Returns:
        Union[str, tuple]: Archivo Excel generado o mensaje de error con código HTTP
        
    Example:
        procesar_admin_cobranza() -> Excel file response or ("Error message", 400)
    """
    # Validación de archivos subidos
    admin_file = request.files.get('adminFile')
    asistencia_file = request.files.get('asistenciaFile')

    if not admin_file or not asistencia_file:
        return "Error: Debes subir ambos archivos.", 400
    
    if not (allowed_file(admin_file.filename, {'xlsx', 'csv'}) and 
            allowed_file(asistencia_file.filename, {'xlsx', 'csv'})):
        return "Error: Formato de archivo no permitido.", 400

    try:
        # Leer archivos y detectar automáticamente cuál es cuál
        try:
            file1_df = pd.read_excel(admin_file, header=0)
        except Exception:
            file1_df = pd.read_excel(admin_file, header=None)
            
        try:
            file2_df = pd.read_excel(asistencia_file, header=0)
        except Exception:
            file2_df = pd.read_excel(asistencia_file, header=None)
        
        # Detectar automáticamente cuál archivo es admin y cuál es asistencia
        file1_es_asistencia = detectar_archivo_asistencia(file1_df)
        file2_es_asistencia = detectar_archivo_asistencia(file2_df)
        
        if file1_es_asistencia and not file2_es_asistencia:
            # Intercambiar archivos
            asistencia_df = file1_df
            admin_df = pd.read_excel(asistencia_file, header=None, dtype={3: str})
        elif file2_es_asistencia and not file1_es_asistencia:
            # Orden correcto
            admin_df = pd.read_excel(admin_file, header=None, dtype={3: str})
            asistencia_df = file2_df
        elif file1_es_asistencia and file2_es_asistencia:
            return "Error: Ambos archivos parecen ser de asistencia. Verifica que uno sea el archivo admin de cobranza.", 400
        else:
            return "Error: No se pudo identificar el archivo de asistencia. Verifica que contenga columnas como 'Primera conexión'.", 400
        
        # Buscar y estandarizar columna de primera conexión
        primera_conexion_col = buscar_columna_conexion(asistencia_df)
        if not primera_conexion_col:
            return f"Error: No se encontró columna de primera conexión. Columnas disponibles: {list(asistencia_df.columns)}", 400
        
        # Renombrar la columna para estandarizar
        if primera_conexion_col != 'Primera conexión':
            asistencia_df = asistencia_df.rename(columns={primera_conexion_col: 'Primera conexión'})
        
        # Validar columnas requeridas
        if not all(col in asistencia_df.columns for col in ['Fecha', 'ID']):
            return "Error: El archivo de asistencia debe tener las columnas 'Fecha' e 'ID'.", 400

        # FILTRAR POR ESTADO: Eliminar registros ausentes antes del cruce
        asistencia_df = filtrar_por_estado_asistencia(asistencia_df)

        # Procesar datos de asistencia
        asistencia_df['Fecha'] = pd.to_datetime(asistencia_df['Fecha'], errors='coerce')
        asistencia_df['day_key'] = asistencia_df['Fecha'].dt.strftime('%Y-%m-%d')
        asistencia_df['ID'] = asistencia_df['ID'].astype(str).str.strip()
        
        # Aplicar filtros de exclusión
        admin_df = aplicar_filtros_exclusion(admin_df)
        
        # Crear columnas necesarias para el procesamiento
        admin_df['ID_str'] = admin_df.get(3, pd.Series(dtype=str)).astype(str).str.strip()
        admin_df['day_key'] = pd.to_datetime(admin_df.get(2, pd.Series(dtype=str)), errors='coerce').dt.strftime('%Y-%m-%d')
        admin_df['ASIGNACION'] = pd.to_numeric(admin_df.get(12, pd.Series(dtype=object)), errors='coerce')
        
        # Aplicar clasificación de carteras
        admin_df['cartera'] = admin_df.apply(
            lambda row: extract_cartera_from_gerencia_and_rango(
                row.get(6, ''), row.get(10, '')
            ), axis=1
        )
        
        # Aplicar ffill y bfill solo si es necesario
        admin_df['cartera'] = admin_df['cartera'].ffill().bfill()
        
        # Validar columna de fecha
        if 2 not in admin_df.columns:
            return "Error: El archivo de Admin no tiene la columna de fecha (columna C).", 400
        
        admin_df[2] = pd.to_datetime(admin_df[2], format='%Y%m%d', errors='coerce')
        admin_df['day_key'] = admin_df[2].dt.strftime('%Y-%m-%d')

        # Hacer merge de datos por ID y fecha
        merged_df = pd.merge(
            admin_df, 
            asistencia_df[['ID', 'Primera conexión', 'day_key']],
            left_on=['ID_str', 'day_key'], 
            right_on=['ID', 'day_key'], 
            how='left'
        )
        
        # Verificar merge exitoso
        if merged_df['Primera conexión'].notna().sum() == 0:
            pass  # No se encontraron coincidencias en el merge

        # Procesar columnas de tiempo
        if 'Primera conexión' in merged_df.columns:
            merged_df['Logueo'] = merged_df['Primera conexión'].apply(convert_time_format)
        else:
            merged_df['Logueo'] = ''  # Columna no encontrada
            
        # Procesar ULTIMO TOQUE
        if 'El último procesamiento' in merged_df.columns:
            merged_df['ULTIMO TOQUE'] = merged_df['El último procesamiento'].apply(convert_time_format)
        else:
            merged_df['ULTIMO TOQUE'] = merged_df.get(28, pd.Series(dtype=object)).apply(convert_time_format)
        
        # Procesar información de gerencia
        gerencia_col = merged_df.get(6, pd.Series(dtype=str))
        
        # Limpiar y extraer team leader - remover prefijos comunes
        team_leader_col = gerencia_col.str.replace(
            r'.*(Team Leader|Team|Back)\s+', '', regex=True, flags=re.IGNORECASE
        ).str.strip()
        
        # Si queda vacío o igual al original, intentar limpiar solo "Team" al inicio
        mask_empty = (team_leader_col == '') | (team_leader_col == gerencia_col)
        team_leader_col[mask_empty] = gerencia_col[mask_empty].str.replace(
            r'Team\s+', '', regex=True, flags=re.IGNORECASE
        ).str.strip()
        
        # Aplicar función de gerente
        merged_df['Gerente'] = merged_df.apply(
            lambda row: get_gerente(
                row.get(6, ''), 
                team_leader_col.iloc[row.name] if row.name < len(team_leader_col) else ""
            ), axis=1
        )
        merged_df['Team Leader'] = team_leader_col
        
        # Aplicar correcciones específicas por cartera
        # REGLA ESPECIAL: M1-1A-BT debe tener Yesid como gerencia y Andres Acevedo como team leader
        mask_m11a_bt = merged_df['cartera'] == 'M1-1A-BT'
        if mask_m11a_bt.any():
            merged_df.loc[mask_m11a_bt, 'Gerente'] = 'Yesid Espitia'
            merged_df.loc[mask_m11a_bt, 'Team Leader'] = 'Andres Acevedo'
        
        # REGLA ESPECIAL: M0-BT debe tener Yesid como gerencia y Andres Acevedo como team leader
        mask_m0_bt = merged_df['cartera'] == 'M0-BT'
        if mask_m0_bt.any():
            merged_df.loc[mask_m0_bt, 'Gerente'] = 'Yesid Espitia'
            merged_df.loc[mask_m0_bt, 'Team Leader'] = 'Andres Acevedo'
        
        # REGLA ESPECIAL: M1-2 no tiene team leader solo para William
        mask_m12_william = (merged_df['cartera'] == 'M1-2') & (merged_df['Gerente'] == 'William Cabiativa')
        if mask_m12_william.any():
            merged_df.loc[mask_m12_william, 'Team Leader'] = ''
        
        merged_df['Ubicación'] = gerencia_col.apply(lambda x: 'Home' if 'home' in str(x).lower() else 'Sede')

        # Asegurar columnas numéricas
        columnas_numericas = [13, 14, 15, 17, 18, 21]
        for col in columnas_numericas:
            if col not in merged_df.columns:
                merged_df[col] = pd.NA
            merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce')
        
        # Calcular porcentajes
        merged_df['% RECUPERADO'] = (merged_df[18].fillna(0) / merged_df[17].replace(0, pd.NA)).fillna(0)
        merged_df['% CUENTAS'] = (merged_df[15].fillna(0) / merged_df[13].replace(0, pd.NA)).fillna(0)
        
        # Ordenar por ASIGNACION, ASIGNADO, RECUPERADO
        merged_df.sort_values(by=[13, 17, 18], ascending=[False, False, False], inplace=True)
        
        # Generar archivo Excel
        return generar_archivo_excel(merged_df)

    except Exception as e:
        return f"Ocurrió un error procesando los archivos: {e}", 500
