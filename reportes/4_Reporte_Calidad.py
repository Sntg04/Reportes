"""
MÃ³dulo para el procesamiento del 4_Reporte_Calidad
Sistema para generaciÃ³n de reportes de calidad con mÃ©tricas y anÃ¡lisis operativo.

Genera un archivo Excel completo con mÃºltiples hojas:
- Consolidado: Resumen de mÃ©tricas principales
- Gerente: Datos de gestiÃ³n y supervisiÃ³n  
- Team: InformaciÃ³n por equipos
- Operativo: Datos operativos detallados del Reporte 3
- Calidad: MÃ©tricas de calidad y seguimiento
- Ausentismo: Control de asistencia con datos biomÃ©tricos
- Asistencia Lideres: Seguimiento de liderazgo
- Planta: InformaciÃ³n de personal y configuraciÃ³n

CaracterÃ­sticas principales:
- ConversiÃ³n automÃ¡tica de formatos de tiempo (24h â†’ AM/PM)
- Procesamiento de datos biomÃ©tricos con validaciÃ³n
- GeneraciÃ³n automÃ¡tica de fÃ³rmulas Excel para mÃ©tricas calculadas
- ValidaciÃ³n de datos y manejo de errores integrado
- ConfiguraciÃ³n centralizada de constantes y umbrales

VersiÃ³n: 3.1 (Con mejoras de cÃ³digo y optimizaciÃ³n)
Ãšltima actualizaciÃ³n: Enero 2025
"""

import pandas as pd
import io
import os
import glob
import logging
from datetime import datetime
from flask import request, jsonify, send_file
from utils.file_utils import allowed_file

# Importaciones de openpyxl consolidadas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# Configurar logging
logger = logging.getLogger(__name__)

# ==========================================
# CONSTANTES Y CONFIGURACIONES
# ==========================================

# Configuraciones de Excel
DEFAULT_EXCEL_RANGE = 1000
PERCENTAGE_FORMAT = '0.00%'
INTEGER_FORMAT = '0'
DATE_FORMAT = '%d/%m/%Y'
TABLE_STYLE_DEFAULT = 'TableStyleMedium2'
TABLE_STYLE_BLUE = 'TableStyleMedium9'
TABLE_STYLE_GREEN = 'TableStyleMedium15'
TABLE_STYLE_YELLOW = 'TableStyleMedium21'

# Meses en espaÃ±ol para nombres de archivo
MESES_ESPANOL = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

# Nombres de hojas
SHEET_NAMES = {
    'CONSOLIDADO': 'Consolidado',
    'GERENTE': 'Gerente', 
    'TEAM': 'Team',
    'OPERATIVO': 'Operativo',
    'CALIDAD': 'Calidad',
    'AUSENTISMO': 'Ausentismo',
    'ASISTENCIA_LIDERES': 'Asistencia Lideres',
    'PLANTA': 'Planta'
}

# Columnas estÃ¡ndar para validaciÃ³n
COLUMNAS_REQUERIDAS_OPERATIVO = [
    'CODIGO', 'Cedula', 'Fecha', 'Team', 'Gerencia'
]

COLUMNAS_REQUERIDAS_CALIDAD = [
    'CODIGO', 'Nota Total', 'Total Monitoreos'
]

# ==========================================
# FUNCIONES DE UTILIDAD
# ==========================================

def validar_columnas_excel(df, columnas_requeridas, sheet_name="hoja"):
    """
    Valida que existan las columnas requeridas en un DataFrame
    
    Args:
        df: DataFrame a validar
        columnas_requeridas: Lista de nombres de columnas requeridas
        sheet_name: Nombre de la hoja para mensajes de error
    
    Returns:
        bool: True si todas las columnas existen, False si alguna falta
    """
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        print(f"ERROR: Columnas faltantes en {sheet_name}: {columnas_faltantes}")
        return False
    
    return True

def encontrar_columna_excel(worksheet, nombre_columna, silent=False):
    """
    Encuentra el Ã­ndice de una columna en una hoja Excel
    
    Args:
        worksheet: Hoja de Excel de openpyxl
        nombre_columna: Nombre de la columna a buscar
        silent: Si True, no imprime mensajes de error
    
    Returns:
        int o None: Ãndice de la columna (1-based) o None si no se encuentra
    """
    for col in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=col).value == nombre_columna:
            return col
    
    if not silent:
        print(f"ERROR: No se encontrÃ³ columna '{nombre_columna}'")
    
    return None

def procesar_fechas_columna(df, columna_fecha, formato_salida='%d/%m/%Y'):
    """
    Procesa y normaliza una columna de fechas con mÃºltiples formatos
    
    Args:
        df: DataFrame que contiene la columna
        columna_fecha: Nombre de la columna de fechas
        formato_salida: Formato de salida para las fechas
    
    Returns:
        tuple: (fechas_procesadas, count_validas)
    """
    try:
        # Intentar conversiÃ³n automÃ¡tica
        df[columna_fecha] = pd.to_datetime(df[columna_fecha], errors='coerce')
        
        # Intentar formatos especÃ­ficos para valores nulos
        mask_nulos = df[columna_fecha].isna()
        if mask_nulos.any():
            # Formato DD/MM/YYYY
            try:
                fechas_temp = pd.to_datetime(df.loc[mask_nulos, columna_fecha], 
                                           format='%d/%m/%Y', errors='coerce')
                df.loc[mask_nulos, columna_fecha] = fechas_temp
            except (ValueError, TypeError) as e:
                logger.debug(f"Error en conversiÃ³n de fecha DD/MM/YYYY: {e}")
                pass
                
            # Formato MM/DD/YYYY  
            mask_nulos = df[columna_fecha].isna()
            if mask_nulos.any():
                try:
                    fechas_temp = pd.to_datetime(df.loc[mask_nulos, columna_fecha], 
                                               format='%m/%d/%Y', errors='coerce')
                    df.loc[mask_nulos, columna_fecha] = fechas_temp
                except (ValueError, TypeError) as e:
                    logger.debug(f"Error en conversiÃ³n de fecha MM/DD/YYYY: {e}")
                    pass
        
        # Convertir a formato de salida
        df[columna_fecha] = df[columna_fecha].dt.strftime(formato_salida)
        count_validas = df[columna_fecha].notna().sum()
        
        return df[columna_fecha], count_validas
        
    except Exception as e:
        logger.warning(f"Error procesando fechas en columna {columna_fecha}: {e}")
        return df[columna_fecha], 0

# Constantes de configuraciÃ³n del reporte
class ConfigReporte:
    """Constantes de configuraciÃ³n centralizadas para el reporte de calidad"""
    
    # Filtros y umbrales
    FILTRO_ASIGNACION_MIN = 5
    VALOR_INDICADOR_CUMPLE = 0.15
    M0_TOQUES_MIN = 120
    OTRAS_CARTERAS_TOQUES_MIN = 160
    PORCENTAJE_ASIGNACION_M0 = 0.9
    ASIGNACION_MIN_GESTION = 45
    VALOR_INDICADOR_TOQUES = 0.2
    
    # Horarios de validaciÃ³n
    HORA_INGRESO_NORMAL = "08:00:00"
    HORA_INGRESO_PAGO = "07:30:00"
    
    # Nombres de hojas del reporte
    HOJAS_REPORTE = [
        "Consolidado", "Gerente", "Team", "Operativo", 
        "Calidad", "Ausentismo", "Asistencia Lideres", "Planta"
    ]
    
    # Columnas para el DataFrame de ausentismo
    COLUMNAS_AUSENTISMO = [
        "Codigo Aus", "Codigo", "Tipo Jornada", "Fecha", "Cedula", "ID",
        "Nombre", "Sede", "Ubicacion", "Logueo Admin", "Ingreso", 
        "Salida", "Horas laboradas", "Novedad Ingreso", "Validacion", "Drive"
    ]



def aplicar_formato_porcentaje(worksheet, columnas_porcentaje, num_filas):
    """
    Aplica formato de porcentaje a columnas especÃ­ficas
    
    Args:
        worksheet: Hoja de Excel de openpyxl
        columnas_porcentaje: Lista de letras de columnas (ej: ['M', 'N', 'P'])
        num_filas: NÃºmero de filas con datos (sin contar encabezados)
    """
    if num_filas > 0:
        for columna in columnas_porcentaje:
            for fila in range(2, num_filas + 2):  # +2 porque empezamos en fila 2
                worksheet[f'{columna}{fila}'].number_format = PERCENTAGE_FORMAT
        
        print(f"âœ… Formato de porcentaje aplicado a {num_filas} filas en columnas {', '.join(columnas_porcentaje)}")

def crear_formula_vlookup(columna_busqueda, hoja_destino, rango_columnas, indice_columna, valor_error=""):
    """
    Crea una fÃ³rmula VLOOKUP estÃ¡ndar con validaciÃ³n
    
    Args:
        columna_busqueda: Columna donde estÃ¡ el valor a buscar (ej: 'B{i}')
        hoja_destino: Nombre de la hoja destino
        rango_columnas: Rango de columnas (ej: 'A:G')
        indice_columna: Ãndice de la columna a devolver
        valor_error: Valor a devolver en caso de error
    
    Returns:
        str: FÃ³rmula VLOOKUP completa
    """
    return f'=IF({columna_busqueda}="",{valor_error},IFERROR(VLOOKUP({columna_busqueda},{hoja_destino}!{rango_columnas},{indice_columna},FALSE),{valor_error}))'

def crear_formula_busqueda_multiple(fecha_col, nombre_col, equipo_col, operacion_col=None, operacion="COUNT"):
    """
    Crea la fÃ³rmula de bÃºsqueda mÃºltiple de palabras para coincidencias de nombres
    
    Args:
        fecha_col: Columna de fecha (ej: 'B{fila}')
        nombre_col: Columna de nombre (ej: 'C{fila}')
        equipo_col: Columna de equipo en Operativo (ej: 'Z' para Team, 'Y' para Gerencia)
        operacion_col: Columna de operaciÃ³n especÃ­fica (ej: 'AI' para infracciones)
        operacion: Tipo de operaciÃ³n ('COUNT', 'SUM', etc.)
    
    Returns:
        str: Parte de la fÃ³rmula para bÃºsqueda mÃºltiple
    """
    condicion_busqueda = f"""((ISNUMBER(SEARCH(LEFT({nombre_col},FIND(" ",{nombre_col}&" ")-1),Operativo!{equipo_col}$2:{equipo_col}$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE({nombre_col}," ",REPT(" ",50)),51,50)),Operativo!{equipo_col}$2:{equipo_col}$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE({nombre_col}," ",REPT(" ",50)),101,50)),Operativo!{equipo_col}$2:{equipo_col}$1000))))>=2"""
    
    condicion_fecha = f"Operativo!C$2:C$1000={fecha_col}"
    
    if operacion == "COUNT":
        return f"SUMPRODUCT(({condicion_fecha})*({condicion_busqueda}))"
    elif operacion == "SUM" and operacion_col:
        return f"SUMPRODUCT(({condicion_fecha})*({condicion_busqueda})*Operativo!{operacion_col}$2:{operacion_col}$1000)"
    elif operacion == "COUNT_MONITORED":
        return f"SUMPRODUCT(({condicion_fecha})*({condicion_busqueda})*(ISNUMBER(MATCH(Operativo!A$2:A$1000,Calidad!A:A,0))))"
    
    return f"SUMPRODUCT(({condicion_fecha})*({condicion_busqueda}))"

def generar_nombre_archivo_calidad(df_reporte3=None):
    """
    Genera nombre dinÃ¡mico para el archivo Excel del Reporte Calidad basado en las fechas de los datos
    
    Args:
        df_reporte3: DataFrame del Reporte 3 con columna 'Fecha' 
        
    Returns:
        str: Nombre del archivo con formato legible
        
    Examples:
        - Un dÃ­a: "Reporte Calidad (9 Agosto 2025).xlsx"
        - Mismo mes: "Reporte Calidad (3-8 Septiembre 2025).xlsx" 
        - Diferente mes: "Reporte Calidad (30 Agosto - 5 Septiembre 2025).xlsx"
    """
    if df_reporte3 is None or df_reporte3.empty or 'Fecha' not in df_reporte3.columns:
        # Usar fecha actual si no hay datos
        fecha_actual = datetime.now()
        mes_actual = MESES_ESPANOL[fecha_actual.month]
        return f"Reporte Calidad ({fecha_actual.day} {mes_actual} {fecha_actual.year}).xlsx"
    
    try:
        # Convertir fechas y obtener Ãºnicas ordenadas
        fechas_unicas = pd.to_datetime(df_reporte3['Fecha'], errors='coerce').dropna().unique()
        fechas_unicas = sorted(fechas_unicas)
        
        if len(fechas_unicas) == 0:
            # Si no hay fechas vÃ¡lidas, usar fecha actual
            fecha_actual = datetime.now()
            mes_actual = MESES_ESPANOL[fecha_actual.month]
            return f"Reporte Calidad ({fecha_actual.day} {mes_actual} {fecha_actual.year}).xlsx"
        
        fecha_inicio = fechas_unicas[0]
        fecha_fin = fechas_unicas[-1]
        
        # Un solo dÃ­a
        if fecha_inicio.date() == fecha_fin.date():
            dia = fecha_inicio.day
            mes = MESES_ESPANOL[fecha_inicio.month]
            aÃ±o = fecha_inicio.year
            return f"Reporte Calidad ({dia} {mes} {aÃ±o}).xlsx"
        
        # Mismo mes y aÃ±o
        if fecha_inicio.month == fecha_fin.month and fecha_inicio.year == fecha_fin.year:
            mes = MESES_ESPANOL[fecha_inicio.month]
            aÃ±o = fecha_inicio.year
            return f"Reporte Calidad ({fecha_inicio.day}-{fecha_fin.day} {mes} {aÃ±o}).xlsx"
        
        # Mismo aÃ±o, diferentes meses
        if fecha_inicio.year == fecha_fin.year:
            mes_inicio = MESES_ESPANOL[fecha_inicio.month]
            mes_fin = MESES_ESPANOL[fecha_fin.month]
            aÃ±o = fecha_inicio.year
            return f"Reporte Calidad ({fecha_inicio.day} {mes_inicio} - {fecha_fin.day} {mes_fin} {aÃ±o}).xlsx"
        
        # Diferentes aÃ±os
        mes_inicio = MESES_ESPANOL[fecha_inicio.month]
        mes_fin = MESES_ESPANOL[fecha_fin.month]
        return f"Reporte Calidad ({fecha_inicio.day} {mes_inicio} {fecha_inicio.year} - {fecha_fin.day} {mes_fin} {fecha_fin.year}).xlsx"
        
    except Exception as e:
        logger.error(f"Error al generar nombre de archivo con fechas: {e}")
        # Fallback a fecha actual
        fecha_actual = datetime.now()
        mes_actual = MESES_ESPANOL[fecha_actual.month]
        return f"Reporte Calidad ({fecha_actual.day} {mes_actual} {fecha_actual.year}).xlsx"

def aplicar_formato_tabla(worksheet, dataframe, table_name):
    """
    Aplica formato de tabla Excel a una hoja de cÃ¡lculo
    
    Args:
        worksheet: Hoja de Excel donde aplicar el formato
        dataframe: DataFrame con los datos
        table_name: Nombre de la tabla
    """
    try:
        # Imports consolidados en el top del archivo
        
        # Definir el rango de la tabla
        if len(dataframe) == 0:
            # Si no hay datos, crear tabla solo con encabezados
            end_column = get_column_letter(len(dataframe.columns))
            table_range = f"A1:{end_column}1"
        else:
            end_column = get_column_letter(len(dataframe.columns))
            end_row = len(dataframe) + 1  # +1 para incluir encabezados
            table_range = f"A1:{end_column}{end_row}"
        
        # Crear tabla con estilo
        table = Table(displayName=table_name, ref=table_range)
        
        # Aplicar estilo de tabla
        style = TableStyleInfo(
            name=TABLE_STYLE_DEFAULT,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Agregar tabla a la hoja
        worksheet.add_table(table)
        
        print(f"OK: Formato de tabla aplicado: {table_name} ({table_range})")
        
    except Exception as e:
        print(f"WARN:  No se pudo aplicar formato de tabla a {table_name}: {str(e)}")

def procesar_archivo_biometricos():
    """
    Procesa archivo biomÃ©trico con estructura especÃ­fica:
    FECHA, CODIGO, CEDULA, NOMBRE, HORA, CARGO, AREA, SEDE, FECHA EN CHINO, TIPO DE DIA
    
    Proceso:
    1. Genera cÃ³digo: CEDULA + DIA + MES de la fecha
    2. Agrupa por cÃ³digo y calcula hora mÃ­nima (ingreso) y mÃ¡xima (salida)
    
    Returns:
        dict: Diccionario con cÃ³digos y horas procesadas para integraciÃ³n
    """
    try:
        print("=== PROCESANDO ARCHIVO BIOMÃ‰TRICOS ===")
        
        # Obtener archivo biomÃ©trico del formulario
        biometricos_file = request.files.get('archivoBiometricos')
        if not biometricos_file:
            print("AVISO: No se encontrÃ³ archivo biomÃ©trico (opcional)")
            return None
        
        print(f"âœ… Archivo biomÃ©trico recibido: {biometricos_file.filename}")
        
        # Validar formato
        if not allowed_file(biometricos_file.filename, {'xlsx', 'xls'}):
            print("ERROR: Formato de archivo no permitido para biomÃ©tricos")
            return None
        
        # Leer archivo Excel - Primero detectar si tiene headers
        print("ðŸ” Analizando estructura del archivo...")
        
        # Intentar leer con headers
        df_test = pd.read_excel(biometricos_file, nrows=5)
        print(f"ðŸ“Š Primeras columnas detectadas: {list(df_test.columns)}")
        
        # Si las columnas son 'Unnamed:', probablemente no hay headers
        if any('Unnamed:' in str(col) for col in df_test.columns):
            print("âš ï¸ Archivo sin headers detectado, aplicando headers estÃ¡ndar...")
            biometricos_file.seek(0)  # Resetear posiciÃ³n
            df_biometricos = pd.read_excel(biometricos_file, header=None)
            
            # Aplicar nombres estÃ¡ndar de columnas segÃºn tu especificaciÃ³n
            if len(df_biometricos.columns) >= 3:
                nuevas_columnas = ['FECHA', 'CODIGO', 'CEDULA', 'NOMBRE', 'HORA', 'CARGO', 'AREA', 'SEDE', 'FECHA EN CHINO', 'TIPO DE DIA']
                # Usar solo las columnas que existen
                num_cols = min(len(df_biometricos.columns), len(nuevas_columnas))
                df_biometricos.columns = nuevas_columnas[:num_cols]
                print(f"âœ… Headers aplicados: {list(df_biometricos.columns)}")
            else:
                print(f"âŒ ERROR: Archivo tiene muy pocas columnas: {len(df_biometricos.columns)}")
                return None
        else:
            # Archivo ya tiene headers vÃ¡lidos
            biometricos_file.seek(0)
            df_biometricos = pd.read_excel(biometricos_file)
        
        print(f"âœ… Archivo procesado: {len(df_biometricos)} registros")
        print(f"âœ… Columnas finales: {list(df_biometricos.columns)}")
        
        # Validar que tenga las columnas requeridas
        columnas_requeridas = ['FECHA', 'CEDULA', 'HORA']
        columnas_faltantes = [col for col in columnas_requeridas if col not in df_biometricos.columns]
        
        if columnas_faltantes:
            print(f"âŒ ERROR: Columnas faltantes: {columnas_faltantes}")
            print(f"ðŸ“‹ Columnas disponibles: {list(df_biometricos.columns)}")
            print("ðŸ’¡ AsegÃºrese de que el archivo tenga las columnas: FECHA, CEDULA, HORA")
            return None
        
        # Limpiar datos nulos
        registros_iniciales = len(df_biometricos)
        df_biometricos = df_biometricos.dropna(subset=['FECHA', 'CEDULA', 'HORA']).copy()
        print(f"âœ… Datos limpiados: {len(df_biometricos)} registros vÃ¡lidos de {registros_iniciales}")
        
        if len(df_biometricos) == 0:
            print("âŒ ERROR: No hay registros vÃ¡lidos")
            return None
        
        # Convertir fecha a datetime
        df_biometricos['FECHA'] = pd.to_datetime(df_biometricos['FECHA'], errors='coerce')
        df_biometricos = df_biometricos.dropna(subset=['FECHA']).copy()
        print(f"âœ… Fechas procesadas: {len(df_biometricos)} registros con fechas vÃ¡lidas")
        
        # Generar cÃ³digo: CEDULA + DIA + MES
        print("ðŸ”„ Generando cÃ³digos biomÃ©tricos...")
        df_biometricos['dia'] = df_biometricos['FECHA'].dt.day.astype(str).str.zfill(2)
        df_biometricos['mes'] = df_biometricos['FECHA'].dt.month.astype(str).str.zfill(2)
        df_biometricos['cedula_str'] = df_biometricos['CEDULA'].astype(str).str.strip()
        df_biometricos['codigo_biometrico'] = df_biometricos['cedula_str'] + df_biometricos['dia'] + df_biometricos['mes']
        
        print(f"âœ… CÃ³digos generados: {len(df_biometricos['codigo_biometrico'].unique())} cÃ³digos Ãºnicos")
        
        # Mostrar muestra de cÃ³digos generados
        print("ðŸ” Muestra de cÃ³digos generados:")
        for i, row in df_biometricos.head(3).iterrows():
            fecha_str = row['FECHA'].strftime('%d/%m/%Y')
            print(f"   CÃ©dula: {row['cedula_str']}, Fecha: {fecha_str} â†’ CÃ³digo: {row['codigo_biometrico']}")
        
        # Agrupar por cÃ³digo y calcular min/max horas
        print("ðŸ“Š Agrupando por cÃ³digo y calculando horas min/max...")
        
        # Convertir HORA a formato datetime para cÃ¡lculos precisos
        try:
            df_biometricos['HORA_DATETIME'] = pd.to_datetime(df_biometricos['HORA'], format='%H:%M:%S', errors='coerce')
            if df_biometricos['HORA_DATETIME'].isna().any():
                # Intentar otros formatos comunes
                mask_nulos = df_biometricos['HORA_DATETIME'].isna()
                df_biometricos.loc[mask_nulos, 'HORA_DATETIME'] = pd.to_datetime(
                    df_biometricos.loc[mask_nulos, 'HORA'], format='%H:%M', errors='coerce'
                )
        except:
            # Si falla, usar la hora original
            df_biometricos['HORA_DATETIME'] = df_biometricos['HORA']
        
        # Agrupar por cÃ³digo biomÃ©trico
        agg_dict = {
            'HORA': ['min', 'max'],  # Para mostrar
            'HORA_DATETIME': ['min', 'max'],  # Para cÃ¡lculos
            'CEDULA': 'first',
            'NOMBRE': 'first' if 'NOMBRE' in df_biometricos.columns else lambda x: 'N/A',
            'FECHA': 'first'
        }
        
        # Incluir CARGO si existe en el DataFrame
        if 'CARGO' in df_biometricos.columns:
            agg_dict['CARGO'] = 'first'
            
        df_agrupado = df_biometricos.groupby('codigo_biometrico').agg(agg_dict).reset_index()
        
        # Aplanar columnas
        base_columns = ['codigo_biometrico', 'hora_ingreso_str', 'hora_salida_str', 
                       'hora_ingreso_dt', 'hora_salida_dt', 'cedula', 'nombre', 'fecha']
        
        # Agregar CARGO si existe
        if 'CARGO' in df_biometricos.columns:
            base_columns.append('cargo')
            
        df_agrupado.columns = base_columns
        
        print(f"âœ… AgrupaciÃ³n completada: {len(df_agrupado)} registros Ãºnicos")
        
        # Convertir horas a formato AM/PM
        print("ðŸ• Convirtiendo horas a formato AM/PM...")
        
        def convertir_a_ampm(hora_str):
            """Convierte hora en formato HH:MM:SS a formato HH:MM AM/PM"""
            try:
                # Convertir a string y limpiar
                hora_str = str(hora_str).strip()
                
                if not hora_str or hora_str == 'nan' or hora_str == 'None' or hora_str == 'NaT':
                    return ''
                
                # Si viene de datetime, podrÃ­a tener formato "1900-01-01 HH:MM:SS"
                if ' ' in hora_str and len(hora_str.split(' ')) == 2:
                    hora_str = hora_str.split(' ')[1]  # Tomar solo la parte de la hora
                
                # Parsear la hora
                if ':' in hora_str:
                    partes = hora_str.split(':')
                    if len(partes) >= 2:
                        try:
                            hora = int(partes[0])
                            minuto = int(partes[1])
                            # Incluir segundos si estÃ¡n disponibles
                            segundo = int(partes[2]) if len(partes) >= 3 else 0
                        except ValueError:
                            return hora_str  # No se puede convertir a int
                    else:
                        return hora_str  # Formato no vÃ¡lido
                else:
                    return hora_str  # No tiene formato de hora
                
                # Convertir a formato AM/PM con segundos
                if hora == 0:
                    return f"12:{minuto:02d}:{segundo:02d} AM"
                elif hora < 12:
                    return f"{hora}:{minuto:02d}:{segundo:02d} AM"
                elif hora == 12:
                    return f"12:{minuto:02d}:{segundo:02d} PM"
                else:
                    return f"{hora-12}:{minuto:02d}:{segundo:02d} PM"
                    
            except Exception as e:
                print(f"âš ï¸  Error convirtiendo '{hora_str}': {e}")
                return str(hora_str)  # Retornar original si hay error
        
        # Convertir a string antes de aplicar conversiÃ³n AM/PM (el groupby puede devolver datetime objects)
        df_agrupado['hora_ingreso_str'] = df_agrupado['hora_ingreso_str'].astype(str)
        df_agrupado['hora_salida_str'] = df_agrupado['hora_salida_str'].astype(str)
        
        # Aplicar conversiÃ³n a las horas
        df_agrupado['hora_ingreso_ampm'] = df_agrupado['hora_ingreso_str'].apply(convertir_a_ampm)
        df_agrupado['hora_salida_ampm'] = df_agrupado['hora_salida_str'].apply(convertir_a_ampm)
        
        # Mostrar resultados con formato AM/PM
        print("ðŸŽ¯ Resultados del procesamiento:")
        for i, row in df_agrupado.head(3).iterrows():
            fecha_str = row['fecha'].strftime('%d/%m/%Y') if pd.notna(row['fecha']) else 'N/A'
            print(f"   CÃ³digo: {row['codigo_biometrico']} | CÃ©dula: {row['cedula']} | Fecha: {fecha_str}")
            print(f"     â†’ Ingreso: {row['hora_ingreso_ampm']} | Salida: {row['hora_salida_ampm']}")
        
        print(f"âœ… Horarios convertidos a formato AM/PM: {len(df_agrupado)} registros")
        
        # Preparar resultado con formato AM/PM
        resultado = {
            'codigos': df_agrupado['codigo_biometrico'].tolist(),
            'cedulas': df_agrupado['cedula'].tolist(),
            'nombres': df_agrupado['nombre'].tolist(),
            'fechas': df_agrupado['fecha'].tolist(),
            'ingresos': df_agrupado['hora_ingreso_ampm'].tolist(),  # Usar formato AM/PM
            'salidas': df_agrupado['hora_salida_ampm'].tolist()     # Usar formato AM/PM
        }
        
        # Incluir CARGO si estÃ¡ disponible
        if 'cargo' in df_agrupado.columns:
            resultado['cargos'] = df_agrupado['cargo'].tolist()
        else:
            resultado['cargos'] = []
        
        print(f"âœ… Procesamiento completado: {len(resultado['codigos'])} registros listos para integraciÃ³n con horarios AM/PM")
        return resultado
        
    except Exception as e:
        print(f"âŒ ERROR procesando archivo biomÃ©trico: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def procesar_reporte_calidad():
    """
    FunciÃ³n principal para procesar el reporte de calidad.
    
    Maneja la recepciÃ³n de archivos del formulario web y genera
    el reporte Excel completo con todas las hojas requeridas.
    
    Returns:
        JSON response con el resultado del procesamiento
    """
    try:
        logger.info("Iniciando procesamiento del Reporte de Calidad...")
        logger.info(f"Datos del formulario: {list(request.form.keys())}")
        logger.info(f"Archivos recibidos: {list(request.files.keys())}")
        
        # Verificar si hay archivo automatico del paso 3
        reporte3_auto_file = request.form.get('reporte3_auto_file')
        archivo_reporte3_content = None
        
        if reporte3_auto_file:
            print(f"INFO: Usando archivo automatico del Paso 3: {reporte3_auto_file}")
            # Usar archivo temporal del paso 3
            temp_filepath = os.path.join('temp_files', reporte3_auto_file)
            if os.path.exists(temp_filepath):
                with open(temp_filepath, 'rb') as f:
                    file_content = f.read()
                archivo_reporte3_content = io.BytesIO(file_content)
                print(f"OK: Archivo temporal cargado: {len(file_content)} bytes")
            else:
                print(f"ERROR: Archivo temporal no encontrado: {temp_filepath}")
                return jsonify({
                    'success': False,
                    'message': 'Archivo temporal del Paso 3 no encontrado'
                }), 400
        else:
            print("INFO: Intentando usar archivo manual...")
            # Usar archivo manual
            if 'excelFileReporte3' not in request.files:
                print("ERROR: No se encontro 'excelFileReporte3' en request.files")
                return jsonify({
                    'success': False,
                    'message': 'Se requiere el archivo de Reporte 3'
                }), 400
            
            archivo_reporte3 = request.files['excelFileReporte3']
            if archivo_reporte3.filename == '':
                print("ERROR: Archivo de Reporte 3 sin nombre valido")
                return jsonify({
                    'success': False,
                    'message': 'Archivo de Reporte 3 sin nombre valido'
                }), 400
            
            archivo_reporte3_content = archivo_reporte3
            print(f"OK: Archivo manual recibido: {archivo_reporte3.filename}")
        
        # Procesar archivo biomÃ©trico (opcional)
        datos_biometricos = procesar_archivo_biometricos()
        
        # Generar el reporte de calidad
        resultado = generar_reporte_calidad(archivo_reporte3_content, datos_biometricos)
        
        return jsonify({
            'success': True,
            'message': 'Reporte de calidad generado exitosamente',
            'filename': resultado['filename'],
            'estadisticas': resultado.get('estadisticas', {}),
            'temp_file': resultado['filename']
        })
        
    except Exception as e:
        logger.error(f"Error en procesar_reporte_calidad: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        }), 500

def descargar_reporte4():
    """
    Descarga el archivo temporal del Reporte 4
    """
    try:
        temp_filename = request.json.get('temp_file')
        if not temp_filename:
            return jsonify({'success': False, 'message': 'Nombre de archivo temporal no proporcionado'}), 400
        
        temp_filepath = os.path.join('temp_files', temp_filename)
        
        if not os.path.exists(temp_filepath):
            return jsonify({'success': False, 'message': 'Archivo temporal no encontrado'}), 404
        
        # Generar nombre de descarga amigable (sin el timestamp tÃ©cnico)
        nombre_descarga = temp_filename
        if temp_filename.startswith('4_Reporte_Calidad_'):
            # Mantener el nombre amigable del archivo
            nombre_descarga = temp_filename
        
        return send_file(
            temp_filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_descarga
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al descargar archivo: {str(e)}'}), 500

def generar_prueba_reporte4():
    """
    Genera un archivo de prueba del Reporte 4 sin necesidad de archivos de entrada
    """
    try:
        # Generar el reporte de calidad sin archivos de entrada
        resultado = generar_reporte_calidad(None)
        
        return jsonify({
            'success': True,
            'message': 'Archivo de prueba generado exitosamente',
            'filename': resultado['filename'],
            'estadisticas': resultado.get('estadisticas', {}),
            'temp_file': resultado['filename']
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generando archivo de prueba: {str(e)}'
        }), 500

def generar_reporte_calidad(archivo_reporte3, datos_biometricos=None):
    """
    Genera el archivo Excel del reporte de calidad con datos del Reporte 3 integrados.
    
    Args:
        archivo_reporte3: Archivo Excel del Reporte 3 con datos base
        
    Returns:
        dict: Resultado con informaciÃ³n del archivo generado y estadÃ­sticas
        
    Raises:
        Exception: Si hay errores en la generaciÃ³n del reporte
    """
    try:
        # Leer datos del Reporte 3 si existe
        df_reporte3 = None
        if archivo_reporte3:
            try:
                # Leer todas las hojas del archivo
                all_sheets = pd.read_excel(archivo_reporte3, sheet_name=None)
                print(f"Hojas encontradas en Reporte 3: {list(all_sheets.keys())}")
                
                # Combinar todas las hojas en un solo DataFrame
                dataframes_list = []
                total_registros = 0
                
                for sheet_name, df_sheet in all_sheets.items():
                    if not df_sheet.empty:
                        print(f"  - Procesando hoja '{sheet_name}': {len(df_sheet)} registros")
                        print(f"    Columnas: {list(df_sheet.columns)}")
                        
                        # Copiar sin modificaciones para preservar todos los datos
                        df_copy = df_sheet.copy()
                        dataframes_list.append(df_copy)
                        total_registros += len(df_copy)
                        print(f"  - OK: Hoja '{sheet_name}' agregada: {len(df_copy)} registros")
                    else:
                        print(f"WARN: Hoja '{sheet_name}' estÃ¡ vacÃ­a, omitida")
                
                print(f"INFO: Hojas procesadas: {len(dataframes_list)}, Registros totales: {total_registros}")
                
                if dataframes_list:
                    # Combinar todos los DataFrames SIN eliminar duplicados
                    df_reporte3 = pd.concat(dataframes_list, ignore_index=True, sort=False)
                    
                    # Verificar que no se perdieron registros
                    if len(df_reporte3) != total_registros:
                        print(f"ERROR: Se perdieron {total_registros - len(df_reporte3)} registros durante pd.concat")
                    else:
                        print(f"OK: Todos los registros preservados correctamente")
                    
                    # Procesar fechas SOLO si existe la columna y DESPUES de verificar integridad
                    if 'Fecha' in df_reporte3.columns:
                        # Procesamiento de fechas
                        print(f"INFO: Procesando fechas - Registros: {len(df_reporte3)}")
                        fechas_originales = df_reporte3['Fecha'].copy()
                        
                        # Convertir la columna Fecha a datetime usando mÃºltiples formatos
                        df_reporte3['Fecha'] = pd.to_datetime(df_reporte3['Fecha'], 
                                                            format='%d/%m/%Y', 
                                                            errors='coerce')
                        
                        # Si fallÃ³, intentar con otros formatos comunes
                        mask_nulos = df_reporte3['Fecha'].isna()
                        if mask_nulos.any():
                            print(f"INFO: Intentando formato alternativo para {mask_nulos.sum()} fechas")
                            # Restaurar valores originales para las fechas que fallaron
                            df_reporte3.loc[mask_nulos, 'Fecha'] = fechas_originales.loc[mask_nulos]
                            fechas_temp = pd.to_datetime(df_reporte3.loc[mask_nulos, 'Fecha'], 
                                                       format='%m/%d/%Y', 
                                                       errors='coerce')
                            df_reporte3.loc[mask_nulos, 'Fecha'] = fechas_temp
                            
                        # Si aun fallo, intentar formato YYYY-MM-DD
                        mask_nulos = df_reporte3['Fecha'].isna()
                        if mask_nulos.any():
                            print(f"  - Intentando formato YYYY-MM-DD para {mask_nulos.sum()} fechas...")
                            # Restaurar valores originales para las fechas que aun fallan
                            df_reporte3.loc[mask_nulos, 'Fecha'] = fechas_originales.loc[mask_nulos]
                            fechas_temp = pd.to_datetime(df_reporte3.loc[mask_nulos, 'Fecha'], 
                                                       format='%Y-%m-%d', 
                                                       errors='coerce')
                            df_reporte3.loc[mask_nulos, 'Fecha'] = fechas_temp
                        
                        # Ordenar por fecha SOLO si se pudieron convertir las fechas
                        fechas_validas = df_reporte3['Fecha'].notna().sum()
                        print(f"  - Fechas validas para ordenamiento: {fechas_validas} de {len(df_reporte3)}")
                        
                        if fechas_validas > 0:
                            df_reporte3 = df_reporte3.sort_values('Fecha', ascending=True, na_position='last')
                            print(f"  - Datos ordenados por fecha (registros sin fecha al final)")
                        
                        print(f"  - Registros despues de procesar fechas: {len(df_reporte3)}")
                        
                        # Convertir fecha SIEMPRE a formato DD/MM/YYYY
                        df_reporte3['Fecha'] = df_reporte3['Fecha'].apply(
                            lambda x: x.strftime('%d/%m/%Y') if pd.notnull(x) else ''
                        )
                        
                        print(f"OK: Fechas procesadas y formato DD/MM/YYYY aplicado")
                    else:
                        print(f"WARN: No se encontrÃ³ columna 'Fecha' para ordenar")
                    
                    # FILTRAR REGISTROS CON ASIGNACION MENOR O IGUAL A 5
                    registros_antes_filtro = len(df_reporte3)
                    
                    # Buscar la columna de asignacion (con o sin acento)
                    columna_asignacion = None
                    for col in df_reporte3.columns:
                        if col.lower().replace('Ã³', 'o').replace('Ã±', 'n') in ['asignacion']:
                            columna_asignacion = col
                            break
                    
                    if columna_asignacion:
                        print(f"INFO: Columna de asignacion encontrada: '{columna_asignacion}'")
                        # Convertir a numerico para poder filtrar
                        df_reporte3[columna_asignacion] = pd.to_numeric(df_reporte3[columna_asignacion], errors='coerce')
                        
                        # Mostrar estadÃ­sticas antes del filtro
                        valores_menores_5 = df_reporte3[df_reporte3[columna_asignacion] <= ConfigReporte.FILTRO_ASIGNACION_MIN]
                        print(f"INFO: Registros con {columna_asignacion} <= {ConfigReporte.FILTRO_ASIGNACION_MIN}: {len(valores_menores_5)}")
                        if len(valores_menores_5) > 0:
                            print(f"INFO: Valores a eliminar: {valores_menores_5[columna_asignacion].value_counts().sort_index().to_dict()}")
                        
                        # Filtrar registros con AsignaciÃ³n > ConfigReporte.FILTRO_ASIGNACION_MIN
                        df_reporte3 = df_reporte3[df_reporte3[columna_asignacion] > ConfigReporte.FILTRO_ASIGNACION_MIN]
                        registros_eliminados = registros_antes_filtro - len(df_reporte3)
                        print(f"INFO: Filtro aplicado - Registros: {registros_antes_filtro} â†’ {len(df_reporte3)} (eliminados: {registros_eliminados})")
                    else:
                        print(f"WARN: No se encontrÃ³ columna de AsignaciÃ³n para filtrar")
                    
                    print(f"OK: Datos procesados: {len(df_reporte3)} registros de {len(dataframes_list)} hojas")
                else:
                    print("WARN: No se encontraron datos en ninguna hoja del Reporte 3")
                    df_reporte3 = None
                    
            except Exception as e:
                print(f"ERROR: Error leyendo Reporte 3: {str(e)}")
                df_reporte3 = None
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Crear hojas en orden inverso para que Planta quede primera
            # Crear hoja "Consolidado" con referencias a cÃ³digos de Operativo
            crear_hoja_consolidado(writer, df_reporte3)
            
            # Crear hoja "Gerente" con datos de gerentes
            crear_hoja_gerente(writer, datos_biometricos)
            
            # Crear hoja "Team" con datos de supervisores
            crear_hoja_team(writer, datos_biometricos)
            
            # Crear hoja "Operativo" con datos del Reporte 3
            crear_hoja_operativo(writer, df_reporte3)
            
            # Crear hoja "Calidad"
            crear_hoja_calidad(writer)
            
            # Crear hoja "Ausentismo" con datos base del Reporte 3
            crear_hoja_ausentismo(writer, df_reporte3)
            
            # Sincronizar cÃ³digos ausentismo con datos biomÃ©tricos
            sincronizar_codigos_ausentismo(writer, datos_biometricos)
            
            # Crear hoja "Asistencia Lideres" con datos biomÃ©tricos
            crear_hoja_asistencia_lideres(writer, datos_biometricos)
            
            # Crear hoja "Planta" con las tres tablas
            crear_hoja_planta(writer)
        
        output.seek(0)
        
        # Generar nombre de archivo amigable basado en fechas de los datos
        filename = generar_nombre_archivo_calidad(df_reporte3)
        
        # Guardar archivo temporalmente
        os.makedirs('temp_files', exist_ok=True)
        temp_filepath = os.path.join('temp_files', filename)
        with open(temp_filepath, 'wb') as f:
            f.write(output.getvalue())
        
        return {
            'filename': filename,
            'estadisticas': {
                'hojas_creadas': ['Consolidado', 'Gerente', 'Team', 'Operativo', 'Calidad', 'Ausentismo', 'Asistencia Lideres', 'Planta'],
                'tablas_planta': 3,
                'total_hojas': 8
            }
        }
        
    except Exception as e:
        raise Exception(f"Error generando reporte: {str(e)}")

def crear_hoja_planta(writer):
    """
    Crea la hoja "Planta" con las tres tablas especificadas
    """
    # Tabla 1: Usuarios
    datos_usuarios = [
        ["William Cabiativa", "1016019347", "Gerente"],
        ["Daniela Arias", "1015447386", "Gerente"],
        ["Yesid Espitia", "1013666144", "Gerente"],
        ["Maria Acero", "1020798229", "Back"],
        ["Luis Aleman", "1003187750", "Team"],
        ["Nancy Rodriguez", "1014269628", "Team"],
        ["Danilo Rodriguez", "1016079907", "Team"],
        ["Camilo Arciniegas", "1003458325", "Team"],
        ["Brayan Murcia", "1010237867", "Team"],
        ["Edgar Parra", "1015456724", "Team"],
        ["Nancy Cruz", "1001296759", "Team"],
        ["Nicolas Briceno", "1013097231", "Back"],
        ["Kebin Bernal", "1033691778", "Back"],
        ["Natalia Quiceno", "1019108899", "Back"],
        ["Neverson Ulloa", "1003777394", "Back"],
        ["Zharik Jimenez", "1070590063", "Back"],
        ["Luisa Arevalo", "1031801240", "Back"],
        ["Paula Rubio", "1007155877", "Team"],
        ["Lizethe Rodriguez", "1105690146", "Back"],
        ["Andres Acevedo", "1000036873", "Back"]
    ]
    
    df_usuarios = pd.DataFrame(datos_usuarios, columns=["Usuario", "Cedula", "Cargo"])
    
    # Tabla 2: Dia Pago
    datos_dia_pago = [
        ["M0-PP", "37,0%"],
        ["M0-VP", "62,0%"],
        ["M1-1A", "9,0%"],
        ["M1-1B", "2,5%"],
        ["M0-PN", "52,0%"],
        ["M0-FRS", "52,0%"],
        ["M0-BT", "52,0%"],
        ["M0-1-PP", "10,0%"],
        ["M1-1A-FRS", "8,0%"],
        ["M1-1A-BT", "8,0%"],
        ["M1-1A-PN", "8,0%"]
    ]
    
    df_dia_pago = pd.DataFrame(datos_dia_pago, columns=["Dia Pago", "Meta"])
    
    # Tabla 3: Dia Normal
    datos_dia_normal = [
        ["M0-PP", "36,0%"],
        ["M0-VP", "58,0%"],
        ["M1-1A", "9,0%"],
        ["M1-1B", "2,5%"],
        ["M0-PN", "52,0%"],
        ["M0-FRS", "52,0%"],
        ["M0-BT", "52,0%"],
        ["M0-1-PP", "10,0%"],
        ["M1-1A-FRS", "8,0%"],
        ["M1-1A-BT", "8,0%"],
        ["M1-1A-PN", "8,0%"]
    ]
    
    df_dia_normal = pd.DataFrame(datos_dia_normal, columns=["Dia Normal", "Meta"])
    
    # Crear la hoja "Planta" directamente con pandas
    # Tabla 1: Usuarios (A1:C25)
    df_usuarios.to_excel(writer, sheet_name="Planta", startrow=0, startcol=0, index=False)
    
    # Obtener la hoja despues de crear la primera tabla
    worksheet = writer.sheets["Planta"]
    
    # Escribir las otras tablas manualmente para control de posicion
    # Tabla 2: Dia Pago (E1:F12)
    worksheet.cell(row=1, column=5, value="Dia Pago")
    worksheet.cell(row=1, column=6, value="Meta")
    for idx, row in enumerate(datos_dia_pago, start=2):
        worksheet.cell(row=idx, column=5, value=row[0])
        worksheet.cell(row=idx, column=6, value=row[1])
    
    # Tabla 3: Dia Normal (H1:I12)
    worksheet.cell(row=1, column=8, value="Dia Normal")
    worksheet.cell(row=1, column=9, value="Meta")
    for idx, row in enumerate(datos_dia_normal, start=2):
        worksheet.cell(row=idx, column=8, value=row[0])
        worksheet.cell(row=idx, column=9, value=row[1])
    
    # Crear tablas de Excel reales
    # Imports consolidados en el top del archivo
    
    # Tabla 1: Usuarios (rango dinÃ¡mico basado en datos reales)
    num_usuarios = len(datos_usuarios) + 1  # +1 para incluir la fila del header
    tabla_usuarios = Table(displayName="TablaUsuarios", ref=f"A1:C{num_usuarios}")
    style_usuarios = TableStyleInfo(name=TABLE_STYLE_BLUE, showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_usuarios.tableStyleInfo = style_usuarios
    worksheet.add_table(tabla_usuarios)
    
    # Tabla 2: Dia Pago  
    tabla_dia_pago = Table(displayName="TablaDiaPago", ref="E1:F12")
    style_dia_pago = TableStyleInfo(name=TABLE_STYLE_GREEN, showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_dia_pago.tableStyleInfo = style_dia_pago
    worksheet.add_table(tabla_dia_pago)
    
    # Tabla 3: Dia Normal
    tabla_dia_normal = Table(displayName="TablaDiaNormal", ref="H1:I12")
    style_dia_normal = TableStyleInfo(name=TABLE_STYLE_YELLOW, showFirstColumn=False,
                                     showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_dia_normal.tableStyleInfo = style_dia_normal
    worksheet.add_table(tabla_dia_normal)
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 20  # Usuario
    worksheet.column_dimensions['B'].width = 12  # Cedula
    worksheet.column_dimensions['C'].width = 10  # Cargo
    worksheet.column_dimensions['E'].width = 15  # Dia Pago
    worksheet.column_dimensions['F'].width = 8   # Meta Pago
    worksheet.column_dimensions['H'].width = 15  # Dia Normal
    worksheet.column_dimensions['I'].width = 8   # Meta Normal

def crear_hoja_asistencia_lideres(writer, datos_biometricos=None):
    """
    Crea la hoja "Asistencia Lideres" con datos filtrados de biomÃ©tricos
    Filtra por CARGO = "SUPERVISOR" y "GERENTE"
    """
    # Crear DataFrame con las columnas especificadas
    columnas_asistencia = [
        "Codigo Aus", "Tipo Jornada", "Fecha", "Usuario", "Cedula", 
        "Cargo", "Ingreso", "Salida", "Horas laboradas", "Novedad Ingreso", "Drive"
    ]
    
    registros_lideres = []
    
    if datos_biometricos is not None:
        print("ðŸ“Š Procesando datos biomÃ©tricos para Asistencia Lideres...")
        
        # Verificar si datos_biometricos tiene informaciÃ³n de cargo
        if 'cargos' not in datos_biometricos or not datos_biometricos['cargos']:
            print("âš ï¸ ADVERTENCIA: No se encontrÃ³ informaciÃ³n de 'CARGO' en datos biomÃ©tricos")
            print("ðŸ’¡ AsegÃºrese de que el archivo biomÃ©trico tenga una columna 'CARGO'")
        else:
            print(f"âœ… InformaciÃ³n de cargo disponible: {len(datos_biometricos['cargos'])} registros")
            
            # Filtrar Ã­ndices donde el cargo es SUPERVISOR o GERENTE
            indices_lideres = []
            for i, cargo in enumerate(datos_biometricos['cargos']):
                if str(cargo).upper() in ['SUPERVISOR', 'GERENTE']:
                    indices_lideres.append(i)
            
            print(f"âœ… Encontrados {len(indices_lideres)} registros de SUPERVISORES y GERENTES")
            
            if len(indices_lideres) > 0:
                # Procesar los registros filtrados
                print("ðŸ“‹ Procesando registros de lideres...")
                
                # Crear lista de registros temporales para facilitar el agrupamiento
                lideres_data = []
                for idx in indices_lideres:
                    lideres_data.append({
                        'fecha': datos_biometricos['fechas'][idx] if idx < len(datos_biometricos.get('fechas', [])) else '',
                        'nombre': datos_biometricos['nombres'][idx] if idx < len(datos_biometricos.get('nombres', [])) else '',
                        'cedula': datos_biometricos['cedulas'][idx] if idx < len(datos_biometricos.get('cedulas', [])) else '',
                        'cargo': datos_biometricos['cargos'][idx],
                        'ingreso': datos_biometricos['ingresos'][idx] if idx < len(datos_biometricos.get('ingresos', [])) else '',
                        'salida': datos_biometricos['salidas'][idx] if idx < len(datos_biometricos.get('salidas', [])) else ''
                    })
                
                # Crear DataFrame temporal para agrupamiento
                df_temp = pd.DataFrame(lideres_data)
                
                if not df_temp.empty:
                    print(f"ðŸ“Š Datos de lideres para procesar: {len(df_temp)} registros")
                    
                    # Agrupar por fecha, nombre, cedula, cargo (en caso de mÃºltiples registros por dÃ­a)
                    # Tomar el primer ingreso y la Ãºltima salida
                    grouped = df_temp.groupby(['fecha', 'nombre', 'cedula', 'cargo']).agg({
                        'ingreso': 'first',  # Primera entrada del dÃ­a
                        'salida': 'last'     # Ãšltima salida del dÃ­a  
                    }).reset_index()
                    
                    print(f"ðŸ“‹ Registros agrupados: {len(grouped)} registros Ãºnicos por dÃ­a")
                    
                    # Convertir cada registro agrupado
                    for _, row in grouped.iterrows():
                        fecha = row['fecha']
                        usuario = row['nombre']
                        cedula = row['cedula']
                        cargo = row['cargo']
                        ingreso = row['ingreso']
                        salida = row['salida']
                        
                        # Formatear fecha si es necesario
                        if pd.notna(fecha):
                            if isinstance(fecha, pd.Timestamp):
                                fecha_str = fecha.strftime('%d/%m/%Y')
                            else:
                                fecha_str = str(fecha)
                        else:
                            fecha_str = ''
                        
                        registro = [
                            "",  # Codigo Aus
                            "",  # Tipo Jornada
                            fecha_str,
                            usuario,
                            str(cedula),
                            cargo,
                            str(ingreso),
                            str(salida),
                            "",  # Horas laboradas (se calcularÃ¡ con fÃ³rmula)
                            "",  # Novedad Ingreso
                            ""   # Drive
                        ]
                        registros_lideres.append(registro)
                    
                    print(f"âœ… Procesados {len(registros_lideres)} registros para Asistencia Lideres")
    
    # Si no hay datos, crear una fila vacÃ­a
    if not registros_lideres:
        registros_lideres = [["", "", "", "", "", "", "", "", "", "", ""]]
        print("ðŸ“‹ No se encontraron datos de SUPERVISORES/GERENTES, creando hoja vacÃ­a")
    
    df_asistencia = pd.DataFrame(registros_lideres, columns=columnas_asistencia)
    
    # Escribir a Excel
    df_asistencia.to_excel(writer, sheet_name="Asistencia Lideres", index=False)
    
    # Obtener la hoja y aplicar formato basico
    worksheet = writer.sheets["Asistencia Lideres"]
    
    # Aplicar formato de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Formatear encabezados
    for col in range(1, 12):  # A hasta K
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 12  # Codigo Aus
    worksheet.column_dimensions['B'].width = 15  # Tipo Jornada
    worksheet.column_dimensions['C'].width = 12  # Fecha
    worksheet.column_dimensions['D'].width = 20  # Usuario
    worksheet.column_dimensions['E'].width = 12  # Cedula
    worksheet.column_dimensions['F'].width = 10  # Cargo
    worksheet.column_dimensions['G'].width = 10  # Ingreso
    worksheet.column_dimensions['H'].width = 10  # Salida
    worksheet.column_dimensions['I'].width = 15  # Horas laboradas
    worksheet.column_dimensions['J'].width = 18  # Novedad Ingreso
    worksheet.column_dimensions['K'].width = 10  # Drive
    
    # Agregar fÃ³rmulas para calcular horas laboradas automÃ¡ticamente
    # Solo aplicar fÃ³rmulas si hay datos reales (no solo la fila vacÃ­a)
    tiene_datos_reales = len(registros_lideres) > 1 or (len(registros_lideres) == 1 and registros_lideres[0][3])  # Verificar si hay usuario en la primera fila
    
    if tiene_datos_reales:
        print(f"ðŸ“Š Agregando fÃ³rmulas para calcular horas laboradas y tipo de jornada en {len(registros_lideres)} registros...")
        
        for row_num in range(2, len(registros_lideres) + 2):  # Empezar desde fila 2 (despuÃ©s del header)
            # Verificar que la fila tenga datos (no estÃ© vacÃ­a)
            if df_asistencia.iloc[row_num-2, 3]:  # Verificar columna Usuario (Ã­ndice 3)
                # Columna A es "Codigo Aus" (Ã­ndice 1)
                # FÃ³rmula: Cedula + DÃ­a(2 dÃ­gitos) + Mes(2 dÃ­gitos) (formato: E2 + TEXT(DAY(C2),"00") + TEXT(MONTH(C2),"00"))
                formula_codigo_aus = f'=IF(AND(E{row_num}<>"",C{row_num}<>""),E{row_num}&TEXT(DAY(C{row_num}),"00")&TEXT(MONTH(C{row_num}),"00"),"")'
                worksheet[f'A{row_num}'].value = formula_codigo_aus
                
                # Columna B es "Tipo Jornada" (Ã­ndice 2)
                # FÃ³rmula para determinar si es dÃ­a de pago (30,31,1,2,15,16,17) o dÃ­a normal
                formula_tipo_jornada = f'=IF(OR(DAY(C{row_num})=30,DAY(C{row_num})=31,DAY(C{row_num})=1,DAY(C{row_num})=2,DAY(C{row_num})=15,DAY(C{row_num})=16,DAY(C{row_num})=17),"Pago","Normal")'
                worksheet[f'B{row_num}'].value = formula_tipo_jornada
                
                # Columna I es "Horas laboradas" (Ã­ndice 9)
                # Columna G es "Ingreso" (Ã­ndice 7), Columna H es "Salida" (Ã­ndice 8)
                formula_horas = f'=IF(AND(G{row_num}<>"",H{row_num}<>""),TEXT(TIMEVALUE(H{row_num})-TIMEVALUE(G{row_num}),"[h]:mm"),"0:00")'
                worksheet[f'I{row_num}'].value = formula_horas
                
                # Columna J es "Novedad Ingreso" (Ã­ndice 10)
                # FÃ³rmula para determinar si llegÃ³ tarde segÃºn tipo de jornada (Normal: >8:00, Pago: >7:30)
                formula_novedad = f'=IFERROR(IF(AND(B{row_num}="Normal",TIMEVALUE(G{row_num})>TIME(8,0,0)),"Llego Tarde",IF(AND(B{row_num}="Pago",TIMEVALUE(G{row_num})>TIME(7,30,0)),"Llego Tarde","Sin Novedad")),"")'
                worksheet[f'J{row_num}'].value = formula_novedad
                
                # Columna K es "Drive" (Ã­ndice 11)
                # FÃ³rmula: Igual a la columna Novedad Ingreso
                formula_drive = f'=J{row_num}'
                worksheet[f'K{row_num}'].value = formula_drive
                
                print(f"  âœ… FÃ³rmulas aplicadas en fila {row_num}")
        
        # Agregar validaciÃ³n de datos (dropdown) para la columna Tipo Jornada
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Crear validaciÃ³n para opciones Pago/Normal
        dv_tipo_jornada = DataValidation(
            type="list",
            formula1='"Pago,Normal"',
            allow_blank=True
        )
        dv_tipo_jornada.error = "Seleccione Pago o Normal"
        dv_tipo_jornada.errorTitle = "Valor invÃ¡lido"
        
        # Aplicar validaciÃ³n a la columna B (Tipo Jornada) para todas las filas con datos
        rango_tipo_jornada = f"B2:B{len(registros_lideres) + 1}"
        dv_tipo_jornada.add(rango_tipo_jornada)
        worksheet.add_data_validation(dv_tipo_jornada)
        
        print(f"âœ… Dropdown agregado para Tipo Jornada en rango {rango_tipo_jornada}")
        print(f"âœ… FÃ³rmulas agregadas para calcular horas laboradas y tipo de jornada")
    else:
        print("ðŸ“‹ No hay datos reales de lÃ­deres, omitiendo fÃ³rmulas")
    
    # Aplicar formato de tabla
    aplicar_formato_tabla(worksheet, df_asistencia, "TablaAsistenciaLideres")

def crear_hoja_ausentismo(writer, df_reporte3=None):
    """
    Crea la hoja "Ausentismo" copiando datos bÃ¡sicos desde el Reporte 3.
    Las columnas Ingreso y Salida se dejan vacÃ­as para ser llenadas despuÃ©s por datos biomÃ©tricos.
    """
    print("=== CREANDO HOJA AUSENTISMO CON DATOS BASE ===")
    
    # Usar las columnas predefinidas para Ausentismo
    columnas_ausentismo = ConfigReporte.COLUMNAS_AUSENTISMO
    
    # Obtener datos bÃ¡sicos desde el DataFrame del Reporte 3
    registros_ausentismo = []
    
    if df_reporte3 is not None and not df_reporte3.empty:
        print(f"âœ… Procesando {len(df_reporte3)} registros del Reporte 3 para Ausentismo")
        print(f"Columnas disponibles: {list(df_reporte3.columns)}")
        
        # Mapeo de columnas del Reporte 3 a Ausentismo
        for index, row in df_reporte3.iterrows():
            # Extraer datos bÃ¡sicos (usando nombres mÃ¡s comunes del Reporte 3)
            codigo = row.get('CODIGO', row.get('Codigo', ''))
            tipo_jornada = row.get('Tipo Jornada', '')
            fecha = row.get('Fecha', '')
            cedula = row.get('Cedula', '')
            id_emp = row.get('ID', '')
            nombre = row.get('Nombre', row.get('NOMBRE', ''))
            sede = row.get('Sede', row.get('SEDE', ''))
            ubicacion = row.get('Ubicacion', row.get('UBICACION', ''))
            logueo = row.get('Logueo', row.get('LOGUEO', ''))
            
            # Crear registro para Ausentismo - TODAS las columnas vacÃ­as
            # Las fÃ³rmulas llenarÃ¡n automÃ¡ticamente los datos
            registro = [
                "",            # Codigo Aus (se llenarÃ¡ con fÃ³rmula: Cedula + DDMM)
                "",            # Codigo (se llenarÃ¡ con copia directa desde Operativo)
                "",            # Tipo Jornada (se llenarÃ¡ con INDEX/MATCH)
                "",            # Fecha (se llenarÃ¡ con INDEX/MATCH)
                "",            # Cedula (se llenarÃ¡ con INDEX/MATCH)
                "",            # ID (se llenarÃ¡ con INDEX/MATCH)
                "",            # Nombre (se llenarÃ¡ con INDEX/MATCH)
                "",            # Sede (se llenarÃ¡ con INDEX/MATCH)
                "",            # Ubicacion (se llenarÃ¡ con INDEX/MATCH)
                "",            # Logueo Admin (se llenarÃ¡ con INDEX/MATCH)
                "",            # Ingreso (vacÃ­o, se llenarÃ¡ con biomÃ©tricos)
                "",            # Salida (vacÃ­o, se llenarÃ¡ con biomÃ©tricos)
                "",            # Horas laboradas
                "",            # Novedad Ingreso
                "",            # Validacion
                ""             # Drive
            ]
            registros_ausentismo.append(registro)
        
        print(f"âœ… {len(registros_ausentismo)} registros preparados para Ausentismo")
        
    else:
        print("âš ï¸ No se encontrÃ³ Reporte 3 vÃ¡lido, creando hoja Ausentismo vacÃ­a")
    
    # Crear DataFrame con los registros (o vacÃ­o si no hay datos)
    if registros_ausentismo:
        df_ausentismo = pd.DataFrame(registros_ausentismo, columns=columnas_ausentismo)
    else:
        # Crear al menos una fila vacÃ­a para la estructura
        df_ausentismo = pd.DataFrame([["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]], columns=columnas_ausentismo)
    
    # Escribir a Excel
    df_ausentismo.to_excel(writer, sheet_name="Ausentismo", index=False)
    
    # Obtener la hoja y aplicar formato basico
    worksheet = writer.sheets["Ausentismo"]
    
    # Aplicar formato de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Formatear encabezados
    for col in range(1, 17):  # A hasta P
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 12  # Codigo Aus
    worksheet.column_dimensions['B'].width = 10  # Codigo
    worksheet.column_dimensions['C'].width = 15  # Tipo Jornada
    worksheet.column_dimensions['D'].width = 12  # Fecha
    worksheet.column_dimensions['E'].width = 12  # Cedula
    worksheet.column_dimensions['F'].width = 8   # ID
    worksheet.column_dimensions['G'].width = 20  # Nombre
    worksheet.column_dimensions['H'].width = 10  # Sede
    worksheet.column_dimensions['I'].width = 12  # Ubicacion
    worksheet.column_dimensions['J'].width = 15  # Logueo Admin
    worksheet.column_dimensions['K'].width = 10  # Ingreso
    worksheet.column_dimensions['L'].width = 10  # Salida
    worksheet.column_dimensions['M'].width = 15  # Horas laboradas
    worksheet.column_dimensions['N'].width = 18  # Novedad Ingreso
    worksheet.column_dimensions['O'].width = 12  # Validacion
    worksheet.column_dimensions['P'].width = 10  # Drive
    
    # Aplicar formato de tabla
    aplicar_formato_tabla(worksheet, df_ausentismo, "TablaAusentismo")
    
    # AGREGAR FÃ“RMULAS DE VLOOKUP para buscar datos en la tabla Operativo
    agregar_formulas_vlookup_ausentismo(worksheet, len(df_ausentismo), df_reporte3)

def agregar_formulas_vlookup_ausentismo(worksheet, num_registros, df_reporte3=None):
    """
    Agrega fÃ³rmulas en la tabla Ausentismo basado en el sistema del backup:
    - Columna Codigo: Copia directa desde Operativo
    - Otras columnas: INDEX/MATCH usando el cÃ³digo como referencia
    - Codigo Aus: Se genera con Cedula + DDMM
    """
    print("=== AGREGANDO FÃ“RMULAS EN TABLA AUSENTISMO (SISTEMA BACKUP) ===")
    
    # Mapeo de columnas segÃºn el backup
    mapeo_columnas = {
        # Ausentismo_col_index: (Operativo_col_letter, descripcion)
        2: ('A', 'Codigo'),           # Codigo <- CODIGO (copia directa)
        3: ('B', 'Tipo Jornada'),     # Tipo Jornada <- Tipo Jornada  
        4: ('C', 'Fecha'),            # Fecha <- Fecha
        5: ('D', 'Cedula'),           # Cedula <- Cedula
        6: ('E', 'ID'),               # ID <- ID
        7: ('H', 'Nombre'),           # Nombre <- Nombre
        8: ('I', 'Sede'),             # Sede <- Sede
        9: ('J', 'Ubicacion'),        # Ubicacion <- Ubicacion
        10: ('K', 'Logueo Admin'),    # Logueo Admin <- Logueo
    }
    
    # Agregar fÃ³rmulas para todas las columnas mapeadas
    for row in range(2, num_registros + 2):  # +2 porque empezamos en fila 2
        for ausentismo_col, (operativo_col_letter, desc) in mapeo_columnas.items():
            if ausentismo_col == 2:  # Columna Codigo - copia directa desde Operativo
                formula = f'=IF(Operativo!{operativo_col_letter}{row}<>"",Operativo!{operativo_col_letter}{row},"")'
            else:  # Otras columnas - busca usando el cÃ³digo como referencia
                # Usa INDEX/MATCH para buscar usando el cÃ³digo de la columna B
                formula = f'=IF(B{row}<>"",IFERROR(INDEX(Operativo!{operativo_col_letter}:{operativo_col_letter},MATCH(B{row},Operativo!A:A,0)),""),"")'
            
            cell = worksheet.cell(row=row, column=ausentismo_col)
            cell.value = formula
            
            if row <= 3:  # Solo mostrar las primeras 3 para debug
                print(f"  âœ… Fila {row}, Col {ausentismo_col}: {desc} -> {formula}")
    
    # Calcular y agregar CODIGO AUS (columna A) directamente como valores - Cedula + DDMM
    print("Calculando y agregando cÃ³digos Aus (Cedula + DDMM) como valores...")
    
    # Usar los datos del df_reporte3 para generar cÃ³digos directamente
    # Esto garantiza que los cÃ³digos estÃ©n disponibles para sincronizaciÃ³n inmediata
    codigos_generados = 0
    
    if df_reporte3 is not None and not df_reporte3.empty:
        # Iterar sobre los datos del DataFrame df_reporte3 
        for idx in range(min(num_registros, len(df_reporte3))):
            row_excel = idx + 2  # Excel empieza en fila 2 (fila 1 son headers)
            
            # Obtener datos del DataFrame
            cedula = df_reporte3.iloc[idx]['Cedula']
            fecha = df_reporte3.iloc[idx]['Fecha']
            
            if cedula and fecha:
                try:
                    # Parsear fecha (viene en formato DD/MM/YYYY del procesamiento previo)
                    fecha_obj = None
                    
                    if isinstance(fecha, str):
                        if '/' in fecha:
                            fecha_obj = datetime.strptime(fecha, '%d/%m/%Y')
                        else:
                            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
                    elif hasattr(fecha, 'day'):
                        fecha_obj = fecha
                    
                    # Solo proceder si tenemos una fecha vÃ¡lida
                    if fecha_obj:
                        # Generar cÃ³digo: CEDULA + DDMM
                        dia = f"{fecha_obj.day:02d}"
                        mes = f"{fecha_obj.month:02d}"
                        codigo_aus = f"{cedula}{dia}{mes}"
                        
                        # Escribir el cÃ³digo directamente como valor en la celda A
                        worksheet.cell(row=row_excel, column=1, value=codigo_aus)
                        codigos_generados += 1
                        
                        # Debug: Mostrar primeros 3 cÃ³digos generados
                        if codigos_generados <= 3:
                            print(f"  âœ… Fila {row_excel}: CÃ³digo '{codigo_aus}' (CÃ©dula: {cedula}, Fecha: {fecha})")
                        
                except Exception as e:
                    if codigos_generados <= 3:  # Solo mostrar errores de las primeras filas para debug
                        print(f"  âŒ Error generando cÃ³digo para fila {row_excel}: {str(e)}")
                    pass  # Continuar con el siguiente registro
        
        print(f"âœ… {codigos_generados} cÃ³digos Aus generados exitosamente como valores")
    else:
        print("âš ï¸ No se encontraron datos del DataFrame para generar cÃ³digos, usando fÃ³rmulas como fallback...")
        # Fallback a fÃ³rmulas si no hay DataFrame disponible
        for row in range(2, num_registros + 2):
            formula_codigo_aus = f'=IF(AND(E{row}<>"",D{row}<>""),E{row}&TEXT(DAY(DATEVALUE(D{row})),"00")&TEXT(MONTH(DATEVALUE(D{row})),"00"),"")'
            worksheet.cell(row=row, column=1, value=formula_codigo_aus)
            
            if row <= 3:
                print(f"  ðŸ”— Fila {row}: Codigo Aus -> {formula_codigo_aus}")
    
    print(f"âœ… Sistema de fÃ³rmulas implementado segÃºn backup:")
    print(f"   - Columna Codigo: Copia directa desde Operativo")
    print(f"   - Otras columnas: INDEX/MATCH usando cÃ³digo como referencia")
    print(f"   - Codigo Aus: Cedula + DDMM automÃ¡tico")
    print(f"   - Columnas Ingreso y Salida quedan libres para biomÃ©tricos")

def sincronizar_codigos_ausentismo(writer, datos_biometricos=None):
    """
    Sincroniza los cÃ³digos biomÃ©tricos con la tabla Ausentismo.
    Compara cÃ³digos generados (CEDULA + DIA + MES) con la columna 'Codigo Aus'
    y actualiza las columnas Ingreso y Salida con las horas min/max correspondientes.
    """
    print("=== SINCRONIZACIÃ“N BIOMÃ‰TRICOS CON AUSENTISMO ===")
    
    # Obtener la hoja de ausentismo
    hoja_ausentismo = writer.sheets["Ausentismo"]
    
    if datos_biometricos is None:
        print("âš ï¸ No hay datos biomÃ©tricos para sincronizar")
        print("Las columnas Ingreso y Salida permanecerÃ¡n vacÃ­as")
        return
    
    print(f"âœ… Datos biomÃ©tricos disponibles: {len(datos_biometricos['codigos'])} cÃ³digos")
    
    # Crear diccionario de cÃ³digos biomÃ©tricos para bÃºsqueda rÃ¡pida
    # cÃ³digo_biomÃ©trico -> (ingreso, salida)
    biometrico_dict = {}
    for i in range(len(datos_biometricos['codigos'])):
        codigo = str(datos_biometricos['codigos'][i]).strip()
        ingreso = datos_biometricos['ingresos'][i]
        salida = datos_biometricos['salidas'][i]
        biometrico_dict[codigo] = (ingreso, salida)
    
    print(f"ðŸ“Š Diccionario de bÃºsqueda creado con {len(biometrico_dict)} entradas")
    
    # Mostrar muestra del diccionario con formato AM/PM
    print("ðŸ” Muestra de cÃ³digos biomÃ©tricos (formato AM/PM):")
    for i, (codigo, horas) in enumerate(list(biometrico_dict.items())[:3]):
        print(f"   {codigo} â†’ Ingreso: {horas[0]}, Salida: {horas[1]}")
    
    # Columnas de Ingreso y Salida en la tabla Ausentismo
    col_ingreso = 11  # Columna K
    col_salida = 12   # Columna L
    col_codigo_aus = 1  # Columna A (Codigo Aus)
    
    coincidencias = 0
    filas_procesadas = 0
    
    # Forzar cÃ¡lculo de fÃ³rmulas antes de sincronizaciÃ³n
    print("âš¡ Forzando cÃ¡lculo de fÃ³rmulas en Excel...")
    try:
        # Guardar temporalmente para que Excel calcule las fÃ³rmulas
        writer.book.calculation.calcMode = 'automatic'
        writer.book.calculation.calcOnSave = True
    except:
        print("âš ï¸ No se pudo configurar el cÃ¡lculo automÃ¡tico")
    
    # Recorrer filas de la tabla Ausentismo (empezar desde fila 2, saltando headers)
    print("ðŸ”„ Procesando filas de la tabla Ausentismo...")
    
    # Crear lista de cÃ³digos debug para investigar el problema
    codigos_encontrados = []
    
    for row in range(2, hoja_ausentismo.max_row + 1):
        filas_procesadas += 1
        
        # Obtener el valor de Codigo Aus
        codigo_aus_celda = hoja_ausentismo.cell(row=row, column=col_codigo_aus)
        
        # Si es una fÃ³rmula, intentar evaluarla manualmente
        if isinstance(codigo_aus_celda.value, str) and codigo_aus_celda.value.startswith('='):
            # Es una fÃ³rmula, necesitamos obtener los valores manualmente
            # Obtener cÃ©dula y fecha para generar el cÃ³digo
            cedula_celda = hoja_ausentismo.cell(row=row, column=5)  # Columna E (Cedula)
            fecha_celda = hoja_ausentismo.cell(row=row, column=4)   # Columna D (Fecha)
            
            # Generar cÃ³digo manualmente siguiendo la misma lÃ³gica de la fÃ³rmula
            cedula = cedula_celda.value
            fecha = fecha_celda.value
            
            if cedula and fecha:
                # Intentar parsear la fecha si es string
                if isinstance(fecha, str):
                    try:
                        from datetime import datetime
                        if '/' in fecha:
                            fecha_obj = datetime.strptime(fecha, '%d/%m/%Y')
                        else:
                            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
                        dia = f"{fecha_obj.day:02d}"
                        mes = f"{fecha_obj.month:02d}"
                        codigo_aus = f"{cedula}{dia}{mes}"
                    except:
                        codigo_aus = None
                elif hasattr(fecha, 'day') and hasattr(fecha, 'month'):
                    dia = f"{fecha.day:02d}"
                    mes = f"{fecha.month:02d}"
                    codigo_aus = f"{cedula}{dia}{mes}"
                else:
                    codigo_aus = None
            else:
                codigo_aus = None
        else:
            codigo_aus = codigo_aus_celda.value
        
        # Validar y procesar cÃ³digo
        if codigo_aus is None or str(codigo_aus).strip() == "":
            continue
            
        codigo_aus_str = str(codigo_aus).strip()
        
        # Debug: Guardar cÃ³digos para anÃ¡lisis
        if len(codigos_encontrados) < 10:
            codigos_encontrados.append(codigo_aus_str)
        
        # Buscar coincidencia con cÃ³digos biomÃ©tricos
        if codigo_aus_str in biometrico_dict:
            ingreso, salida = biometrico_dict[codigo_aus_str]
            
            # Actualizar columnas Ingreso y Salida
            hoja_ausentismo.cell(row=row, column=col_ingreso).value = ingreso
            hoja_ausentismo.cell(row=row, column=col_salida).value = salida
            
            coincidencias += 1
            
            # Mostrar primeras 3 coincidencias para debug
            if coincidencias <= 3:
                print(f"   âœ… Fila {row}: CÃ³digo '{codigo_aus_str}' â†’ Ingreso: {ingreso} | Salida: {salida}")
        else:
            # No hay coincidencia biomÃ©trica, llenar con 00:00:00
            hoja_ausentismo.cell(row=row, column=col_ingreso).value = "00:00:00"
            hoja_ausentismo.cell(row=row, column=col_salida).value = "00:00:00"
    
    # Resumen de la sincronizaciÃ³n
    print(f"\nðŸ“ˆ RESUMEN DE SINCRONIZACIÃ“N:")
    print(f"   - Filas procesadas en Ausentismo: {filas_procesadas}")
    print(f"   - Coincidencias encontradas: {coincidencias}")
    print(f"   - CÃ³digos biomÃ©tricos disponibles: {len(biometrico_dict)}")
    if len(biometrico_dict) > 0:
        print(f"   - Porcentaje de coincidencia: {(coincidencias/len(biometrico_dict)*100):.1f}%")
    else:
        print(f"   - Porcentaje de coincidencia: 0.0% (sin datos biomÃ©tricos)")
    
    # Debug adicional para investigar falta de coincidencias
    if coincidencias == 0 and len(codigos_encontrados) > 0:
        print("\nðŸ” DEBUG - COMPARACIÃ“N DE CÃ“DIGOS:")
        print("   CÃ³digos de Ausentismo encontrados (primeros 5):")
        for i, codigo in enumerate(codigos_encontrados[:5]):
            print(f"     {i+1}. '{codigo}'")
        
        print("   CÃ³digos biomÃ©tricos disponibles (primeros 5):")
        for i, codigo in enumerate(list(biometrico_dict.keys())[:5]):
            print(f"     {i+1}. '{codigo}'")
        
        # Verificar longitudes y formatos
        if len(codigos_encontrados) > 0 and len(biometrico_dict) > 0:
            codigo_aus_sample = codigos_encontrados[0]
            codigo_bio_sample = list(biometrico_dict.keys())[0]
            print(f"   Longitud cÃ³digo Ausentismo: {len(codigo_aus_sample)} chars")
            print(f"   Longitud cÃ³digo BiomÃ©trico: {len(codigo_bio_sample)} chars")
    
    if coincidencias == 0:
        print("âš ï¸ ADVERTENCIA: No se encontraron coincidencias entre cÃ³digos")
        print("   Verificar que los cÃ³digos 'Codigo Aus' tengan el formato CEDULA+DDMM")
    else:
        print(f"âœ… Ã‰XITO: {coincidencias} registros actualizados con horarios biomÃ©tricos")
    
    # Agregar fÃ³rmulas de "Horas laboradas" en la columna M (Ã­ndice 13)
    print("ðŸ“ Agregando fÃ³rmulas de Horas Laboradas...")
    col_horas_laboradas = 13  # Columna M
    
    for row in range(2, filas_procesadas + 2):  # Desde fila 2 hasta la Ãºltima fila procesada
        # FÃ³rmula para calcular diferencia de tiempo en formato HH:MM
        # =IFERROR(TEXT(TIMEVALUE(L{row})-TIMEVALUE(K{row}),"[h]:mm"),"0:00")
        formula = f'=IFERROR(TEXT(TIMEVALUE(L{row})-TIMEVALUE(K{row}),"[h]:mm"),"0:00")'
        hoja_ausentismo.cell(row=row, column=col_horas_laboradas).value = formula
    
    print(f"âœ… FÃ³rmulas de Horas Laboradas agregadas: {filas_procesadas} fÃ³rmulas (formato HH:MM)")
    
    # Agregar fÃ³rmulas de "Novedad Ingreso" en la columna N (Ã­ndice 14)
    print("ðŸ“ Agregando fÃ³rmulas de Novedad Ingreso...")
    col_novedad_ingreso = 14  # Columna N
    
    for row in range(2, filas_procesadas + 2):  # Desde fila 2 hasta la Ãºltima fila procesada
        # FÃ³rmula para validar horarios segÃºn tipo de jornada:
        # Si Tipo Jornada = "Normal" y Ingreso > 8:00 AM â†’ "Llego Tarde"
        # Si Tipo Jornada = "Pago" y Ingreso > 7:30 AM â†’ "Llego Tarde"  
        # Si cumple horario â†’ "Sin Novedad"
        # Si hay error â†’ ""
        formula = f'''=IFERROR(IF(AND(C{row}="Normal",TIMEVALUE(K{row})>TIME(8,0,0)),"Llego Tarde",IF(AND(C{row}="Pago",TIMEVALUE(K{row})>TIME(7,30,0)),"Llego Tarde","Sin Novedad")),"")'''
        hoja_ausentismo.cell(row=row, column=col_novedad_ingreso).value = formula
    
    # Agregar validaciÃ³n de datos (checklist) para la columna Novedad Ingreso
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Crear validaciÃ³n de datos con lista desplegable
    dv = DataValidation(
        type="list",
        formula1='"Sin Novedad,Llego Tarde"',
        allow_blank=True
    )
    dv.error = "Por favor seleccione una opciÃ³n vÃ¡lida"
    dv.errorTitle = "Entrada no vÃ¡lida"
    dv.prompt = "Seleccione: Sin Novedad o Llego Tarde"
    dv.promptTitle = "Novedad de Ingreso"
    
    # Aplicar validaciÃ³n a todo el rango de la columna N (desde fila 2 hasta la Ãºltima)
    rango_validacion = f"N2:N{filas_procesadas + 1}"
    dv.add(rango_validacion)
    hoja_ausentismo.add_data_validation(dv)
    
    print(f"âœ… FÃ³rmulas de Novedad Ingreso agregadas: {filas_procesadas} fÃ³rmulas")
    print(f"âœ… ValidaciÃ³n de datos (checklist) aplicada al rango: {rango_validacion}")
    
    # Agregar fÃ³rmulas de "Validacion" en la columna O (Ã­ndice 15)
    print("ðŸ“ Agregando fÃ³rmulas de ValidaciÃ³n...")
    col_validacion = 15  # Columna O
    
    for row in range(2, filas_procesadas + 2):  # Desde fila 2 hasta la Ãºltima fila procesada
        # FÃ³rmula para comparar Logueo Admin < Ingreso
        # =[@[Logueo Admin]]<[@Ingreso] adaptada a referencias de celda
        formula = f"=J{row}<K{row}"
        hoja_ausentismo.cell(row=row, column=col_validacion).value = formula
    
    print(f"âœ… FÃ³rmulas de ValidaciÃ³n agregadas: {filas_procesadas} fÃ³rmulas")
    
    # Agregar fÃ³rmulas de "Drive" en la columna P (Ã­ndice 16)
    print("ðŸ“ Agregando fÃ³rmulas de Drive...")
    col_drive = 16  # Columna P
    
    for row in range(2, filas_procesadas + 2):  # Desde fila 2 hasta la Ãºltima fila procesada
        # FÃ³rmula para referenciar la columna Novedad Ingreso
        formula = f"=N{row}"
        hoja_ausentismo.cell(row=row, column=col_drive).value = formula
    
    print(f"âœ… FÃ³rmulas de Drive agregadas: {filas_procesadas} fÃ³rmulas")

def leer_archivo_monitoreos():
    """
    Lee el archivo de monitoreos desde la carpeta 'Monitoreos'
    
    Returns:
        pd.DataFrame: DataFrame con los datos de monitoreos o DataFrame vacÃ­o si no encuentra el archivo
    """
    try:
        # Crear carpeta Monitoreos si no existe (con M mayÃºscula)
        if not os.path.exists('Monitoreos'):
            os.makedirs('Monitoreos')
            print("INFO: Carpeta 'Monitoreos' creada")
        
        # Buscar archivos en la carpeta Monitoreos
        patrones_archivos = [
            'Monitoreos/*.xlsx',
            'Monitoreos/*.xls', 
            'Monitoreos/*.csv'
        ]
        
        archivo_encontrado = None
        for patron in patrones_archivos:
            archivos = glob.glob(patron)
            if archivos:
                # Tomar el archivo mÃ¡s reciente
                archivo_encontrado = max(archivos, key=os.path.getmtime)
                break
        
        if not archivo_encontrado:
            print("ADVERTENCIA: No se encontrÃ³ archivo de monitoreos en la carpeta 'Monitoreos'")
            return pd.DataFrame()
        
        print(f"INFO: Cargando archivo de monitoreos: {archivo_encontrado}")
        
        # Leer el archivo segÃºn su extensiÃ³n
        if archivo_encontrado.endswith('.csv'):
            df_monitoreos = pd.read_csv(archivo_encontrado)
        else:
            df_monitoreos = pd.read_excel(archivo_encontrado)
        
        # Limpiar columnas unnamed (columnas vacÃ­as)
        df_monitoreos = df_monitoreos.loc[:, ~df_monitoreos.columns.str.contains('^Unnamed')]
        
        # Verificar que tenga las columnas requeridas
        columnas_requeridas = ['Fecha Monitoreo', 'ID Asesor', 'VOZ', 'SMS', 'TERCERO', 'Nota Total', 'Total Monitoreos']
        columnas_disponibles = [col for col in columnas_requeridas if col in df_monitoreos.columns]
        columnas_faltantes = [col for col in columnas_requeridas if col not in df_monitoreos.columns]
        
        if columnas_faltantes:
            print(f"ADVERTENCIA: Columnas faltantes en archivo de monitoreos: {columnas_faltantes}")
            
        print(f"INFO: Columnas encontradas: {columnas_disponibles}")
        print(f"INFO: Archivo de monitoreos cargado exitosamente: {len(df_monitoreos)} registros")
        
        return df_monitoreos
        
    except Exception as e:
        print(f"ERROR: Error al leer archivo de monitoreos: {str(e)}")
        return pd.DataFrame()


def crear_hoja_calidad(writer):
    """
    Crea la hoja "Calidad" con datos del archivo de monitoreos
    """
    # Leer datos de monitoreos
    df_monitoreos = leer_archivo_monitoreos()
    
    # Definir columnas esperadas (agregando Codigo como primera columna)
    columnas_calidad = [
        "Codigo", "Fecha Monitoreo", "ID Asesor", "VOZ", "SMS", 
        "TERCERO", "Nota Total", "Total Monitoreos"
    ]
    
    # Si tenemos datos de monitoreos, usarlos; si no, crear DataFrame vacÃ­o con las columnas
    if not df_monitoreos.empty:
        # Filtrar y reorganizar columnas segÃºn lo requerido (sin la columna Codigo del archivo original)
        columnas_disponibles = [col for col in columnas_calidad[1:] if col in df_monitoreos.columns]  # Excluir "Codigo" de la bÃºsqueda
        df_calidad = df_monitoreos[columnas_disponibles].copy()
        
        # Agregar columna Codigo como primera columna (vacÃ­a inicialmente, se llenarÃ¡ con fÃ³rmulas)
        df_calidad.insert(0, 'Codigo', '')
        
        # Procesar fecha para que se muestre correctamente
        if 'Fecha Monitoreo' in df_calidad.columns:
            df_calidad['Fecha Monitoreo'] = pd.to_datetime(df_calidad['Fecha Monitoreo'], errors='coerce').dt.strftime('%d/%m/%Y')
        
        # Asegurar que ID Asesor sea string sin decimales
        if 'ID Asesor' in df_calidad.columns:
            df_calidad['ID Asesor'] = df_calidad['ID Asesor'].fillna(0).astype(int).astype(str)
        
        # Procesar columnas de porcentaje - convertir a decimal para formato de Excel
        columnas_porcentaje = ['VOZ', 'SMS', 'TERCERO', 'Nota Total']
        for col in columnas_porcentaje:
            if col in df_calidad.columns:
                # Convertir a numÃ©rico y mantener como decimal (ya estÃ¡n en formato 0-1)
                df_calidad[col] = pd.to_numeric(df_calidad[col], errors='coerce')
        
        # Agregar columnas faltantes como vacÃ­as
        for col in columnas_calidad:
            if col not in df_calidad.columns:
                df_calidad[col] = ""
        
        # Reorganizar columnas en el orden correcto
        df_calidad = df_calidad[columnas_calidad]
        
        # FILTRAR FILAS CON VALORES PROBLEMÃTICOS
        # Eliminar filas que tengan exactamente el nÃºmero 0 o un solo nÃºmero en columnas importantes
        filas_antes = len(df_calidad)
        
        def es_fila_problematica(row):
            """Verifica si una fila tiene solo 0s o nÃºmeros Ãºnicos problemÃ¡ticos"""
            # Verificar columnas numÃ©ricas importantes
            columnas_verificar = ['ID Asesor', 'VOZ', 'SMS', 'TERCERO', 'Nota Total', 'Total Monitoreos']
            
            for col in columnas_verificar:
                if col in row.index:
                    valor = row[col]
                    # Si el valor es string y es exactamente "0"
                    if isinstance(valor, str) and valor.strip() == '0':
                        return True
                    # Si el valor es numÃ©rico y es exactamente 0
                    elif pd.notna(valor) and valor == 0:
                        # Solo eliminar si TODAS las columnas importantes son 0 o vacÃ­as
                        valores_importantes = [row[c] for c in columnas_verificar if c in row.index]
                        valores_validos = [v for v in valores_importantes if pd.notna(v) and v != 0 and str(v).strip() != '0']
                        if len(valores_validos) == 0:
                            return True
            
            return False
        
        # Aplicar filtro
        mask_mantener = ~df_calidad.apply(es_fila_problematica, axis=1)
        df_calidad = df_calidad[mask_mantener].reset_index(drop=True)
        
        filas_despues = len(df_calidad)
        filas_eliminadas = filas_antes - filas_despues
        
        if filas_eliminadas > 0:
            print(f"INFO: Eliminadas {filas_eliminadas} filas con valores problemÃ¡ticos (0s o nÃºmeros Ãºnicos)")
        
        # GENERAR CÃ“DIGOS antes de escribir a Excel
        print("INFO: Generando cÃ³digos en hoja Calidad...")
        
        def generar_codigo(row):
            try:
                id_asesor = str(row['ID Asesor']).strip()
                fecha_str = str(row['Fecha Monitoreo']).strip()
                
                if id_asesor and fecha_str and id_asesor != '0' and fecha_str != '':
                    # Parsear la fecha en formato DD/MM/YYYY
                    if '/' in fecha_str:
                        parts = fecha_str.split('/')
                        if len(parts) == 3:
                            dia = parts[0].zfill(2)
                            mes = parts[1].zfill(2)
                            return f"{id_asesor}{dia}{mes}"
                return ""
            except:
                return ""
        
        # Aplicar la funciÃ³n para generar cÃ³digos
        df_calidad['Codigo'] = df_calidad.apply(generar_codigo, axis=1)
        
        print(f"INFO: CÃ³digos generados para {len(df_calidad[df_calidad['Codigo'] != ''])} registros")
        print(f"INFO: Hoja Calidad creada con {len(df_calidad)} registros de monitoreos")
    else:
        # Crear DataFrame vacÃ­o con las columnas requeridas
        df_calidad = pd.DataFrame(columns=columnas_calidad)
        print("INFO: Hoja Calidad creada sin datos (archivo de monitoreos no encontrado)")
    
    # Escribir a Excel
    df_calidad.to_excel(writer, sheet_name="Calidad", index=False)
    
    # Obtener la hoja y aplicar formato bÃ¡sico
    worksheet = writer.sheets["Calidad"]
    
    # Aplicar formato de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Formatear encabezados
    for col in range(1, len(columnas_calidad) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 15  # Codigo
    worksheet.column_dimensions['B'].width = 18  # Fecha Monitoreo
    worksheet.column_dimensions['C'].width = 12  # ID Asesor
    worksheet.column_dimensions['D'].width = 8   # VOZ
    worksheet.column_dimensions['E'].width = 8   # SMS
    worksheet.column_dimensions['F'].width = 10  # TERCERO
    worksheet.column_dimensions['G'].width = 12  # Nota Total
    worksheet.column_dimensions['H'].width = 18  # Total Monitoreos
    
    # Aplicar formato de porcentaje a las columnas especificadas
    porcentaje_format = '0.00%'
    columnas_porcentaje = ['VOZ', 'SMS', 'TERCERO', 'Nota Total']
    
    # Obtener las letras de columna para las columnas de porcentaje
    columnas_porcentaje_indices = {}
    for idx, col_name in enumerate(columnas_calidad, 1):
        if col_name in columnas_porcentaje:
            columnas_porcentaje_indices[col_name] = idx
    
    # Aplicar formato de porcentaje a todas las celdas de esas columnas
    if not df_calidad.empty:
        for col_name, col_idx in columnas_porcentaje_indices.items():
            col_letter = chr(64 + col_idx)  # Convertir nÃºmero a letra (A=1, B=2, etc.)
            for row_num in range(2, len(df_calidad) + 2):  # Empezar desde fila 2 (despuÃ©s del encabezado)
                cell = worksheet[f"{col_letter}{row_num}"]
                cell.number_format = porcentaje_format
    
    # Aplicar formato de tabla si hay datos
    if not df_calidad.empty:
        aplicar_formato_tabla(worksheet, df_calidad, "TablaCalidad")

def verificar_integridad_datos(df_operativo, mapeo_columnas):
    """
    Verifica la integridad de los datos transferidos del Reporte 3 al Operativo
    
    Args:
        df_operativo (pd.DataFrame): DataFrame del operativo con datos mapeados
        mapeo_columnas (dict): Diccionario de mapeo de columnas
    """
    print("\nINFO: Verificando integridad de datos:")
    
    # Verificar columnas clave
    columnas_clave = ['ID', 'EXT', 'VOIP', 'Nombre', 'Fecha']
    for col in columnas_clave:
        if col in df_operativo.columns and col in mapeo_columnas:
            valores_vacios = df_operativo[col].isna().sum() + (df_operativo[col] == '').sum()
            total = len(df_operativo)
            pct_completo = (total - valores_vacios) / total * 100 if total > 0 else 0
            
            if pct_completo >= 90:
                status = "OK:"
            elif pct_completo >= 70:
                status = "WARN: "
            else:
                status = "ERROR:"
                
            print(f"  {status} {col}: {pct_completo:.1f}% completo ({total - valores_vacios}/{total})")
            
            # Mostrar muestra de datos para columnas importantes
            if col in ['VOIP', 'EXT'] and valores_vacios < total:
                muestra = df_operativo[col].dropna().head(3).tolist()
                print(f"    Muestra: {muestra}")
    
    # Verificar datos numericos
    columnas_numericas = ['Asignacion', 'PAGOS', 'Total toques', 'Llamadas Microsip', 'Total Llamadas']
    for col in columnas_numericas:
        if col in df_operativo.columns and col in mapeo_columnas:
            try:
                suma_total = pd.to_numeric(df_operativo[col], errors='coerce').sum()
                print(f"  INFO: {col}: Suma total = {suma_total:,.0f}")
            except:
                print(f"  WARN:  {col}: Error al procesar valores numericos")

def obtener_mapeo_columnas_operativo():
    """
    Retorna el mapeo de columnas entre Reporte 3 y Operativo
    """
    return {
        # Mapeos directos (nombres exactos)
        'Fecha': 'Fecha',
        'Cedula': 'Cedula', 
        'ID': 'ID',
        'EXT': 'EXT',
        'VOIP': 'VOIP',
        'Nombre': 'Nombre',
        'Sede': 'Sede',
        'Ubicacion': 'Ubicacion',
        'Logueo': 'Logueo',
        'Mora': 'Mora',
        'Asignacion': 'Asignacion',
        'PAGOS': 'PAGOS',
        'Total toques': 'Total toques',
        'Ultimo Toque': 'Ultimo Toque',
        'Llamadas Microsip': 'Llamadas Microsip',
        'Llamadas VOIP': 'Llamadas VOIP',
        'Total Llamadas': 'Total Llamadas',
        'Gerencia': 'Gerencia',
        'Team': 'Team',
        'Meta': 'Meta',
        'Ejecucion': 'Ejecucion',
        
        # Mapeos con nombres diferentes
        'Cliente gestionados 11 am': 'Clientes gestionados 11 am',
        'Capital Asignado': 'Capital Asignado',
        'Capital Recuperado': 'Capital Recuperado',
        '% Recuperado': '% Recuperado',
        '% Cuentas': '% Cuentas',
        'Ind Logueo': 'Ind Logueo',
        'Ind Ultimo': 'Ind Ultimo', 
        'Ind Ges Medio': 'Ind Ges Medio',
        'Ind Llamadas': 'Ind Llamadas',
        'Indicador Toques': 'Indicador Toques',
        'Ind Pausa': 'Ind Pausa',
        'Total Infracciones': 'Total Infracciones',
        'Total Operativo': 'Total Operativo'
    }

def obtener_columnas_operativo():
    """
    Retorna las columnas de la hoja Operativo
    """
    return [
        "CODIGO", "Tipo Jornada", "Fecha", "Cedula", "ID", "EXT", "VOIP", "Nombre", "Sede", "Ubicacion",
        "Logueo", "Mora", "Asignacion", "Clientes gestionados 11 am", "Capital Asignado", "Capital Recuperado",
        "PAGOS", "% Recuperado", "% Cuentas", "Total toques", "Ultimo Toque", "Llamadas Microsip",
        "Llamadas VOIP", "Total Llamadas", "Gerencia", "Team", "Meta", "Ejecucion", "Ind Logueo",
        "Ind Ultimo", "Ind Ges Medio", "Ind Llamadas", "Indicador Toques", "Ind Pausa", "Total Infracciones", "Total Operativo"
    ]

def crear_hoja_operativo(writer, df_reporte3=None):
    """
    Crea la hoja "Operativo" con la tabla especificada e integra datos del Reporte 3
    """
    # Obtener configuraciones usando las funciones auxiliares
    columnas_operativo = obtener_columnas_operativo()
    mapeo_explicito = obtener_mapeo_columnas_operativo()
    
    # Crear DataFrame con datos del Reporte 3 si existe
    if df_reporte3 is not None and not df_reporte3.empty:
        print("Integrando datos del Reporte 3 en hoja Operativo...")
        print(f"Columnas disponibles en Reporte 3: {list(df_reporte3.columns)}")
        
        # Mapeo automÃ¡tico inteligente
        mapeo_columnas = {}
        
        # Primero aplicar mapeos explicitos
        for col_operativo, col_reporte3_esperado in mapeo_explicito.items():
            if col_reporte3_esperado in df_reporte3.columns:
                mapeo_columnas[col_operativo] = col_reporte3_esperado
                print(f"OK: Mapeo directo: {col_operativo} â† {col_reporte3_esperado}")
        
        # FunciÃ³n auxiliar para normalizar texto y mejorar coincidencias
        def normalizar_texto(texto):
            """Normaliza texto para mejorar coincidencias de mapeo"""
            import unicodedata
            # Quitar acentos y convertir a minÃºsculas
            texto_normalizado = unicodedata.normalize('NFD', texto)
            texto_sin_acentos = ''.join(c for c in texto_normalizado if unicodedata.category(c) != 'Mn')
            return texto_sin_acentos.lower().strip()
        
        # Luego buscar coincidencias por similitud para columnas no mapeadas
        for col_reporte3 in df_reporte3.columns:
            col_r3_lower = col_reporte3.lower().strip()
            col_r3_normalizado = normalizar_texto(col_reporte3)
            
            for col_operativo in columnas_operativo:
                if col_operativo in mapeo_columnas:
                    continue  # Ya esta mapeado
                    
                col_op_lower = col_operativo.lower().strip()
                col_op_normalizado = normalizar_texto(col_operativo)
                
                # Coincidencias exactas (sin espacios/case)
                if col_r3_lower == col_op_lower:
                    mapeo_columnas[col_operativo] = col_reporte3
                    print(f"OK: Mapeo exacto: {col_operativo} <- {col_reporte3}")
                    continue
                
                # Coincidencias normalizadas (sin acentos)
                if col_r3_normalizado == col_op_normalizado:
                    mapeo_columnas[col_operativo] = col_reporte3
                    print(f"OK: Mapeo normalizado: {col_operativo} <- {col_reporte3}")
                    continue
                
                # Coincidencias parciales inteligentes
                palabras_operativo = set(col_op_normalizado.replace('%', 'porcentaje').split())
                palabras_reporte3 = set(col_r3_normalizado.replace('%', 'porcentaje').split())
                
                # Si al menos 80% de palabras coinciden
                if len(palabras_operativo) > 0:
                    coincidencia = len(palabras_operativo & palabras_reporte3) / len(palabras_operativo)
                    if coincidencia >= 0.8:
                        mapeo_columnas[col_operativo] = col_reporte3
                        print(f"OK: Mapeo por similitud ({coincidencia:.1%}): {col_operativo} <- {col_reporte3}")
        
        print(f"\nMapeo final encontrado ({len(mapeo_columnas)} columnas):")
        for k, v in mapeo_columnas.items():
            print(f"  {k} <- {v}")
        
        # Identificar columnas no mapeadas
        no_mapeadas = [col for col in columnas_operativo if col not in mapeo_columnas]
        if no_mapeadas:
            print(f"\nWARN:  Columnas no mapeadas ({len(no_mapeadas)}): {no_mapeadas}")
        
        # Identificar columnas del Reporte 3 no utilizadas
        utilizadas = set(mapeo_columnas.values())
        no_utilizadas = [col for col in df_reporte3.columns if col not in utilizadas]
        if no_utilizadas:
            print(f"\nINFO: Columnas del Reporte 3 no utilizadas ({len(no_utilizadas)}): {no_utilizadas}")
        
        print(f"\nINFO: Resumen del mapeo:")
        print(f"  - Columnas del Reporte 3: {len(df_reporte3.columns)}")
        print(f"  - Columnas del Operativo: {len(columnas_operativo)}")
        print(f"  - Columnas mapeadas: {len(mapeo_columnas)}")
        print(f"  - Eficiencia de mapeo: {len(mapeo_columnas)/len(columnas_operativo)*100:.1f}%")
        
        # Crear DataFrame con los datos mapeados - MEJORADO
        print("\nINFO: Procesando datos del Reporte 3...")
        print(f"INFO: Registros originales en Reporte 3: {len(df_reporte3)}")
        
        # Crear DataFrame directamente con mapeo de columnas (mas eficiente)
        df_operativo = pd.DataFrame()
        
        for col_operativo in columnas_operativo:
            if col_operativo in mapeo_columnas:
                col_reporte3 = mapeo_columnas[col_operativo]
                # Preservar el tipo de datos original, pero manejar fechas especialmente
                if col_operativo in ['Fecha']:
                    # Para columnas de fecha, copiar directamente (ya estan en formato DD/MM/YYYY)
                    df_operativo[col_operativo] = df_reporte3[col_reporte3].copy()
                    print(f"  INFO: Columna {col_operativo}: copiada directamente con formato DD/MM/YYYY")
                else:
                    df_operativo[col_operativo] = df_reporte3[col_reporte3].copy()
                
                # Verificar si hay valores nulos y reportar
                nulos = df_operativo[col_operativo].isna().sum()
                if nulos > 0:
                    print(f"  WARN:  {col_operativo}: {nulos} valores nulos de {len(df_operativo)}")
            else:
                # Columna no mapeada - inicializar con valores vacios
                df_operativo[col_operativo] = ""
        
        print(f"INFO: Registros procesados en hoja Operativo: {len(df_operativo)}")
        
        # Verificar que no se hayan perdido registros
        if len(df_operativo) != len(df_reporte3):
            print(f"INFO: ADVERTENCIA: Se perdieron {len(df_reporte3) - len(df_operativo)} registros durante el procesamiento")
        else:
            print(f"OK: Todos los registros fueron procesados correctamente")
        
        print(f"OK: Datos integrados: {len(df_operativo)} registros, {len([c for c in columnas_operativo if c in mapeo_columnas])} columnas con datos")
        
        # Verificar integridad de datos clave
        verificar_integridad_datos(df_operativo, mapeo_columnas)
        
    else:
        # Crear DataFrame con una fila de ejemplo si no hay datos del Reporte 3
        datos_ejemplo = [[""] * len(columnas_operativo)]
        df_operativo = pd.DataFrame(datos_ejemplo, columns=columnas_operativo)
        print("No hay datos del Reporte 3, creando hoja con estructura vacia")
    
    # Verificacion final de formatos antes de escribir a Excel
    # VerificaciÃ³n final de datos de fecha
    for col in ['Fecha']:
        if col in df_operativo.columns:
            # Asegurar que las fechas sean strings y estÃ©n en formato DD/MM/YYYY
            df_operativo[col] = df_operativo[col].astype(str)
            # Limpiar valores que no sean fechas vÃ¡lidas
            df_operativo[col] = df_operativo[col].replace('NaT', '').replace('nan', '')
            print(f"INFO: Formato de fechas verificado para columna {col}")
            df_operativo[col] = df_operativo[col].replace(['nan', 'NaT', 'None'], '')
            
            # Validar y corregir formato DD/MM/YYYY si es necesario
            import re
            for idx, fecha_val in enumerate(df_operativo[col]):
                if fecha_val and fecha_val != '':
                    # Verificar si esta en formato MM/DD/YYYY y convertir a DD/MM/YYYY
                    if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', fecha_val):
                        try:
                            # Intentar parsear como DD/MM/YYYY primero
                            fecha_dt = pd.to_datetime(fecha_val, format='%d/%m/%Y', errors='raise')
                            df_operativo.iloc[idx, df_operativo.columns.get_loc(col)] = fecha_dt.strftime('%d/%m/%Y')
                        except:
                            try:
                                # Si falla, intentar como MM/DD/YYYY y convertir a DD/MM/YYYY
                                fecha_dt = pd.to_datetime(fecha_val, format='%m/%d/%Y', errors='raise')
                                df_operativo.iloc[idx, df_operativo.columns.get_loc(col)] = fecha_dt.strftime('%d/%m/%Y')
                                print(f"    - Corregida fecha {fecha_val} â†’ {fecha_dt.strftime('%d/%m/%Y')}")
                            except:
                                print(f"    - No se pudo parsear fecha: {fecha_val}")
            
            print(f"  - Despues de correccion - Valores vacios: {df_operativo[col].eq('').sum()}")
            print(f"  - Muestra final: {df_operativo[col].head(3).tolist()}")
            print(f"OK: Columna {col} convertida y validada en formato DD/MM/YYYY")
    
    # Escribir a Excel
    df_operativo.to_excel(writer, sheet_name="Operativo", index=False)
    
    # Obtener la hoja y aplicar formato basico
    worksheet = writer.sheets["Operativo"]
    
    # AGREGAR FORMULAS EN LA COLUMNA CODIGO (primera columna)
    # Encontrar las posiciones de las columnas ID y Fecha
    id_col = None
    fecha_col = None
    for i, col_name in enumerate(columnas_operativo):
        if col_name == "ID":
            id_col = i + 1  # Excel columnas empiezan en 1
        elif col_name == "Fecha":
            fecha_col = i + 1
    
    if id_col and fecha_col:
        from openpyxl.utils import get_column_letter
        id_letter = get_column_letter(id_col)
        fecha_letter = get_column_letter(fecha_col)
        
        # Agregar formulas en la columna CODIGO para cada fila de datos
        # Formato: ID + DDMM (ej: 2020 + 0908 = 20200908)
        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
            # Formula Excel para extraer dia y mes de fecha DD/MM/YYYY con validacion
            # Si fecha es 09/08/2025, resultado debe ser: ID + "0908"
            formula = f'=IF(AND({id_letter}{row}<>"",{fecha_letter}{row}<>""),{id_letter}{row}&TEXT(DAY(DATEVALUE({fecha_letter}{row})),"00")&TEXT(MONTH(DATEVALUE({fecha_letter}{row})),"00"),"")'
            cell = worksheet.cell(row=row, column=1)  # Columna A (CODIGO)
            cell.value = formula
        print(f"Formulas agregadas en columna CODIGO: {len(df_operativo)} formulas con DATEVALUE, validacion de celdas vacias y formato ID+DDMM")
        
        # AGREGAR FORMULAS EN LA COLUMNA TIPO JORNADA (segunda columna)
        # Encontrar la posicion de la columna Tipo Jornada
        tipo_jornada_col = None
        for i, col_name in enumerate(columnas_operativo):
            if col_name == "Tipo Jornada":
                tipo_jornada_col = i + 1  # Excel columnas empiezan en 1
                break
        
        if tipo_jornada_col and fecha_col:
            tipo_jornada_letter = get_column_letter(tipo_jornada_col)
            
            # Agregar formulas en la columna Tipo Jornada para cada fila de datos
            # Si dia es 30, 31, 1, 2, 15, 16, 17 â†’ "Pago", sino â†’ "Normal"
            for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
                formula = f'=IF(OR(DAY({fecha_letter}{row})=30,DAY({fecha_letter}{row})=31,DAY({fecha_letter}{row})=1,DAY({fecha_letter}{row})=2,DAY({fecha_letter}{row})=15,DAY({fecha_letter}{row})=16,DAY({fecha_letter}{row})=17),"Pago","Normal")'
                cell = worksheet.cell(row=row, column=tipo_jornada_col)
                cell.value = formula
            print(f"Formulas agregadas en columna Tipo Jornada: {len(df_operativo)} formulas con logica Pago/Normal (dias: 30,31,1,2,15,16,17)")
            
            # Agregar validacion de datos (dropdown) en la columna Tipo Jornada
            from openpyxl.worksheet.datavalidation import DataValidation
            
            # Crear validacion con lista desplegable
            dv = DataValidation(type="list", formula1='"Pago,Normal"', allow_blank=False)
            dv.error = 'Seleccione: Pago o Normal'
            dv.errorTitle = 'Valor invalido'
            dv.prompt = 'Seleccione el tipo de jornada'
            dv.promptTitle = 'Tipo Jornada'
            
            # Aplicar validacion a todas las filas de datos
            range_validation = f"{tipo_jornada_letter}2:{tipo_jornada_letter}{len(df_operativo) + 1}"
            dv.add(range_validation)
            worksheet.add_data_validation(dv)
            print(f"Validacion de datos agregada en columna Tipo Jornada: rango {range_validation}")
        else:
            print("ERROR: No se encontro columna Tipo Jornada para las formulas")
            
        # AGREGAR FORMULAS EN LA COLUMNA META
        # Encontrar las posiciones de las columnas necesarias
        meta_col = None
        mora_col = None
        capital_asignado_col = None
        for i, col_name in enumerate(columnas_operativo):
            if col_name == "Meta":
                meta_col = i + 1  # Excel columnas empiezan en 1
            elif col_name == "Mora":
                mora_col = i + 1
            elif col_name == "Capital Asignado":
                capital_asignado_col = i + 1
        
        if meta_col and mora_col and tipo_jornada_col and capital_asignado_col:
            meta_letter = get_column_letter(meta_col)
            mora_letter = get_column_letter(mora_col)
            capital_asignado_letter = get_column_letter(capital_asignado_col)
            
            # Agregar formulas en la columna Meta para cada fila de datos
            for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
                # Formula con validacion de celdas vacias:
                # 1. Verificar que Mora, Capital Asignado y Tipo Jornada no esten vacios
                # 2. Si Mora = "M1-2" â†’ calcular 500000 / Capital Asignado como porcentaje
                # 3. Si Tipo Jornada = "Pago" â†’ buscar en Planta!E:F (Dia Pago)
                # 4. Si Tipo Jornada = "Normal" â†’ buscar en Planta!H:I (Dia Normal)
                formula = f'=IF(AND({mora_letter}{row}<>"",{capital_asignado_letter}{row}<>"",{tipo_jornada_letter}{row}<>""),IF({mora_letter}{row}="M1-2",ROUND(500000/{capital_asignado_letter}{row},4),IF({tipo_jornada_letter}{row}="Pago",VLOOKUP({mora_letter}{row},Planta!E:F,2,FALSE),IF({tipo_jornada_letter}{row}="Normal",VLOOKUP({mora_letter}{row},Planta!H:I,2,FALSE),""))),"")'
                cell = worksheet.cell(row=row, column=meta_col)
                cell.value = formula
            print(f"Formulas agregadas en columna Meta: {len(df_operativo)} formulas con validacion de celdas vacias, VLOOKUP a tablas de Planta y calculo especial para M1-2")
            
            # AGREGAR FORMULAS EN LA COLUMNA EJECUCION
            # Encontrar las posiciones de las columnas necesarias para Ejecucion
            ejecucion_col = None
            percent_recuperado_col = None
            percent_cuentas_col = None
            for i, col_name in enumerate(columnas_operativo):
                if col_name == "Ejecucion":
                    ejecucion_col = i + 1
                elif col_name == "% Recuperado":
                    percent_recuperado_col = i + 1
                elif col_name == "% Cuentas":
                    percent_cuentas_col = i + 1
            
            if ejecucion_col and percent_recuperado_col and percent_cuentas_col:
                ejecucion_letter = get_column_letter(ejecucion_col)
                percent_recuperado_letter = get_column_letter(percent_recuperado_col)
                percent_cuentas_letter = get_column_letter(percent_cuentas_col)
                
                # Agregar formulas en la columna Ejecucion para cada fila de datos
                for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
                    # Formula con validacion de celdas vacias:
                    # 1. Verificar que Mora, Meta, % Recuperado y % Cuentas no esten vacios
                    # 2. Si Mora="M0-PP" â†’ % Cuentas / Meta, sino â†’ % Recuperado / Meta
                    formula = f'=IF(AND({mora_letter}{row}<>"",{meta_letter}{row}<>"",{percent_recuperado_letter}{row}<>"",{percent_cuentas_letter}{row}<>""),IF({mora_letter}{row}="M0-PP",{percent_cuentas_letter}{row}/{meta_letter}{row},{percent_recuperado_letter}{row}/{meta_letter}{row}),"")'
                    cell = worksheet.cell(row=row, column=ejecucion_col)
                    cell.value = formula
                print(f"Formulas agregadas en columna Ejecucion: {len(df_operativo)} formulas con validacion de celdas vacias y logica M0-PP")
                
                # AGREGAR FORMULAS EN LA COLUMNA IND LOGUEO
                # Encontrar las posiciones de las columnas necesarias para Ind Logueo
                ind_logueo_col = None
                logueo_col = None
                for i, col_name in enumerate(columnas_operativo):
                    if col_name == "Ind Logueo":
                        ind_logueo_col = i + 1
                    elif col_name == "Logueo":
                        logueo_col = i + 1
                
                if ind_logueo_col and logueo_col:
                    ind_logueo_letter = get_column_letter(ind_logueo_col)
                    logueo_letter = get_column_letter(logueo_col)
                    
                    # Agregar formulas en la columna Ind Logueo para cada fila de datos
                    for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
                        # Formula con validacion de celdas vacias:
                        # 1. Verificar que Logueo y Tipo Jornada no esten vacios
                        # 2. Si Tipo Jornada="Pago" y Logueoâ‰¤7:30 AM â†’ 0.15
                        # 3. Si Tipo Jornada="Normal" y Logueoâ‰¤8:00 AM â†’ 0.15, sino â†’ 0
                        formula = f'=IF(AND({logueo_letter}{row}<>"",{tipo_jornada_letter}{row}<>""),IF({tipo_jornada_letter}{row}="Pago",IF(TIMEVALUE({logueo_letter}{row})<=TIMEVALUE("7:30:00 AM"),0.15,0),IF({tipo_jornada_letter}{row}="Normal",IF(TIMEVALUE({logueo_letter}{row})<=TIMEVALUE("8:00:00 AM"),0.15,0),0)),0)'
                        cell = worksheet.cell(row=row, column=ind_logueo_col)
                        cell.value = formula
                    print(f"Formulas agregadas en columna Ind Logueo: {len(df_operativo)} formulas con TIMEVALUE para formato AM/PM")
                    
                    # AGREGAR FORMULAS EN LA COLUMNA IND ULTIMO
                    # Encontrar las posiciones de las columnas necesarias para Ind Ultimo
                    ind_ultimo_col = None
                    ultimo_toque_col = None
                    fecha_col = None
                    for i, col_name in enumerate(columnas_operativo):
                        if col_name == "Ind Ultimo":
                            ind_ultimo_col = i + 1
                        elif col_name == "Ultimo Toque":
                            ultimo_toque_col = i + 1
                        elif col_name == "Fecha":
                            fecha_col = i + 1
                    
                    if ind_ultimo_col and ultimo_toque_col and fecha_col:
                        ind_ultimo_letter = get_column_letter(ind_ultimo_col)
                        ultimo_toque_letter = get_column_letter(ultimo_toque_col)
                        fecha_letter = get_column_letter(fecha_col)
                        
                        # Agregar formulas en la columna Ind Ultimo para cada fila de datos
                        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
                            # Formula con validacion de celdas vacias y conversion de fecha:
                            # 1. Verificar que Fecha, Ultimo Toque y Tipo Jornada no esten vacios
                            # 2. Convertir fecha DD/MM/YYYY a formato que Excel entienda para WEEKDAY
                            # 3. Si es sabado: Ultimo Toque >= 12:20 PM â†’ 0.15
                            # 4. Si es pago (no sabado): Ultimo Toque >= 6:50 PM â†’ 0.15  
                            # 5. Si es normal (no sabado): Ultimo Toque >= 5:20 PM â†’ 0.15
                            # 6. Sino â†’ 0
                            formula = f'=IF(AND({fecha_letter}{row}<>"",{ultimo_toque_letter}{row}<>"",{tipo_jornada_letter}{row}<>""),IF(WEEKDAY(DATEVALUE({fecha_letter}{row}))=7,IF(TIMEVALUE({ultimo_toque_letter}{row})>=TIMEVALUE("12:20:00 PM"),0.15,0),IF({tipo_jornada_letter}{row}="Pago",IF(TIMEVALUE({ultimo_toque_letter}{row})>=TIMEVALUE("6:50:00 PM"),0.15,0),IF({tipo_jornada_letter}{row}="Normal",IF(TIMEVALUE({ultimo_toque_letter}{row})>=TIMEVALUE("5:20:00 PM"),0.15,0),0))),0)'
                            cell = worksheet.cell(row=row, column=ind_ultimo_col)
                            cell.value = formula
                        print(f"Formulas agregadas en columna Ind Ultimo: {len(df_operativo)} formulas con DATEVALUE para conversion de fecha, validacion de celdas vacias, sabados (12:20 PM) y horarios PM")
                    else:
                        print("ERROR: No se encontraron columnas Ind Ultimo, Ultimo Toque y/o Fecha para las formulas")
                else:
                    print("ERROR: No se encontraron columnas Ind Logueo y/o Logueo para las formulas")
            else:
                print("ERROR: No se encontraron columnas Ejecucion, % Recuperado y/o % Cuentas para las formulas")
        else:
            print("ERROR: No se encontraron columnas Meta, Mora, Tipo Jornada y/o Capital Asignado para las formulas")
    else:
        print("ERROR: No se encontraron columnas ID y/o Fecha para las formulas")
    
    # AGREGAR VALORES FIJOS EN LA COLUMNA IND PAUSA
    # Encontrar la posicion de la columna Ind Pausa
    ind_pausa_col = None
    for i, col_name in enumerate(columnas_operativo):
        if col_name == "Ind Pausa":
            ind_pausa_col = i + 1
            break
    
    if ind_pausa_col:
        # Agregar valor fijo ConfigReporte.VALOR_INDICADOR_CUMPLE en la columna Ind Pausa para cada fila de datos
        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
            cell = worksheet.cell(row=row, column=ind_pausa_col)
            cell.value = ConfigReporte.VALOR_INDICADOR_CUMPLE
        print(f"Valores fijos {ConfigReporte.VALOR_INDICADOR_CUMPLE} agregados en columna Ind Pausa: {len(df_operativo)} filas")
    else:
        print("ERROR: No se encontro columna Ind Pausa")
    
    # AGREGAR FORMULAS EN LA COLUMNA IND GES MEDIO
    # Encontrar las posiciones de las columnas necesarias para Ind Ges Medio
    ind_ges_medio_col = None
    asignacion_col = None
    clientes_gestionados_col = None
    
    for i, col_name in enumerate(columnas_operativo):
        if col_name == "Ind Ges Medio":
            ind_ges_medio_col = i + 1
        elif col_name == "Asignacion":
            asignacion_col = i + 1
        elif col_name == "Clientes gestionados 11 am":
            clientes_gestionados_col = i + 1
    
    if ind_ges_medio_col and asignacion_col and clientes_gestionados_col:
        ind_ges_medio_letter = get_column_letter(ind_ges_medio_col)
        asignacion_letter = get_column_letter(asignacion_col)
        clientes_gestionados_letter = get_column_letter(clientes_gestionados_col)
        
        # Agregar formulas en la columna Ind Ges Medio para cada fila de datos
        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
            # Formula con validacion de celdas vacias y logica condicional:
            # 1. Verificar que Asignacion y Clientes gestionados no esten vacios
            # 2. Si Asignacion < 45: Clientes gestionados >= ROUND(Asignacion*0.9,0) â†’ 0.15, sino â†’ 0
            # 3. Si Asignacion >= 45: Clientes gestionados >= 45 â†’ 0.15, sino â†’ 0
            # Aplicar redondeo al 90% de asignacion: 35.1â†’35, 35.6â†’36
            formula = f'=IF(AND({asignacion_letter}{row}<>"",{clientes_gestionados_letter}{row}<>""),IF({asignacion_letter}{row}<45,IF({clientes_gestionados_letter}{row}>=ROUND({asignacion_letter}{row}*0.9,0),0.15,0),IF({clientes_gestionados_letter}{row}>=45,0.15,0)),0)'
            cell = worksheet.cell(row=row, column=ind_ges_medio_col)
            cell.value = formula
        print(f"Formulas agregadas en columna Ind Ges Medio: {len(df_operativo)} formulas con validacion de gestion media (<45: >=90% Asignacion redondeado, >=45: >=45)")
    else:
        print("ERROR: No se encontraron columnas Ind Ges Medio, Asignacion y/o Clientes gestionados 11 am")
    
    # AGREGAR FORMULAS EN LA COLUMNA INDICADOR TOQUES
    # Encontrar las posiciones de las columnas necesarias para Indicador Toques
    indicador_toques_col = None
    mora_col = None
    total_toques_col = None
    
    for i, col_name in enumerate(columnas_operativo):
        if col_name == "Indicador Toques":
            indicador_toques_col = i + 1
        elif col_name == "Mora":
            mora_col = i + 1
        elif col_name == "Total toques":
            total_toques_col = i + 1
    
    if indicador_toques_col and mora_col and total_toques_col:
        mora_letter = get_column_letter(mora_col)
        total_toques_letter = get_column_letter(total_toques_col)
        
        # Agregar formulas en la columna Indicador Toques para cada fila de datos
        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
            # Formula con validacion de cartera que CONTENGA M0 vs otras carteras:
            # 1. Verificar que Mora y Total toques no esten vacios
            # 2. Si Mora CONTIENE "M0": Total toques >= 120 â†’ 0.2, sino â†’ 0
            # 3. Si Mora NO CONTIENE "M0": Total toques >= 160 â†’ 0.2, sino â†’ 0
            # Usar ISNUMBER(SEARCH()) para buscar "M0" dentro del texto
            formula = f'=IF(AND({mora_letter}{row}<>"",{total_toques_letter}{row}<>""),IF(ISNUMBER(SEARCH("M0",{mora_letter}{row})),IF({total_toques_letter}{row}>=120,0.2,0),IF({total_toques_letter}{row}>=160,0.2,0)),0)'
            cell = worksheet.cell(row=row, column=indicador_toques_col)
            cell.value = formula
        print(f"Formulas agregadas en columna Indicador Toques: {len(df_operativo)} formulas con validacion de toques (CONTIENE M0: >=120, Otras: >=160)")
    else:
        print("ERROR: No se encontraron columnas Indicador Toques, Mora y/o Total toques")
    
    # AGREGAR FORMULAS EN LA COLUMNA IND LLAMADAS
    # Encontrar las posiciones de las columnas necesarias para Ind Llamadas
    ind_llamadas_col = None
    total_llamadas_col = None
    
    for i, col_name in enumerate(columnas_operativo):
        if col_name == "Ind Llamadas":
            ind_llamadas_col = i + 1
        elif col_name == "Total Llamadas":
            total_llamadas_col = i + 1
    
    if ind_llamadas_col and total_llamadas_col:
        total_llamadas_letter = get_column_letter(total_llamadas_col)
        
        # Agregar formulas en la columna Ind Llamadas para cada fila de datos
        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
            # Formula con validacion de Total Llamadas >= 150:
            # 1. Verificar que Total Llamadas no este vacio
            # 2. Si Total Llamadas >= 150 â†’ 0.2, sino â†’ 0
            formula = f'=IF({total_llamadas_letter}{row}<>"",IF({total_llamadas_letter}{row}>=150,0.2,0),0)'
            cell = worksheet.cell(row=row, column=ind_llamadas_col)
            cell.value = formula
        print(f"Formulas agregadas en columna Ind Llamadas: {len(df_operativo)} formulas con validacion de llamadas (>=150)")
    else:
        print("ERROR: No se encontraron columnas Ind Llamadas y/o Total Llamadas")
    
    # AGREGAR FORMULAS EN LA COLUMNA TOTAL INFRACCIONES
    # Encontrar las posiciones de las columnas necesarias
    total_infracciones_col = None
    ind_logueo_col = None
    ind_pausa_col = None
    
    for i, col_name in enumerate(columnas_operativo):
        if col_name == "Total Infracciones":
            total_infracciones_col = i + 1
        elif col_name == "Ind Logueo":
            ind_logueo_col = i + 1
        elif col_name == "Ind Pausa":
            ind_pausa_col = i + 1
    
    if total_infracciones_col and ind_logueo_col and ind_pausa_col:
        ind_logueo_letter = get_column_letter(ind_logueo_col)
        ind_pausa_letter = get_column_letter(ind_pausa_col)
        
        # Agregar formulas en la columna Total Infracciones para cada fila de datos
        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
            # Formula: Contar valores que NO son 0.15 (es decir, contar infracciones/ceros)
            # Como los indicadores usan 0.15 para "cumple" y 0 para "no cumple", contamos los 0s
            formula = f'=COUNTIF({ind_logueo_letter}{row}:{ind_pausa_letter}{row},0)'
            cell = worksheet.cell(row=row, column=total_infracciones_col)
            cell.value = formula
        print(f"Formulas agregadas en columna Total Infracciones: {len(df_operativo)} formulas COUNTIF para contar ceros (infracciones)")
    else:
        print("ERROR: No se encontraron columnas Total Infracciones, Ind Logueo y/o Ind Pausa")
    
    # AGREGAR FORMULAS EN LA COLUMNA TOTAL OPERATIVO
    # Encontrar la posicion de la columna Total Operativo
    total_operativo_col = None
    for i, col_name in enumerate(columnas_operativo):
        if col_name == "Total Operativo":
            total_operativo_col = i + 1
            break
    
    if total_operativo_col and ind_logueo_col and ind_pausa_col:
        total_operativo_letter = get_column_letter(total_operativo_col)
        
        # Agregar formulas en la columna Total Operativo para cada fila de datos
        for row in range(2, len(df_operativo) + 2):  # Empezar en fila 2 (despues del header)
            # Formula: SUM del rango Ind Logueo hasta Ind Pausa (resultado en porcentaje)
            formula = f'=SUM({ind_logueo_letter}{row}:{ind_pausa_letter}{row})'
            cell = worksheet.cell(row=row, column=total_operativo_col)
            cell.value = formula
        print(f"Formulas agregadas en columna Total Operativo: {len(df_operativo)} formulas SUM (formato porcentaje)")
    else:
        print("ERROR: No se encontraron columnas Total Operativo, Ind Logueo y/o Ind Pausa")
    
    # Aplicar formato de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    
    # Formatear encabezados
    for col in range(1, len(columnas_operativo) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Aplicar formato de porcentaje a las columnas % Recuperado y % Cuentas
    from openpyxl.utils import get_column_letter
    
    # Crear diccionario de columnas para encontrar las posiciones
    column_letters = {}
    for col_num, col_name in enumerate(columnas_operativo, 1):
        column_letters[col_name] = get_column_letter(col_num)
    
    # Formato de porcentaje para las columnas de porcentaje
    if '% Recuperado' in column_letters and '% Cuentas' in column_letters:
        percent_format = '0.00%'
        # Aplicar formato a todas las filas de datos existentes
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            # Columna % Recuperado
            percent_recup_cell = f"{column_letters['% Recuperado']}{row_num}"
            worksheet[percent_recup_cell].number_format = percent_format
            
            # Columna % Cuentas
            percent_cuentas_cell = f"{column_letters['% Cuentas']}{row_num}"
            worksheet[percent_cuentas_cell].number_format = percent_format
        print(f"OK: Formato de porcentaje aplicado a {max_row-1} filas en columnas % Recuperado y % Cuentas")
    
    # Aplicar formato de porcentaje a la columna Meta
    if 'Meta' in column_letters:
        percent_format = '0.00%'
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            meta_cell = f"{column_letters['Meta']}{row_num}"
            worksheet[meta_cell].number_format = percent_format
        print(f"OK: Formato de porcentaje aplicado a {max_row-1} filas en columna Meta")
    
    # Aplicar formato de porcentaje a la columna Ejecucion
    if 'Ejecucion' in column_letters:
        percent_format = '0.00%'
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            ejecucion_cell = f"{column_letters['Ejecucion']}{row_num}"
            worksheet[ejecucion_cell].number_format = percent_format
        print(f"OK: Formato de porcentaje aplicado a {max_row-1} filas en columna Ejecucion")
    
    # Aplicar formato de porcentaje a la columna Total Operativo
    if 'Total Operativo' in column_letters:
        percent_format = '0.00%'
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            total_operativo_cell = f"{column_letters['Total Operativo']}{row_num}"
            worksheet[total_operativo_cell].number_format = percent_format
        print(f"OK: Formato de porcentaje aplicado a {max_row-1} filas en columna Total Operativo")
    
    # Aplicar formato numerico a la columna Ind Logueo
    if 'Ind Logueo' in column_letters:
        decimal_format = '0.00'
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            ind_logueo_cell = f"{column_letters['Ind Logueo']}{row_num}"
            worksheet[ind_logueo_cell].number_format = decimal_format
        print(f"OK: Formato decimal aplicado a {max_row-1} filas en columna Ind Logueo")
    
    # Aplicar formato numerico a la columna Ind Ultimo
    if 'Ind Ultimo' in column_letters:
        decimal_format = '0.00'
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            ind_ultimo_cell = f"{column_letters['Ind Ultimo']}{row_num}"
            worksheet[ind_ultimo_cell].number_format = decimal_format
        print(f"OK: Formato decimal aplicado a {max_row-1} filas en columna Ind Ultimo")
    
    # Aplicar formato numerico a la columna Ind Pausa
    if 'Ind Pausa' in column_letters:
        decimal_format = '0.00'
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            ind_pausa_cell = f"{column_letters['Ind Pausa']}{row_num}"
            worksheet[ind_pausa_cell].number_format = decimal_format
        print(f"OK: Formato decimal aplicado a {max_row-1} filas en columna Ind Pausa")
    
    # Aplicar formato numerico a la columna Ind Ges Medio
    if 'Ind Ges Medio' in column_letters:
        decimal_format = '0.00'
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            ind_ges_medio_cell = f"{column_letters['Ind Ges Medio']}{row_num}"
            worksheet[ind_ges_medio_cell].number_format = decimal_format
        print(f"OK: Formato decimal aplicado a {max_row-1} filas en columna Ind Ges Medio")
    
    # Aplicar formato numerico a la columna Total Infracciones
    if 'Total Infracciones' in column_letters:
        integer_format = '0'  # Formato entero sin decimales
        max_row = max(len(df_operativo) + 1, 100)  # Minimo 100 filas para permitir datos manuales
        for row_num in range(2, max_row + 1):  # Empezar en fila 2 (despues del header)
            total_infracciones_cell = f"{column_letters['Total Infracciones']}{row_num}"
            worksheet[total_infracciones_cell].number_format = integer_format
        print(f"OK: Formato entero aplicado a {max_row-1} filas en columna Total Infracciones")
    # Ajustar ancho de columnas
    from openpyxl.utils import get_column_letter
    anchos_operativo = [10, 12, 10, 12, 8, 8, 8, 20, 10, 12, 10, 8, 12, 18, 15, 15, 10, 12, 12, 12, 15, 15, 15, 15, 12, 10, 8, 12, 12, 12, 12, 12, 15, 12, 18, 15]
    for i, ancho in enumerate(anchos_operativo, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = ancho
    
    # Aplicar formato de tabla
    aplicar_formato_tabla(worksheet, df_operativo, "TablaOperativo")

def crear_hoja_team(writer, datos_biometricos=None):
    """
    Crea la hoja "Team" con la tabla especificada
    Incluye datos de supervisores desde datos biomÃ©tricos
    """
    # Crear DataFrame con las columnas especificadas
    columnas_team = [
        "Codigo Aus", "Fecha", "Usuario", "Cedula", "Asistencia", "Asesores", 
        "Monitoreos", "% Calidad", "Infracciones", "% Operativo", "Cargo"
    ]
    
    # Crear registros para supervisores desde datos biomÃ©tricos
    registros_team = []
    
    if datos_biometricos is not None:
        print("ðŸ“Š Procesando datos de supervisores para hoja Team...")
        
        # Filtrar solo supervisores (no gerentes)
        supervisores_data = []
        for i, cargo in enumerate(datos_biometricos['cargos']):
            if str(cargo).upper() == 'SUPERVISOR':
                supervisores_data.append({
                    'codigo': datos_biometricos['codigos'][i],
                    'cedula': datos_biometricos['cedulas'][i],
                    'nombre': datos_biometricos['nombres'][i],
                    'fecha': datos_biometricos['fechas'][i],
                    'ingreso': datos_biometricos['ingresos'][i],
                    'salida': datos_biometricos['salidas'][i],
                    'cargo': cargo
                })
        
        if supervisores_data:
            print(f"ðŸ“‹ Encontrados {len(supervisores_data)} supervisores para Team")
            
            # Crear DataFrame temporal para agrupamiento
            df_temp = pd.DataFrame(supervisores_data)
            
            # Agrupar por fecha, nombre, cedula (en caso de mÃºltiples registros por dÃ­a)
            grouped = df_temp.groupby(['fecha', 'nombre', 'cedula']).agg({
                'ingreso': 'first',
                'salida': 'last',
                'cargo': 'first'
            }).reset_index()
            
            print(f"ðŸ“Š Registros agrupados para Team: {len(grouped)}")
            
            # Generar cÃ³digos de ausentismo para cada supervisor
            for _, row in grouped.iterrows():
                fecha = row['fecha']
                usuario = row['nombre']
                cedula = row['cedula']
                cargo = row['cargo']
                
                # Generar cÃ³digo de ausentismo: cedula + dÃ­a(00) + mes(00)
                if pd.notna(fecha) and pd.notna(cedula):
                    if isinstance(fecha, pd.Timestamp):
                        dia = fecha.strftime('%d')
                        mes = fecha.strftime('%m')
                        fecha_str = fecha.strftime('%d/%m/%Y')
                    else:
                        # Intentar parsear la fecha si es string
                        try:
                            fecha_obj = pd.to_datetime(fecha)
                            dia = fecha_obj.strftime('%d')
                            mes = fecha_obj.strftime('%m')
                            fecha_str = fecha_obj.strftime('%d/%m/%Y')
                        except:
                            dia = "00"
                            mes = "00"
                            fecha_str = str(fecha)
                    
                    codigo_aus = f"{cedula}{dia}{mes}"
                else:
                    codigo_aus = ""
                    fecha_str = ""
                
                registro = [
                    codigo_aus,    # Codigo Aus
                    fecha_str,     # Fecha
                    usuario,       # Usuario
                    str(cedula),   # Cedula
                    "",           # Asistencia (vacÃ­o para llenar manualmente)
                    "",           # Asesores (vacÃ­o para llenar manualmente)
                    "",           # Monitoreos (vacÃ­o para llenar manualmente)
                    "",           # % Calidad (vacÃ­o para llenar manualmente)
                    "",           # Infracciones (vacÃ­o para llenar manualmente)
                    "",           # % Operativo (vacÃ­o para llenar manualmente)
                    cargo         # Cargo (temporal, serÃ¡ reemplazado por fÃ³rmula VLOOKUP)
                ]
                registros_team.append(registro)
                print(f"  ðŸ“ Registro Team: {usuario} - Cedula: {cedula} - Cargo: {cargo}")
            
            print(f"âœ… Generados {len(registros_team)} registros para hoja Team")
    
    # Si no hay supervisores, crear una fila vacÃ­a
    if not registros_team:
        registros_team = [[""] * len(columnas_team)]
        print("ðŸ“‹ No se encontraron supervisores, creando hoja Team vacÃ­a")
    
    df_team = pd.DataFrame(registros_team, columns=columnas_team)
    
    # Escribir a Excel
    df_team.to_excel(writer, sheet_name="Team", index=False)
    
    # Obtener la hoja y aplicar formato basico
    worksheet = writer.sheets["Team"]
    
    # Aplicar formato de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    
    # Formatear encabezados
    for col in range(1, len(columnas_team) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 12  # Codigo Aus
    worksheet.column_dimensions['B'].width = 12  # Fecha
    worksheet.column_dimensions['C'].width = 20  # Usuario
    worksheet.column_dimensions['D'].width = 12  # Cedula
    worksheet.column_dimensions['E'].width = 12  # Asistencia
    worksheet.column_dimensions['F'].width = 10  # Asesores
    worksheet.column_dimensions['G'].width = 12  # Monitoreos
    worksheet.column_dimensions['H'].width = 12  # % Calidad
    worksheet.column_dimensions['I'].width = 15  # Infracciones
    worksheet.column_dimensions['J'].width = 15  # % Operativo
    worksheet.column_dimensions['K'].width = 10  # Cargo
    
    # Agregar fÃ³rmulas VLOOKUP para buscar cargo en hoja Planta
    print("ðŸ“Š Agregando fÃ³rmulas VLOOKUP para buscar cargo desde hoja Planta...")
    
    # Solo aplicar fÃ³rmulas si hay registros reales (no solo la fila vacÃ­a)
    tiene_datos_reales = len(registros_team) > 1 or (len(registros_team) == 1 and registros_team[0][2])  # Verificar si hay usuario en la primera fila
    
    if tiene_datos_reales:
        print(f"ðŸ“‹ Aplicando fÃ³rmulas VLOOKUP a {len(registros_team)} registros de Team...")
        
        for row_num in range(2, len(registros_team) + 2):  # Empezar desde fila 2 (despuÃ©s del header)
            # Verificar que la fila tenga datos (cÃ©dula no estÃ© vacÃ­a)
            cedula_valor = df_team.iloc[row_num-2, 3]  # Verificar columna Cedula (Ã­ndice 3)
            if cedula_valor and str(cedula_valor).strip():
                # Columna E es "Asistencia" (Ã­ndice 5)
                # VLOOKUP busca el cÃ³digo de ausentismo (columna A) en Asistencia Lideres y trae Novedad Ingreso (columna J)
                formula_asistencia = f'=IFERROR(VLOOKUP(A{row_num},\'Asistencia Lideres\'!A:J,10,FALSE),0)'
                worksheet[f'E{row_num}'].value = formula_asistencia
                
                # Columna F es "Asesores" (Ã­ndice 6)  
                # Contar registros POR DÃA ESPECÃFICO donde coincida fecha exacta Y mÃ­nimo 2 palabras del nombre
                # Operativo: C=Fecha (C2:C1000), Z=Team (Z2:Z1000)
                # IMPORTANTE: Solo cuenta para la fecha especÃ­fica de cada fila (B{row_num})
                
                # FÃ³rmula que cuenta SOLO para el dÃ­a especÃ­fico con validaciÃ³n de 2+ coincidencias
                # La condiciÃ³n (Operativo!C$2:C$1000=B{row_num}) asegura que solo cuenta registros del mismo dÃ­a
                formula_asesores = f'''=SUMPRODUCT((Operativo!C$2:C$1000=B{row_num})*(((ISNUMBER(SEARCH(LEFT(C{row_num},FIND(" ",C{row_num}&" ")-1),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),51,50)),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),101,50)),Operativo!Z$2:Z$1000))))>=2))'''
                
                print(f"  ðŸ“… FÃ³rmula asesores: cuenta solo para fecha B{row_num} (dÃ­a especÃ­fico)")
                worksheet[f'F{row_num}'].value = formula_asesores
                
                # Columna G es "Monitoreos" (Ã­ndice 7)
                # LÃ³gica compleja: buscar registros del dÃ­a con 2+ coincidencias de nombre, obtener sus cÃ³digos, verificar si aparecen en Calidad
                # Usar SUMPRODUCT para contar cÃ³digos de Operativo que tambiÃ©n aparecen en Calidad
                formula_monitoreos = f'''=SUMPRODUCT((Operativo!C$2:C$1000=B{row_num})*(((ISNUMBER(SEARCH(LEFT(C{row_num},FIND(" ",C{row_num}&" ")-1),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),51,50)),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),101,50)),Operativo!Z$2:Z$1000))))>=2)*(ISNUMBER(MATCH(Operativo!A$2:A$1000,Calidad!A:A,0))))'''
                worksheet[f'G{row_num}'].value = formula_monitoreos
                
                print(f"  ðŸ“Š FÃ³rmula monitoreos: cÃ³digos de Operativo que aparecen en Calidad para fecha B{row_num}")
                
                # Columna H es "% Calidad" (Ã­ndice 8)
                # PROCESO: igual que Monitoreos pero sacando promedio de las notas
                # 1. Encontrar cÃ³digos de Operativo del dÃ­a especÃ­fico que tengan 2+ coincidencias de nombre
                # 2. Verificar cuÃ¡les de esos cÃ³digos aparecen en Calidad 
                # 3. Sumar las Nota Total (columna G) de esos cÃ³digos en Calidad
                # 4. Dividir entre la cantidad de cÃ³digos para sacar el promedio
                # 5. Mostrar como porcentaje
                
                # Numerador: suma de las notas totales de los cÃ³digos que cumplen criterios
                # Los cÃ³digos de Operativo que coinciden con criterios Y aparecen en Calidad
                suma_notas = f'SUMPRODUCT((Operativo!C$2:C$1000=B{row_num})*(((ISNUMBER(SEARCH(LEFT(C{row_num},FIND(" ",C{row_num}&" ")-1),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),51,50)),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),101,50)),Operativo!Z$2:Z$1000))))>=2)*(SUMIF(Calidad!A:A,Operativo!A$2:A$1000,Calidad!G:G)))'
                
                # Denominador: cantidad de cÃ³digos que cumplen criterios (igual que Monitoreos)
                cantidad_codigos = f'SUMPRODUCT((Operativo!C$2:C$1000=B{row_num})*(((ISNUMBER(SEARCH(LEFT(C{row_num},FIND(" ",C{row_num}&" ")-1),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),51,50)),Operativo!Z$2:Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{row_num}," ",REPT(" ",50)),101,50)),Operativo!Z$2:Z$1000))))>=2)*(ISNUMBER(MATCH(Operativo!A$2:A$1000,Calidad!A:A,0))))'
                
                # FÃ³rmula completa: promedio como porcentaje
                formula_porcentaje_calidad = f'=IFERROR({suma_notas}/{cantidad_codigos},0)'
                worksheet[f'H{row_num}'].value = formula_porcentaje_calidad
                
                print(f"  ðŸ“ˆ FÃ³rmula % Calidad: promedio de Nota Total para cÃ³digos coincidentes fecha B{row_num}")
                
                # Columna I es "Infracciones" (Ã­ndice 9) - FÃ“RMULA CON BÃšSQUEDA MÃšLTIPLE
                # Suma las infracciones de la columna AI (Total Infracciones) para registros que coincidan en fecha y supervisor
                formula_infracciones = f'''=SUMPRODUCT((Operativo!$C$2:$C$1000=$B{row_num})*(((ISNUMBER(SEARCH(LEFT($C{row_num},FIND(" ",$C{row_num}&" ")-1),Operativo!$Z$2:$Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{row_num}," ",REPT(" ",50)),51,50)),Operativo!$Z$2:$Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{row_num}," ",REPT(" ",50)),101,50)),Operativo!$Z$2:$Z$1000))))>=2)*Operativo!$AI$2:$AI$1000)'''
                worksheet[f'I{row_num}'].value = formula_infracciones
                
                # Columna J es "% Operativo" (Ã­ndice 10) - FÃ“RMULA CON BÃšSQUEDA MÃšLTIPLE
                # Calcula el promedio de Total Operativo (columna AJ) para registros que coincidan en fecha y supervisor
                # Numerador: suma de valores de Total Operativo con bÃºsqueda mÃºltiple
                suma_operativo = f'SUMPRODUCT((Operativo!$C$2:$C$1000=$B{row_num})*(((ISNUMBER(SEARCH(LEFT($C{row_num},FIND(" ",$C{row_num}&" ")-1),Operativo!$Z$2:$Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{row_num}," ",REPT(" ",50)),51,50)),Operativo!$Z$2:$Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{row_num}," ",REPT(" ",50)),101,50)),Operativo!$Z$2:$Z$1000))))>=2)*Operativo!$AJ$2:$AJ$1000)'
                # Denominador: cantidad de registros que coinciden con bÃºsqueda mÃºltiple
                cantidad_registros = f'SUMPRODUCT((Operativo!$C$2:$C$1000=$B{row_num})*(((ISNUMBER(SEARCH(LEFT($C{row_num},FIND(" ",$C{row_num}&" ")-1),Operativo!$Z$2:$Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{row_num}," ",REPT(" ",50)),51,50)),Operativo!$Z$2:$Z$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{row_num}," ",REPT(" ",50)),101,50)),Operativo!$Z$2:$Z$1000))))>=2)*1)'
                # FÃ³rmula completa: promedio como porcentaje
                formula_porcentaje_operativo = f'=IFERROR({suma_operativo}/{cantidad_registros},0)'
                worksheet[f'J{row_num}'].value = formula_porcentaje_operativo
                
                print(f"  ðŸ“Š FÃ³rmula Infracciones: suma Total Infracciones de Operativo para fecha B{row_num}")
                print(f"  ðŸ“ˆ FÃ³rmula % Operativo: promedio de Total Operativo para registros coincidentes fecha B{row_num}")
                
                # Columna K es "Cargo" (Ã­ndice 11)
                # VLOOKUP busca la cÃ©dula (columna D) en la tabla de usuarios de Planta (B:C) 
                # Cambiando el rango para que sea mÃ¡s especÃ­fico: B=Cedula (buscar), C=Cargo (devolver)
                formula_cargo = f'=IFERROR(VLOOKUP(D{row_num},Planta!B:C,2,FALSE),"No encontrado")'
                worksheet[f'K{row_num}'].value = formula_cargo
                
                print(f"  âœ… FÃ³rmulas aplicadas en fila {row_num} - CÃ©dula: {cedula_valor}")
        
        # Agregar validaciÃ³n programÃ¡tica para filtrar registros con 0 asesores
        print("ðŸ“Š Implementando validaciÃ³n para filtrar registros con 0 asesores...")
        
        # Como las fÃ³rmulas se calculan en Excel, necesitamos filtrar durante la generaciÃ³n
        # Vamos a hacer una pre-validaciÃ³n aproximada basada en datos disponibles
        
        # Modificar la lÃ³gica para pre-filtrar registros que probablemente no tengan asesores
        # Esto se puede hacer verificando si existe el nombre en la fuente de datos original
        
        # El filtro automÃ¡tico ya estÃ¡ incluido en la tabla de Excel que se crea despuÃ©s
        # No necesitamos aplicar auto_filter manualmente
        max_row = len(registros_team) + 1
        
        # Aplicar formatos usando funciones optimizadas
        aplicar_formato_porcentaje(worksheet, ['E', 'H', 'J'], max_row - 1)
        
        # Aplicar formato de nÃºmero entero a toda la columna I (Infracciones)
        print("ðŸ“Š Aplicando formato de nÃºmero entero a columna Infracciones...")
        for row_num in range(2, max_row + 1):
            try:
                worksheet[f'I{row_num}'].number_format = "0"
            except:
                pass
        print(f"âœ… Formato de nÃºmero entero aplicado a I2:I{max_row}")
        
        # Agregar formato condicional para identificar visualmente filas con 0 asesores
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill, Font
        
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        red_font = Font(color="990000")
        
        # Crear regla para colorear filas donde columna F (Asesores) = 0
        rule_zero = CellIsRule(operator='equal', formula=['0'], fill=red_fill, font=red_font)
        worksheet.conditional_formatting.add(f'F2:F{max_row}', rule_zero)
        
        # TambiÃ©n agregar una regla para filas vacÃ­as en Asesores
        rule_empty = CellIsRule(operator='equal', formula=['""'], fill=red_fill, font=red_font)
        worksheet.conditional_formatting.add(f'F2:F{max_row}', rule_empty)
        
        print("âœ… ValidaciÃ³n implementada:")
        print("  - Formato condicional: filas con 0 asesores se colorean en rojo")
        print("  - La tabla de Excel incluye filtros automÃ¡ticos")
        print("  - Use los filtros de tabla para ocultar/eliminar filas con 0 asesores")
        
        print(f"âœ… FÃ³rmulas VLOOKUP agregadas para buscar cargo desde Planta")
    else:
        print("ðŸ“‹ No hay datos reales en Team, omitiendo fÃ³rmulas VLOOKUP")
    
    # Aplicar formato de tabla Excel a la hoja Team
    print("ðŸ“Š Aplicando formato de tabla Excel a la hoja Team...")
    
    # Crear tabla Excel con filtros automÃ¡ticos y estilo
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    max_row = len(registros_team) + 1
    tabla_team = Table(displayName="TablaTeam", ref=f"A1:{get_column_letter(len(columnas_team))}{max_row}")
    
    # Aplicar estilo de tabla
    style_team = TableStyleInfo(
        name=TABLE_STYLE_BLUE, 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=True
    )
    tabla_team.tableStyleInfo = style_team
    
    # Agregar la tabla a la hoja
    worksheet.add_table(tabla_team)
    
    print(f"âœ… Formato de tabla aplicado: TablaTeam (A1:{get_column_letter(len(columnas_team))}{max_row})")

def crear_hoja_gerente(writer, datos_biometricos=None):
    """
    Crea la hoja "Gerente" con la tabla especificada
    Incluye datos de gerentes desde datos biomÃ©tricos
    """
    # Crear DataFrame con las columnas especificadas
    columnas_gerente = [
        "Codigo Aus", "Fecha", "Usuario", "Cedula", "Asistencia", "Asesores", 
        "Monitoreos", "% Calidad", "Infracciones", "% Operativo", "Cargo"
    ]
    
    # Crear registros para gerentes desde datos biomÃ©tricos
    registros_gerente = []
    
    if datos_biometricos is not None:
        print("ðŸ“Š Procesando datos de gerentes para hoja Gerente...")
        
        # Filtrar solo gerentes
        gerentes_data = []
        for i, cargo in enumerate(datos_biometricos['cargos']):
            if str(cargo).upper() == 'GERENTE':
                gerentes_data.append({
                    'codigo': datos_biometricos['codigos'][i] if i < len(datos_biometricos.get('codigos', [])) else '',
                    'cedula': datos_biometricos['cedulas'][i] if i < len(datos_biometricos.get('cedulas', [])) else '',
                    'nombre': datos_biometricos['nombres'][i] if i < len(datos_biometricos.get('nombres', [])) else '',
                    'fecha': datos_biometricos['fechas'][i] if i < len(datos_biometricos.get('fechas', [])) else '',
                    'ingreso': datos_biometricos['ingresos'][i] if i < len(datos_biometricos.get('ingresos', [])) else '',
                    'salida': datos_biometricos['salidas'][i] if i < len(datos_biometricos.get('salidas', [])) else '',
                    'cargo': cargo
                })
        
        if gerentes_data:
            print(f"ðŸ“‹ Encontrados {len(gerentes_data)} gerentes para Gerente")
            
            # Crear DataFrame temporal para agrupamiento
            df_temp = pd.DataFrame(gerentes_data)
            
            # Agrupar por fecha, nombre, cedula (en caso de mÃºltiples registros por dÃ­a)
            grouped = df_temp.groupby(['fecha', 'nombre', 'cedula']).agg({
                'ingreso': 'first',
                'salida': 'last',
                'cargo': 'first'
            }).reset_index()
            
            print(f"ðŸ“Š Registros agrupados para Gerente: {len(grouped)}")
            
            # Generar cÃ³digos de ausentismo para cada gerente
            for _, row in grouped.iterrows():
                fecha = row['fecha']
                usuario = row['nombre']
                cedula = row['cedula']
                cargo = row['cargo']
                
                # Generar cÃ³digo de ausentismo: cedula + dÃ­a(00) + mes(00)
                if pd.notna(fecha) and pd.notna(cedula):
                    if isinstance(fecha, pd.Timestamp):
                        dia = fecha.strftime('%d')
                        mes = fecha.strftime('%m')
                        fecha_str = fecha.strftime('%d/%m/%Y')
                    else:
                        # Intentar parsear la fecha si es string
                        try:
                            fecha_obj = pd.to_datetime(fecha)
                            dia = fecha_obj.strftime('%d')
                            mes = fecha_obj.strftime('%m')
                            fecha_str = fecha_obj.strftime('%d/%m/%Y')
                        except:
                            dia = "00"
                            mes = "00"
                            fecha_str = str(fecha)
                    
                    codigo_aus = f"{cedula}{dia}{mes}"
                else:
                    codigo_aus = ""
                    fecha_str = ""
                
                registro = [
                    codigo_aus,    # Codigo Aus
                    fecha_str,     # Fecha
                    usuario,       # Usuario
                    str(cedula),   # Cedula
                    "",           # Asistencia (vacÃ­o para llenar manualmente)
                    "",           # Asesores (vacÃ­o para llenar manualmente)
                    "",           # Monitoreos (vacÃ­o para llenar manualmente)
                    "",           # % Calidad (vacÃ­o para llenar manualmente)
                    "",           # Infracciones (vacÃ­o para llenar manualmente)
                    "",           # % Operativo (vacÃ­o para llenar manualmente)
                    cargo         # Cargo (temporal, serÃ¡ reemplazado por fÃ³rmula VLOOKUP)
                ]
                registros_gerente.append(registro)
                print(f"  ðŸ“ Registro Gerente: {usuario} - Cedula: {cedula} - Cargo: {cargo}")
            
            print(f"âœ… Generados {len(registros_gerente)} registros para hoja Gerente")
    
    # Si no hay gerentes, crear una fila vacÃ­a
    if not registros_gerente:
        registros_gerente = [[""] * len(columnas_gerente)]
        print("ðŸ“‹ No se encontraron gerentes, creando hoja Gerente vacÃ­a")
    
    df_gerente = pd.DataFrame(registros_gerente, columns=columnas_gerente)
    
    # Escribir a Excel
    df_gerente.to_excel(writer, sheet_name="Gerente", index=False)
    
    # Obtener la hoja y aplicar formato basico
    worksheet = writer.sheets["Gerente"]
    
    # Aplicar formato de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    
    # Formatear encabezados
    for col in range(1, len(columnas_gerente) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 12  # Codigo Aus
    worksheet.column_dimensions['B'].width = 12  # Fecha
    worksheet.column_dimensions['C'].width = 20  # Usuario
    worksheet.column_dimensions['D'].width = 12  # Cedula
    worksheet.column_dimensions['E'].width = 12  # Asistencia
    worksheet.column_dimensions['F'].width = 10  # Asesores
    worksheet.column_dimensions['G'].width = 12  # Monitoreos
    worksheet.column_dimensions['H'].width = 12  # % Calidad
    worksheet.column_dimensions['I'].width = 15  # Infracciones
    worksheet.column_dimensions['J'].width = 15  # % Operativo
    worksheet.column_dimensions['K'].width = 10  # Cargo
    
    # Agregar fÃ³rmulas para gerentes (comenzando en fila 2 ya que fila 1 son encabezados)
    print("ðŸ“ Agregando fÃ³rmulas VLOOKUP para hoja Gerente...")
    
    num_filas = len(df_gerente) + 1  # +1 por el encabezado
    
    for fila in range(2, num_filas + 1):
        # E: Asistencia - VLOOKUP en 'Asistencia Lideres'!A:J, columna 10 (0% si vacÃ­o)
        formula_asistencia = f'=IFERROR(VLOOKUP(A{fila},\'Asistencia Lideres\'!A:J,10,FALSE),0)'
        worksheet[f'E{fila}'] = formula_asistencia
        
        # F: Asesores - SUMPRODUCT contando asesores en columna Gerencia (Y) de Operativo con bÃºsqueda mÃºltiple
        formula_asesores = f'''=SUMPRODUCT((Operativo!C$2:C$1000=B{fila})*(((ISNUMBER(SEARCH(LEFT(C{fila},FIND(" ",C{fila}&" ")-1),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),51,50)),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),101,50)),Operativo!Y$2:Y$1000))))>=2))'''
        worksheet[f'F{fila}'] = formula_asesores
        
        # G: Monitoreos - SUMPRODUCT con coincidencia en Gerencia y existencia en Calidad con bÃºsqueda mÃºltiple
        formula_monitoreos = f'''=SUMPRODUCT((Operativo!C$2:C$1000=B{fila})*(((ISNUMBER(SEARCH(LEFT(C{fila},FIND(" ",C{fila}&" ")-1),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),51,50)),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),101,50)),Operativo!Y$2:Y$1000))))>=2)*(ISNUMBER(MATCH(Operativo!A$2:A$1000,Calidad!A:A,0))))'''
        worksheet[f'G{fila}'] = formula_monitoreos
        
        # H: % Calidad - Porcentaje de calidad con bÃºsqueda mÃºltiple de palabras (0% si vacÃ­o)
        formula_porcentaje_calidad = f'''=IFERROR(SUMPRODUCT((Operativo!C$2:C$1000=B{fila})*(((ISNUMBER(SEARCH(LEFT(C{fila},FIND(" ",C{fila}&" ")-1),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),51,50)),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),101,50)),Operativo!Y$2:Y$1000))))>=2)*(SUMIF(Calidad!A:A,Operativo!A$2:A$1000,Calidad!G:G)))/SUMPRODUCT((Operativo!C$2:C$1000=B{fila})*(((ISNUMBER(SEARCH(LEFT(C{fila},FIND(" ",C{fila}&" ")-1),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),51,50)),Operativo!Y$2:Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE(C{fila}," ",REPT(" ",50)),101,50)),Operativo!Y$2:Y$1000))))>=2)*(ISNUMBER(MATCH(Operativo!A$2:A$1000,Calidad!A:A,0)))),0)'''
        worksheet[f'H{fila}'] = formula_porcentaje_calidad
        
        # I: Infracciones - SUMPRODUCT sumando infracciones en columna Gerencia (Y) con bÃºsqueda mÃºltiple de palabras
        formula_infracciones = f'''=SUMPRODUCT((Operativo!$C$2:$C$1000=$B{fila})*(((ISNUMBER(SEARCH(LEFT($C{fila},FIND(" ",$C{fila}&" ")-1),Operativo!$Y$2:$Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{fila}," ",REPT(" ",50)),51,50)),Operativo!$Y$2:$Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{fila}," ",REPT(" ",50)),101,50)),Operativo!$Y$2:$Y$1000))))>=2)*Operativo!$AI$2:$AI$1000)'''
        worksheet[f'I{fila}'] = formula_infracciones
        
        # J: % Operativo - Promedio de Total Operativo en columna Gerencia (Y) con bÃºsqueda mÃºltiple de palabras
        # Numerador: suma de valores de Total Operativo (AJ) con bÃºsqueda mÃºltiple
        suma_operativo = f'SUMPRODUCT((Operativo!$C$2:$C$1000=$B{fila})*(((ISNUMBER(SEARCH(LEFT($C{fila},FIND(" ",$C{fila}&" ")-1),Operativo!$Y$2:$Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{fila}," ",REPT(" ",50)),51,50)),Operativo!$Y$2:$Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{fila}," ",REPT(" ",50)),101,50)),Operativo!$Y$2:$Y$1000))))>=2)*Operativo!$AJ$2:$AJ$1000)'
        # Denominador: cantidad de registros que coinciden con bÃºsqueda mÃºltiple
        cantidad_registros = f'SUMPRODUCT((Operativo!$C$2:$C$1000=$B{fila})*(((ISNUMBER(SEARCH(LEFT($C{fila},FIND(" ",$C{fila}&" ")-1),Operativo!$Y$2:$Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{fila}," ",REPT(" ",50)),51,50)),Operativo!$Y$2:$Y$1000)))+(ISNUMBER(SEARCH(TRIM(MID(SUBSTITUTE($C{fila}," ",REPT(" ",50)),101,50)),Operativo!$Y$2:$Y$1000))))>=2)*1)'
        # FÃ³rmula completa: promedio como porcentaje
        formula_porcentaje_operativo = f'=IFERROR({suma_operativo}/{cantidad_registros},0)'
        worksheet[f'J{fila}'] = formula_porcentaje_operativo
        
        print(f"  ðŸ“‹ Fila {fila}: FÃ³rmulas Asistencia, Asesores, Monitoreos, % Calidad, Infracciones y % Operativo agregadas")
    
    print("âœ… FÃ³rmulas VLOOKUP agregadas correctamente para hoja Gerente")
    
    # Aplicar formato de porcentaje usando funciÃ³n optimizada
    aplicar_formato_porcentaje(worksheet, ['E', 'H', 'J'], num_filas - 1)
    
    # Aplicar formato de tabla
    aplicar_formato_tabla(worksheet, df_gerente, "TablaGerente")

def crear_hoja_consolidado(writer, df_operativo=None):
    """
    Crea la hoja "Consolidado" con la tabla especificada
    Extrae cÃ³digos de la hoja Operativo y usa VLOOKUP para buscar informaciÃ³n
    """
    # Crear DataFrame con las columnas especificadas
    columnas_consolidado = [
        "Codigo_Asis", "CODIGO", "Tipo Jornada", "Fecha", "Cedula", "ID", "Nombre", "Sede", "Ubicacion",
        "Asistencia", "Mora", "Monitoreos", "Nota Calidad", "Ejecucion", "# Infracciones", 
        "% Operativo", "Team", "Gerente"
    ]
    
    # Crear filas solo para los registros que realmente existen en Operativo
    print("ðŸ“Š Creando hoja Consolidado solo con filas que tienen datos...")
    
    registros_consolidado = []
    
    # Determinar cuÃ¡ntas filas realmente tiene Operativo
    if df_operativo is not None and not df_operativo.empty:
        num_filas_operativo = len(df_operativo)
        print(f"ðŸ“‹ Operativo tiene {num_filas_operativo} registros reales")
        
        # Crear solo las filas que corresponden a datos reales
        for i in range(2, num_filas_operativo + 2):  # +2 porque empezamos en fila 2 y len() da cantidad
            registro = [
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:D,4,FALSE)&TEXT(VLOOKUP(B{i},Operativo!A:C,3,FALSE),"DDMM"),""))',  # Codigo_Asis - cedula + dÃ­a + mes usando CODIGO
                f'=Operativo!A{i}',  # CODIGO - cÃ³digo directo
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:B,2,FALSE),""))',  # Tipo Jornada - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:C,3,FALSE),""))',  # Fecha - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:D,4,FALSE),""))',  # Cedula - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:E,5,FALSE),""))',  # ID - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:H,8,FALSE),""))',  # Nombre - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:I,9,FALSE),""))',  # Sede - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:J,10,FALSE),""))', # Ubicacion - VLOOKUP con CODIGO
                f'=IF(A{i}="",0,IFERROR(VLOOKUP(A{i},Ausentismo!A:P,16,FALSE),0))',  # % Asistencia - VLOOKUP Drive con Codigo_Asis, 0 si vacÃ­o
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:L,12,FALSE),""))', # Mora - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Calidad!A:H,8,FALSE),0))',  # Monitoreos - VLOOKUP Total Monitoreos (columna H), 0 si no encuentra
                f'=IF(B{i}="",0,IFERROR(VLOOKUP(B{i},Calidad!A:G,7,FALSE),0))',   # Nota Calidad - VLOOKUP Nota Total (columna G) como porcentaje, 0 si vacÃ­o
                f'=IF(B{i}="",0,IFERROR(VLOOKUP(B{i},Operativo!A:AB,28,FALSE),0))', # Ejecucion - VLOOKUP en columna AB (EjecuciÃ³n) de Operativo, 0% si vacÃ­o
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:AI,35,FALSE),""))', # # Infracciones - mantener lÃ³gica especial
                f'=IF(B{i}="",0,IFERROR(VLOOKUP(B{i},Operativo!A:AJ,36,FALSE),0))', # % Operativo - VLOOKUP con 0 si vacÃ­o o no encuentra
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:Z,26,FALSE),""))',  # Team - VLOOKUP con CODIGO
                f'=IF(B{i}="","",IFERROR(VLOOKUP(B{i},Operativo!A:Y,25,FALSE),""))'   # Gerente - VLOOKUP con CODIGO
            ]
            registros_consolidado.append(registro)
        
        print(f"âœ… Generados exactamente {len(registros_consolidado)} registros (igual que Operativo)")
    else:
        # Si no hay datos de Operativo, crear solo una fila vacÃ­a
        registro_vacio = [""] * len(columnas_consolidado)
        registros_consolidado.append(registro_vacio)
        print("ðŸ“‹ No hay datos de Operativo, creando hoja Consolidado vacÃ­a")
    
    # Crear DataFrame con los registros exactos
    df_consolidado = pd.DataFrame(registros_consolidado, columns=columnas_consolidado)
    
    # Escribir a Excel
    df_consolidado.to_excel(writer, sheet_name="Consolidado", index=False)
    
    # Obtener la hoja y aplicar formato basico
    worksheet = writer.sheets["Consolidado"]
    
    # Aplicar formato de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    
    # Formatear encabezados
    for col in range(1, len(columnas_consolidado) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 12  # Codigo_Asis
    worksheet.column_dimensions['B'].width = 10  # CODIGO
    worksheet.column_dimensions['C'].width = 15  # Tipo Jornada
    worksheet.column_dimensions['D'].width = 12  # Fecha
    worksheet.column_dimensions['E'].width = 12  # Cedula
    worksheet.column_dimensions['F'].width = 8   # ID
    worksheet.column_dimensions['G'].width = 20  # Nombre
    worksheet.column_dimensions['H'].width = 10  # Sede
    worksheet.column_dimensions['I'].width = 12  # Ubicacion
    worksheet.column_dimensions['J'].width = 12  # Asistencia
    worksheet.column_dimensions['K'].width = 8   # Mora
    worksheet.column_dimensions['L'].width = 12  # Monitoreos
    worksheet.column_dimensions['M'].width = 15  # Nota Calidad
    worksheet.column_dimensions['N'].width = 12  # Ejecucion
    worksheet.column_dimensions['O'].width = 15  # # Infracciones
    worksheet.column_dimensions['P'].width = 15  # % Operativo
    worksheet.column_dimensions['Q'].width = 10  # Team
    worksheet.column_dimensions['R'].width = 12  # Gerente
    
    # Aplicar formato de porcentaje usando funciÃ³n optimizada
    aplicar_formato_porcentaje(worksheet, ['J', 'M', 'N', 'P'], len(registros_consolidado))
    
    # Aplicar formato de tabla
    aplicar_formato_tabla(worksheet, df_consolidado, "TablaConsolidado")
