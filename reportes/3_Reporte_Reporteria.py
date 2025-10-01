"""
Módulo para el procesamiento del 3_Reporte_Reporteria
Sistema optimizado y limpio para generación de reportes de cobranza.
"""
import io
import json
import os
import re
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from flask import request, send_file, jsonify
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from utils.file_utils import allowed_file

# ============================================================================
# CONSTANTES DE CONFIGURACIÓN
# ============================================================================
BASE_ASESORES_FILE = 'base_asesores.json'
EXCEL_EPOCH_THRESHOLD = 40000
DEFAULT_SHEET_NAME = 'Información'
TABLE_STYLE = "TableStyleMedium12"
MAX_COLUMN_WIDTH = 50
FORMATO_HORA = 'h:mm:ss AM/PM'

# Columnas requeridas para validación
REQUIRED_ASESOR_COLUMNS = [
    'Fecha Ingreso', 'Fecha', 'Cedula', 'ID', 'EXT', 
    'VOIP', 'Nombre', 'Sede', 'Ubicación'
]

# Mapeo de columnas de archivos de entrada
ADMIN_COLUMNS = {
    'ID': 'ID',
    'NOMBRE': 'Nombre',
    'LOGUEO': 'Logueo',
    'CARTERA': 'CARTERA',
    'ASIGNACION': 'ASIGNACION',
    'TOCADAS_11AM': 'TOCADAS 11 AM',
    'ASIGNADO': 'ASIGNADO',
    'RECUPERADO': 'RECUPERADO',
    'PAGOS': 'PAGOS',
    'PORCENTAJE_RECUPERADO': '% RECUPERADO',
    'PORCENTAJE_CUENTAS': '% CUENTAS',
    'TOQUES': 'TOQUES',
    'ULTIMO_TOQUE': 'ULTIMO TOQUE',
    'GERENTE': 'Gerente',
    'TEAM_LEADER': 'Team Leader',
    'UBICACION': 'Ubicación'
}

LLAMADAS_COLUMNS = {
    'EXTENSION': 'Número Extensión',
    'TOTAL_LLAMADAS': 'Total Llamadas'
}

VOIP_COLUMNS = {
    'VOIP_NUMBER': 'Extensión',
    'TOTAL_VOIP': 'Total'
}

# Correcciones específicas para valores de mora
MORA_CORRECTIONS = {
    'M0-1 PP': 'M0-1-PP',
    'M1-1A FRS': 'M1-1A-FRS',
    'M1-1A BT': 'M1-1A-BT',
    'M1-1A PX': 'M1-1A-PX'
}

# Nombres de meses en español
MESES_ESPANOL = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

# ============================================================================
# FUNCIONES UTILITARIAS
# ============================================================================


def generar_nombre_archivo_reporteria(hojas_procesadas: int, admin_sheets: Dict[str, Any]) -> str:
    """
    Genera nombre dinámico de archivo basado en las fechas de las hojas (formato igual a reportes 1 y 2)
    
    Args:
        hojas_procesadas: Número de hojas procesadas
        admin_sheets: Diccionario con las hojas administrativas
        
    Returns:
        str: Nombre del archivo con formato legible (sin extensión)
        
    Examples:
        - Un día: "3_Reporte_Reporteria (9 Agosto 2025)"
        - Mismo mes: "3_Reporte_Reporteria (3-8 Septiembre 2025)" 
        - Diferente mes: "3_Reporte_Reporteria (30 Agosto - 5 Septiembre 2025)"
    """
    try:
        fecha_partes = []

        
        for i, sheet_name in enumerate(list(admin_sheets.keys())[:hojas_procesadas]):

            fecha_obj = None
            # Intentar múltiples formatos de fecha
            formatos = ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%y', '%d/%m/%y']
            
            for formato in formatos:
                try:
                    fecha_obj = datetime.strptime(sheet_name, formato)
                    fecha_partes.append(fecha_obj)

                    break
                except ValueError:
                    continue
            
            # Si no funciona con datetime, intentar extraer números que parezcan fecha
            if fecha_obj is None:
                import re
                # Buscar patrones como dd-mm-yyyy, dd/mm/yyyy, etc.
                patron_fecha = re.search(r'(\d{1,2})[-/](\d{1,2})[-/](\d{4})', sheet_name)
                if patron_fecha:
                    try:
                        dia, mes, año = patron_fecha.groups()
                        fecha_obj = datetime(int(año), int(mes), int(dia))
                        fecha_partes.append(fecha_obj)

                    except ValueError:
                        pass

        
        if not fecha_partes:
            # Si no hay fechas válidas, usar timestamp
            timestamp = datetime.now().strftime("%Y-%m-%dT%H%M%S.%f")[:-3]

            return f"3_Reporte_Reporteria - {timestamp}"
        
        # Ordenar las fechas y obtener rango
        fecha_partes = sorted(fecha_partes)
        fecha_inicio = fecha_partes[0]
        fecha_fin = fecha_partes[-1]
        
        # Un solo día
        if len(fecha_partes) == 1 or fecha_inicio == fecha_fin:
            dia = fecha_inicio.day
            mes = MESES_ESPANOL[fecha_inicio.month]
            año = fecha_inicio.year
            nombre_final = f"3_Reporte_Reporteria ({dia} {mes} {año})"

            return nombre_final
        
        # Múltiples días - analizar rangos
        dia_inicio, dia_fin = fecha_inicio.day, fecha_fin.day
        año_inicio, año_fin = fecha_inicio.year, fecha_fin.year
        mes_inicio_num, mes_fin_num = fecha_inicio.month, fecha_fin.month
        
        mes_inicio = MESES_ESPANOL[mes_inicio_num]
        mes_fin = MESES_ESPANOL[mes_fin_num]
        
        # Mismo mes y año
        if mes_inicio_num == mes_fin_num and año_inicio == año_fin:
            nombre_final = f"3_Reporte_Reporteria ({dia_inicio}-{dia_fin} {mes_inicio} {año_inicio})"

            return nombre_final
        
        # Mismo año, diferente mes  
        elif año_inicio == año_fin:
            nombre_final = f"3_Reporte_Reporteria ({dia_inicio} {mes_inicio} - {dia_fin} {mes_fin} {año_inicio})"

            return nombre_final
        
        # Diferente año
        else:
            nombre_final = f"3_Reporte_Reporteria ({dia_inicio} {mes_inicio} {año_inicio} - {dia_fin} {mes_fin} {año_fin})"

            return nombre_final
    
    except Exception as e:

        timestamp = datetime.now().strftime("%Y-%m-%dT%H%M%S.%f")[:-3]
        return f"3_Reporte_Reporteria - {timestamp}"

def safe_get_value(row: Optional[Dict[str, Any]], key: str, default: Any = '') -> Any:
    """Obtiene valor de manera segura de un diccionario."""
    return row.get(key, default) if row else default

def safe_float_conversion(value: Any, default: float = 0) -> float:
    """Convierte valor a float de manera segura."""
    try:
        return float(value) if value is not None and str(value).strip() else default
    except (ValueError, TypeError):
        return default

def limpiar_fecha(fecha_str: str) -> str:
    """Limpia y normaliza fechas en formato dd/mm/yyyy."""
    if not fecha_str or pd.isna(fecha_str):
        return ''
    
    fecha_str = str(fecha_str).strip()
    
    # Si es un timestamp de pandas, convertir directamente
    if isinstance(fecha_str, pd.Timestamp):
        return fecha_str.strftime('%d/%m/%Y')
    
    # Extraer fecha con regex si contiene ruido
    fecha_clean = re.sub(r'[^\d/\-.]', ' ', fecha_str)
    fecha_clean = re.sub(r'\s+', ' ', fecha_clean).strip()
    
    if not fecha_clean:
        return ''
    
    # Intentar parsear diferentes formatos
    formatos = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%d/%m/%y']
    
    for formato in formatos:
        try:
            fecha_obj = datetime.strptime(fecha_clean, formato)
            return fecha_obj.strftime('%d/%m/%Y')
        except ValueError:
            continue
    
    return fecha_str

def convertir_fecha_formato(fecha_str: str) -> str:
    """Convierte fechas a formato dd-mm-yyyy para nombres de hojas."""
    if not fecha_str or pd.isna(fecha_str):
        return ''
    
    fecha_str = str(fecha_str).strip()
    
    # Patrones de fechas con nombres de meses
    patron_mes = r'(\d{1,2})\s*de\s*(\w+)\s*de\s*(\d{4})'
    
    meses_map = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
    }
    
    try:
        match = re.search(patron_mes, fecha_str.lower())
        if match:
            dia = match.group(1).zfill(2)
            mes_nombre = match.group(2)
            año = match.group(3)
            if mes_nombre in meses_map:
                mes = meses_map[mes_nombre]
                return f"{dia}-{mes}-{año}"
    except Exception:
        pass
    
    # Formatos numéricos comunes
    try:
        if re.match(r'\d{4}-\d{2}-\d{2}', fecha_str):
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
            return fecha_obj.strftime('%d-%m-%Y')
        
        if re.match(r'\d{2}-\d{2}-\d{4}', fecha_str):
            fecha_obj = datetime.strptime(fecha_str, '%d-%m-%Y')
            return fecha_obj.strftime('%d-%m-%Y')
        
        # Manejo de fechas numéricas de Excel
        try:
            fecha_num = float(fecha_str)
            if fecha_num > EXCEL_EPOCH_THRESHOLD:
                fecha_excel = pd.to_datetime('1900-01-01') + pd.Timedelta(days=fecha_num-2)
                return fecha_excel.strftime('%d-%m-%Y')
        except (ValueError, TypeError):
            pass
        
    except Exception:
        pass
    
    return fecha_str

def formatear_valor_monetario(valor: Any) -> str:
    """Formatea valores monetarios con símbolo peso y separadores de miles."""
    valor_num = safe_float_conversion(valor)
    
    if valor_num == 0:
        return '$0'
    
    # Formatear con puntos como separador de miles
    valor_str = f"{valor_num:.0f}"
    if len(valor_str) <= 3:
        return f"${valor_str}"
    
    valor_formateado = ""
    for i, digit in enumerate(reversed(valor_str)):
        if i > 0 and i % 3 == 0:
            valor_formateado = "." + valor_formateado
        valor_formateado = digit + valor_formateado
    
    return f"${valor_formateado}"

def convertir_porcentaje_a_decimal(valor_porcentaje: Any) -> float:
    """Convierte porcentaje a decimal para cálculos."""
    try:
        if pd.isna(valor_porcentaje) or valor_porcentaje in ['', None]:
            return 0.0
        
        valor_str = str(valor_porcentaje).strip()
        
        # Manejar porcentajes con símbolo %
        if '%' in valor_str:
            numero_str = valor_str.replace('%', '').strip()
            numero = float(numero_str)
            return numero / 100.0
        else:
            numero = float(valor_str)
            # Si es mayor que 1, asumir que está en formato porcentaje
            return numero / 100.0 if numero > 1 else numero
    except (ValueError, TypeError):
        return 0.0

def formatear_porcentaje_con_coma(valor: float, decimales: int = 1) -> str:
    """Formatea un valor decimal como porcentaje usando coma como separador decimal."""
    try:
        if pd.isna(valor) or valor == '':
            return '0,0%'
        
        # Convertir a porcentaje y formatear con coma
        porcentaje = valor * 100
        valor_formateado = f"{porcentaje:.{decimales}f}%"
        # Reemplazar punto por coma
        return valor_formateado.replace('.', ',')
    except (ValueError, TypeError):
        return '0,0%'

def normalizar_valor_mora(valor_mora: Any) -> Any:
    """Normaliza valores de mora aplicando correcciones específicas."""
    if pd.isna(valor_mora) or valor_mora == '':
        return valor_mora
    
    valor_str = str(valor_mora).strip()
    return MORA_CORRECTIONS.get(valor_str, valor_mora)

def convertir_hora_formato(hora_str: Any) -> str:
    """Convierte horas a formato estándar HH:MM AM/PM."""
    if pd.isna(hora_str) or not hora_str:
        return ''
    
    hora_str = str(hora_str).strip()
    
    try:
        # Si ya está en formato correcto
        if re.match(r'\d{1,2}:\d{2} (AM|PM)', hora_str):
            return hora_str
        
        # Si es formato 24 horas
        if re.match(r'\d{1,2}:\d{2}(:\d{2})?$', hora_str):
            try:
                if ':' in hora_str and len(hora_str.split(':')) >= 2:
                    partes = hora_str.split(':')
                    hora = int(partes[0])
                    minuto = int(partes[1])
                    
                    periodo = 'AM' if hora < 12 else 'PM'
                    hora_12 = hora if hora <= 12 else hora - 12
                    if hora_12 == 0:
                        hora_12 = 12
                    
                    return f"{hora_12}:{minuto:02d} {periodo}"
            except (ValueError, IndexError):
                pass
    except Exception:
        pass
    
    return hora_str

# ============================================================================
# FUNCIONES DE DATOS
# ============================================================================
def load_base_asesores() -> List[Dict[str, Any]]:
    """Carga la base de datos de asesores desde archivo JSON."""
    try:
        if os.path.exists(BASE_ASESORES_FILE):
            with open(BASE_ASESORES_FILE, 'r', encoding='utf-8') as f:
                asesores = json.load(f)

                return asesores
        else:

            return []
    except Exception as e:

        return []

def create_data_map(data: Optional[pd.DataFrame], key_column: str) -> Dict[str, Dict[str, Any]]:
    """Crea mapeo de datos usando columna clave especificada."""
    if data is None or data.empty:
        return {}
    
    if key_column not in data.columns:

        return {}
    
    try:
        # Crear mapeo convirtiendo claves a string
        data_map = {}
        for _, row in data.iterrows():
            key_value = str(row[key_column]).strip()
            if key_value and key_value != 'nan':
                data_map[key_value] = row.to_dict()
        

        return data_map
    
    except Exception as e:

        return {}

# ============================================================================
# FUNCIONES DE FORMATEO EXCEL
# ============================================================================
def apply_excel_formatting(worksheet: Any, dataframe: pd.DataFrame, sheet_name: str) -> None:
    """Aplica formato profesional a hojas Excel."""
    try:
        # Ajustar anchos de columnas
        for i, column in enumerate(dataframe.columns, 1):
            col_letter = get_column_letter(i)
            max_length = max(
                len(str(column)),
                max(len(str(val)) for val in dataframe.iloc[:, i-1].astype(str)) if not dataframe.empty else 0
            )
            adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Crear tabla estructurada
        if not dataframe.empty:
            clean_name = re.sub(r'[^a-zA-Z0-9_]', '_', sheet_name)
            table_name = f"Tabla_{clean_name}"
            
            end_col = get_column_letter(len(dataframe.columns))
            end_row = len(dataframe) + 1
            table_ref = f"A1:{end_col}{end_row}"
            
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name=TABLE_STYLE,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            worksheet.add_table(table)
        
        # Aplicar formato a columnas de tiempo
        time_columns = [col for col in dataframe.columns if 'toque' in col.lower()]
        for col_name in time_columns:
            if col_name in dataframe.columns:
                col_idx = dataframe.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                
                for row in range(2, len(dataframe) + 2):
                    time_cell = f"{col_letter}{row}"
                    worksheet[time_cell].alignment = Alignment(horizontal='center')
                    worksheet[time_cell].number_format = FORMATO_HORA
    except Exception as e:
        pass


# ============================================================================
# FUNCIÓN PRINCIPAL DE PROCESAMIENTO
# ============================================================================
def procesar_reporteria_cobranza():
    """Procesa archivos de reportería y genera el 3_Reporte_Reporteria."""
    # Validar archivos de entrada
    admin_file = request.files.get('reporteAdminFile')
    llamadas_file = request.files.get('reporteLlamadasFile')
    voip_file = request.files.get('reporteVoipFile')

    if not admin_file or not llamadas_file:
        return "Error: Debes subir al menos los archivos de reporte Admin y Llamadas.", 400
    
    if not (allowed_file(admin_file.filename, {'xlsx'}) and 
            allowed_file(llamadas_file.filename, {'xlsx'})):
        return "Error: Formato de archivo no permitido (solo .xlsx).", 400
    
    if voip_file and not allowed_file(voip_file.filename, {'xlsx'}):
        return "Error: Formato de archivo VOIP no permitido (solo .xlsx).", 400
    
    try:
        # Cargar datos
        admin_sheets = pd.read_excel(admin_file, sheet_name=None, dtype=str)
        llamadas_sheets = pd.read_excel(llamadas_file, sheet_name=None, dtype=str)
        
        voip_sheets = {}
        if voip_file:
            voip_sheets = pd.read_excel(voip_file, sheet_name=None, dtype=str)

        

        
        # Cargar base de asesores
        base_asesores = load_base_asesores()
        if not base_asesores:
            return "Error: No hay asesores en la base de datos.", 400
        
        # Debug: mostrar asesores con VOIP
        asesores_con_voip = [asesor for asesor in base_asesores if asesor.get('VOIP')]

        if asesores_con_voip:

            voip_numbers = [asesor.get('VOIP') for asesor in asesores_con_voip]
            voip_numbers_sorted = sorted(voip_numbers)[:10]

        
        output = io.BytesIO()
        hojas_procesadas = 0
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name in admin_sheets.keys():
                if sheet_name in llamadas_sheets:

                    
                    # Obtener datos de cada archivo
                    admin_data = admin_sheets[sheet_name]
                    llamadas_data = llamadas_sheets[sheet_name]
                    voip_data = voip_sheets.get(sheet_name) if voip_sheets else None
                    
                    if admin_data.empty:

                        continue
                    
                    # Crear mapas de datos
                    admin_map = create_data_map(admin_data, ADMIN_COLUMNS['ID'])
                    
                    if admin_map:
                        sample_id = list(admin_map.keys())[0]
                        sample_admin = admin_map[sample_id]
                        print(f"   🗂️  Mapa de admin creado: {len(admin_map)} registros")
                        print(f"   📋 Ejemplo ID {sample_id} - columnas disponibles: {list(sample_admin.keys())}")
                        print(f"   💼 Valores ejemplo: CARTERA={sample_admin.get('CARTERA')}, ASIGNACION={sample_admin.get('ASIGNACION')}, RECUPERADO={sample_admin.get('RECUPERADO')}")
                    else:
                        print(f"   ❌ No se pudo crear mapa de admin para hoja '{sheet_name}'")
                    
                    # Procesar llamadas
                    llamadas_map = {}
                    if llamadas_data is not None and not llamadas_data.empty:
                        llamadas_map = create_data_map(llamadas_data, LLAMADAS_COLUMNS['EXTENSION'])
                        print(f"✅ Datos de llamadas cargados para hoja '{sheet_name}': {len(llamadas_data)} registros")
                        
                        if llamadas_map:
                            print(f"   📞 Mapa de llamadas creado con {len(llamadas_map)} extensiones")
                            sample_extensions = list(llamadas_map.keys())[:5]
                            print(f"   📝 Primeras 5 extensiones mapeadas: {sample_extensions}")
                            for ext in sample_extensions:
                                total_llamadas = safe_get_value(llamadas_map[ext], LLAMADAS_COLUMNS['TOTAL_LLAMADAS'], 0)
                                print(f"   🔢 Ext {ext}: {total_llamadas} llamadas")
                        else:
                            print(f"   ❌ No se pudo crear mapa de llamadas para hoja '{sheet_name}'")
                    else:
                        print(f"❌ No hay datos de llamadas para hoja '{sheet_name}' - todas las llamadas serán 0")
                    
                    # Procesar VOIP
                    voip_map = {}
                    if voip_data is not None and not voip_data.empty:
                        print(f"   📋 Columnas disponibles en archivo VOIP: {list(voip_data.columns)}")
                        
                        voip_map = create_data_map(voip_data, VOIP_COLUMNS['VOIP_NUMBER'])
                        print(f"✅ Datos de VOIP cargados para hoja '{sheet_name}': {len(voip_data)} registros")
                        
                        if voip_map:
                            print(f"   📱 Mapa de VOIP creado con {len(voip_map)} registros usando columna '{VOIP_COLUMNS['VOIP_NUMBER']}'")
                            
                            all_voip_keys = sorted(list(voip_map.keys()))
                            print(f"   📝 TODOS los números VOIP disponibles ({len(all_voip_keys)}): {all_voip_keys}")
                            
                            sample_voips = list(voip_map.keys())[:5]
                            for voip_key in sample_voips:
                                voip_value = safe_get_value(voip_map[voip_key], VOIP_COLUMNS['TOTAL_VOIP'], 0)
                                print(f"   🔢 Clave '{voip_key}': {voip_value} llamadas")
                        else:
                            print(f"   ⚠️  No se pudo crear mapa de VOIP para hoja '{sheet_name}'")
                            print(f"   🔍 Verificar que la columna '{VOIP_COLUMNS['VOIP_NUMBER']}' existe en: {list(voip_data.columns)}")
                    else:
                        print(f"❌ No hay datos de VOIP para hoja '{sheet_name}' - todas las llamadas VOIP serán 0")
                    
                    # Generar filas del reporte
                    report_rows = generate_report_rows(base_asesores, admin_map, llamadas_map, voip_map, sheet_name)
                    
                    if report_rows:
                        df_report = pd.DataFrame(report_rows)
                        sheet_name_clean = convertir_fecha_formato(sheet_name)
                        df_report.to_excel(writer, sheet_name=sheet_name_clean, index=False)
                        apply_excel_formatting(writer.sheets[sheet_name_clean], df_report, sheet_name_clean)
                        hojas_procesadas += 1
                        print(f"✅ Tabla creada: Tabla_{sheet_name_clean}")
                    else:
                        print(f"❌ No se generaron filas para la hoja '{sheet_name}'")
                else:
                    print(f"⚠️ Hoja '{sheet_name}' no encontrada en archivo de llamadas")
            
            if hojas_procesadas == 0:
                create_default_sheet(writer, admin_sheets)
        
        # Preparar respuesta
        output.seek(0)
        
        # Debug: mostrar hojas procesadas para generar nombre
        print(f"📁 Generando nombre archivo:")
        print(f"   📊 Hojas procesadas: {hojas_procesadas}")
        print(f"   📋 Nombres de hojas disponibles: {list(admin_sheets.keys())}")
        print(f"   🎯 Primeras {hojas_procesadas} hojas: {list(admin_sheets.keys())[:hojas_procesadas]}")
        
        filename = generar_nombre_archivo_reporteria(hojas_procesadas, admin_sheets)
        print(f"   📝 Nombre generado: {filename}")
        
        # Guardar archivo temporal
        temp_dir = 'temp_files'
        os.makedirs(temp_dir, exist_ok=True)
        temp_file_path = os.path.join(temp_dir, f"{filename}.xlsx")
        
        with open(temp_file_path, 'wb') as f:
            f.write(output.getvalue())
        
        # Enviar respuesta JSON con información del archivo en lugar de descargar directamente
        return jsonify({
            'success': True,
            'message': 'Reporte 3 procesado correctamente',
            'temp_file': f"{filename}.xlsx",
            'download_name': f"{filename}.xlsx",
            'hojas_procesadas': hojas_procesadas
        })
    
    except pd.errors.EmptyDataError:
        return jsonify({'success': False, 'message': 'Error: Uno o más archivos están vacíos o no tienen datos válidos.'}), 400
    except pd.errors.ParserError:
        return jsonify({'success': False, 'message': 'Error: No se pudo leer el archivo Excel. Verifica el formato.'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': f'Ocurrió un error procesando los archivos: {e}'}), 500

def convertir_sheet_name_a_fecha(sheet_name: str) -> str:
    """Convierte el nombre de la hoja (formato dd-mm-yyyy) a fecha dd/mm/yyyy."""
    try:
        if not sheet_name:
            return ''
        
        # Si ya está en formato dd/mm/yyyy, devolverlo tal como está
        if '/' in sheet_name and len(sheet_name.split('/')) == 3:
            return sheet_name
        
        # Si está en formato dd-mm-yyyy, convertir a dd/mm/yyyy
        if '-' in sheet_name and len(sheet_name.split('-')) == 3:
            partes = sheet_name.split('-')
            if len(partes) == 3:
                return f"{partes[0]}/{partes[1]}/{partes[2]}"
        
        # Intentar parsear como fecha y reformatear
        try:
            fecha_obj = datetime.strptime(sheet_name, '%d-%m-%Y')
            return fecha_obj.strftime('%d/%m/%Y')
        except ValueError:
            try:
                fecha_obj = datetime.strptime(sheet_name, '%Y-%m-%d')
                return fecha_obj.strftime('%d/%m/%Y')
            except ValueError:
                pass
        
        # Si no se puede convertir, devolver el sheet_name original
        return sheet_name
    except Exception:
        return sheet_name

def generate_report_rows(
    base_asesores: List[Dict[str, Any]], 
    admin_map: Dict[str, Dict[str, Any]], 
    llamadas_map: Dict[str, Dict[str, Any]], 
    voip_map: Dict[str, Dict[str, Any]], 
    sheet_name: str
) -> List[Dict[str, Any]]:
    """Genera las filas del reporte para una hoja específica."""
    report_rows = []
    debug_count = 0
    
    for asesor in base_asesores:
        asesor_id = str(asesor.get('ID', ''))
        asesor_ext = str(asesor.get('EXT', ''))
        asesor_voip = str(asesor.get('VOIP', ''))
        
        admin_row = admin_map.get(asesor_id, {})
        llamadas_row = llamadas_map.get(asesor_ext, {})
        
        # Mapeo VOIP por número VOIP del asesor
        asesor_nombre = str(asesor.get('Nombre', '')).title()
        voip_row = voip_map.get(asesor_voip, {})
        
        # Debug limitado
        if debug_count < 5:
            if voip_row:
                print(f"   🎯 VOIP encontrado para {asesor_nombre} (VOIP: {asesor_voip}): {safe_get_value(voip_row, VOIP_COLUMNS['TOTAL_VOIP'], 0)} llamadas")
            elif asesor_nombre and len(voip_map) > 0:
                print(f"   ❌ VOIP no encontrado para {asesor_nombre} (VOIP: {asesor_voip})")
                voip_exists = asesor_voip in voip_map if asesor_voip else False
                print(f"   🔍 ¿VOIP '{asesor_voip}' está en mapa? {voip_exists}")
                if voip_exists:
                    print(f"   ⚠️  ERROR: El número VOIP SÍ existe pero no se está mapeando correctamente")
                print(f"   📋 Total claves en voip_map: {len(voip_map)}")
            debug_count += 1
        
        # Calcular llamadas
        llamadas_microsip = safe_float_conversion(safe_get_value(llamadas_row, LLAMADAS_COLUMNS['TOTAL_LLAMADAS'], 0))
        llamadas_voip = safe_float_conversion(safe_get_value(voip_row, VOIP_COLUMNS['TOTAL_VOIP'], 0))
        total_llamadas = llamadas_microsip + llamadas_voip
        
        # Convertir sheet_name a fecha del reporte (dd/mm/yyyy)
        fecha_reporte = convertir_sheet_name_a_fecha(sheet_name)
        
        # Crear fila del reporte
        new_row = {
            'Fecha Ingreso': limpiar_fecha(asesor.get('Fecha Ingreso', '')), 
            'Fecha': fecha_reporte,  # Usar la fecha del reporte (sheet_name)
            'Cedula': asesor.get('Cedula', ''),
            'ID': asesor.get('ID', ''),
            'EXT': asesor.get('EXT', ''),
            'VOIP': asesor.get('VOIP', ''),
            'Nombre': str(asesor.get('Nombre', '')).title(),
            'Sede': asesor.get('Sede', ''),
            'Ubicacion': asesor.get('Ubicación', ''),
            'Logueo': convertir_hora_formato(safe_get_value(admin_row, ADMIN_COLUMNS['LOGUEO'], '')),
            'Mora': normalizar_valor_mora(safe_get_value(admin_row, ADMIN_COLUMNS['CARTERA'], '')),
            'Asignacion': safe_get_value(admin_row, ADMIN_COLUMNS['ASIGNACION'], ''),
            'Clientes gestionados 11 am': safe_get_value(admin_row, ADMIN_COLUMNS['TOCADAS_11AM'], ''),
            'Capital Asignado': formatear_valor_monetario(safe_get_value(admin_row, ADMIN_COLUMNS['ASIGNADO'], 0)),
            'Capital Recuperado': formatear_valor_monetario(safe_get_value(admin_row, ADMIN_COLUMNS['RECUPERADO'], 0)),
            'PAGOS': safe_get_value(admin_row, ADMIN_COLUMNS['PAGOS'], ''),
            '% Recuperado': formatear_porcentaje_con_coma(convertir_porcentaje_a_decimal(safe_get_value(admin_row, ADMIN_COLUMNS['PORCENTAJE_RECUPERADO'], 0))),
            '% Cuentas': formatear_porcentaje_con_coma(convertir_porcentaje_a_decimal(safe_get_value(admin_row, ADMIN_COLUMNS['PORCENTAJE_CUENTAS'], 0))),
            'Total toques': safe_get_value(admin_row, ADMIN_COLUMNS['TOQUES'], ''),
            'Ultimo Toque': convertir_hora_formato(safe_get_value(admin_row, ADMIN_COLUMNS['ULTIMO_TOQUE'], '')),
            'Llamadas Microsip': int(llamadas_microsip),
            'Llamadas VOIP': int(llamadas_voip),
            'Total Llamadas': int(total_llamadas),
            'Gerencia': safe_get_value(admin_row, ADMIN_COLUMNS['GERENTE'], ''),
            'Team': safe_get_value(admin_row, ADMIN_COLUMNS['TEAM_LEADER'], '')
        }
        report_rows.append(new_row)
    
    return report_rows

def create_default_sheet(writer: Any, available_sheets: Dict[str, Any]) -> None:
    """Crea hoja por defecto cuando no se procesan datos."""
    try:
        default_data = pd.DataFrame({
            'Mensaje': ['No se encontraron hojas coincidentes para procesar'],
            'Hojas disponibles': [', '.join(available_sheets.keys())],
            'Fecha procesamiento': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        })
        default_data.to_excel(writer, sheet_name=DEFAULT_SHEET_NAME, index=False)
        apply_excel_formatting(writer.sheets[DEFAULT_SHEET_NAME], default_data, DEFAULT_SHEET_NAME)
        print(f"✅ Hoja por defecto creada: {DEFAULT_SHEET_NAME}")
    except Exception as e:
        print(f"❌ Error creando hoja por defecto: {e}")

# ============================================================================
# ENDPOINTS ADICIONALES
# ============================================================================
def descargar_reporte3():
    """Descarga el archivo temporal del Reporte 3."""
    try:
        data = request.get_json()
        if not data or 'temp_file' not in data:
            return "Error: Nombre de archivo no proporcionado.", 400
        
        filename = data['temp_file']
        temp_dir = 'temp_files'
        temp_file_path = os.path.join(temp_dir, filename)
        
        if not os.path.exists(temp_file_path):
            return "Error: Archivo no encontrado.", 404
        
        return send_file(
            temp_file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return f"Error descargando archivo: {e}", 500

def continuar_a_paso4():
    """Continúa al paso 4 con el archivo del Reporte 3."""
    try:
        data = request.get_json()
        if not data or 'temp_file' not in data:
            return jsonify({'success': False, 'message': 'Nombre de archivo no proporcionado.'}), 400
        
        filename = data['temp_file']
        temp_dir = 'temp_files'
        temp_file_path = os.path.join(temp_dir, filename)
        
        if not os.path.exists(temp_file_path):
            return jsonify({'success': False, 'message': 'Archivo no encontrado.'}), 404
        
        return jsonify({
            'success': True,
            'message': 'Continuando al paso 4',
            'filename': filename,
            'redirect_url': '/paso4'
        })
    except Exception as e:
        return f"Error: {e}", 500

def actualizar_base_asesores():
    """Actualiza la base de asesores desde archivo Excel."""
    try:
        archivo = request.files.get('archivoAsesores')
        if not archivo:
            return "Error: No se proporcionó archivo.", 400
        
        if not allowed_file(archivo.filename, {'xlsx'}):
            return "Error: Solo se permiten archivos .xlsx", 400
        
        df = pd.read_excel(archivo, dtype=str)
        df = df.fillna('')
        
        missing_columns = [col for col in REQUIRED_ASESOR_COLUMNS if col not in df.columns]
        if missing_columns:
            return f"Error: Faltan columnas requeridas: {missing_columns}", 400
        
        asesores_data = []
        for _, row in df.iterrows():
            asesor = {}
            for campo in REQUIRED_ASESOR_COLUMNS:
                valor = row.get(campo, '')
                if campo in ['Fecha Ingreso', 'Fecha'] and valor:
                    valor = limpiar_fecha(str(valor))
                elif campo == 'Nombre' and valor:
                    valor = str(valor).title()
                asesor[campo] = valor
            asesores_data.append(asesor)
        
        if asesores_data:
            backup_name = f"{BASE_ASESORES_FILE}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            if os.path.exists(BASE_ASESORES_FILE):
                shutil.copy2(BASE_ASESORES_FILE, backup_name)
            
            with open(BASE_ASESORES_FILE, 'w', encoding='utf-8') as f:
                json.dump(asesores_data, f, indent=2, ensure_ascii=False)
            
            return jsonify({
                'status': 'success',
                'message': f'Base de asesores actualizada: {len(asesores_data)} registros',
                'backup_created': backup_name
            })
        else:
            return "Error: No se encontraron datos válidos en el archivo.", 400
    
    except Exception as e:
        return f"Error procesando archivo: {e}", 500